from __future__ import annotations

import copy
import csv
import json
import re
import sys
import tempfile
import unittest
from pathlib import Path


SKILL_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SKILL_ROOT / "scripts"))

import compile_fill  # noqa: E402
import execute_batch  # noqa: E402
import mod_nominate  # noqa: E402
import promote_output  # noqa: E402
from _mod_catalog import parse_mod_index  # noqa: E402
from _probe_fixtures import (  # noqa: E402
    BASE_SPEC,
    make_all_missing_lookup_workdir,
    make_empty_lookup_workdir,
    make_egypt_workdir,
    make_header_row_workdir,
    make_multiproduct_block_workdir,
    make_preformatted_quotation_workdir,
    make_probe_inplace_workdir as make_inplace_workdir,
    make_probe_workdir as make_workdir,
    make_single_block_workdir,
    make_styled_anchor_workdir,
)


# ── Shared fixtures ────────────────────────────────────────────────────

def spec_with(wd: dict, **mutations) -> dict:
    import copy
    spec = copy.deepcopy(BASE_SPEC)
    spec["fingerprints"] = {
        "source_structure": wd["manifest"]["fingerprints"]["source_structure"],
        "target_structure": wd["manifest"]["fingerprints"]["target_structure"],
    }
    for path, value in mutations.items():
        node = spec
        parts = path.split(".")
        for p in parts[:-1]:
            node = node[p] if isinstance(node, dict) else node[int(p)]
        if isinstance(node, list):
            node[int(parts[-1])] = value
        else:
            node[parts[-1]] = value
    return spec


def compile_spec_with(wd: dict, spec: dict) -> dict:
    return compile_fill.compile_spec(spec, wd["manifest"], wd["workdir"])


def compile_fail_codes(wd: dict, spec: dict) -> list[str]:
    """Compile; return emitted defect codes (exit 3) or [] on success."""
    from io import StringIO
    buf = StringIO()
    old = sys.stderr
    sys.stderr = buf
    try:
        compile_fill.compile_spec(spec, wd["manifest"], wd["workdir"])
        return []
    except SystemExit as e:
        assert e.code == 3
        return re.findall(r'"code": "([A-Z_]+)"', buf.getvalue())
    finally:
        sys.stderr = old


def lookup_column_spec(wd: dict, field: str = "compressor",
                       missing: str = "empty", key_column: str = "C",
                       target: str = "F") -> dict:
    """Base spec + one lookup-only column mapping (Q6/Q15 tests)."""
    spec = spec_with(wd)
    spec["mapping"]["targets"][0]["columns"].append(
        {"target": target, "lookup": {"name": "fields",
                                      "field": field, "missing": missing}})
    spec["mapping"]["targets"][0]["lookups"] = [
        {"name": "fields", "from": "inheritance.json", "key_column": key_column,
         "fields": [field], "missing": missing}]
    return spec


INPLACE_BASE_SPEC = {
    "task": {"intent": "t", "selected_mod": "NONE", "selected_mod_revision": None},
    "inputs": {"sources": ["source_maoli.xlsx"], "target": "target.xlsx",
               "source_sheets": [{"source": "source_maoli.xlsx", "sheets": ["毛利表"]}],
               "target_sheet": "S"},
    "fingerprints": {"source_structure": None, "target_structure": None},
    "mapping": {"targets": [{
        "sheet": "S", "base_last_row": 14,
        "clone_roles": [
            {"role": "data", "mode": "inplace", "start_row": 7, "capacity": 4,
             "template_row": 8},
        ],
        "rows": {"source": "source_maoli"},
        "columns": [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E"},
        ],
        "group_merges": [{"col": "A", "group_by": "A", "style": "label"}],
        "nulls": [{"col": "D", "rows": "all"}, {"col": "F", "rows": "all"}],
    }]},
    "decisions": [], "gaps": [],
    "lineage": [{"source": "source_maoli_flat.csv", "role": "primary", "note": ""}],
    "validation": {"required_coverage": [], "required_empty": [], "key_outputs": ["A7"]},
}


class CompileTests(unittest.TestCase):
    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.workdir = make_workdir(self.tmp)
        self.workdir["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    # ── FillSpec schema validation ──
    def test_schema_requires_top_keys(self):
        spec = spec_with(self.workdir)
        del spec["lineage"]
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Fingerprint match ──
    def test_stale_fingerprint_rejected(self):
        spec = spec_with(self.workdir)
        spec["fingerprints"]["target_structure"] = "deadbeef"
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Input hash binding ──
    def test_plan_binds_staged_input_hashes(self):
        """plan.input_hashes = compile-time sha256 of the staged sources+target
        (recomputed, not copied from the manifest's outline-stage snapshot)."""
        plan = compile_spec_with(self.workdir, spec_with(self.workdir))
        self.assertEqual(set(plan["input_hashes"]),
                         {"source_maoli.xlsx", "target.xlsx"})
        # probe workdir has no real staged xlsx → binding is None (unverifiable;
        # execute fails closed on a None binding)
        self.assertIsNone(plan["input_hashes"]["source_maoli.xlsx"])
        # with real staged files → sha256 bound
        (self.tmp / "source_maoli.xlsx").write_bytes(b"SRC-BYTES")
        (self.tmp / "target.xlsx").write_bytes(b"TARGET-BYTES")
        plan2 = compile_spec_with(self.workdir, spec_with(self.workdir))
        self.assertEqual(plan2["input_hashes"]["source_maoli.xlsx"],
                         promote_output.sha256_file(self.tmp / "source_maoli.xlsx"))
        self.assertEqual(plan2["input_hashes"]["target.xlsx"],
                         promote_output.sha256_file(self.tmp / "target.xlsx"))

    # ── Selectors ──
    def test_selectors_filter_rows(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["rows"]["selectors"] = [
            {"column": "A", "pattern": "家用*"}]
        spec["validation"]["key_outputs"] = ["A7"]
        plan = compile_spec_with(self.workdir, spec)
        self.assertEqual(plan["source_coverage"][0]["matched"], 2)
        self.assertEqual([[s, o, t] for s, o, t in plan["row_map"]],
                         [["source_maoli_flat.csv", 101, 7], ["source_maoli_flat.csv", 102, 8]])

    def test_selector_no_match_fails(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["rows"]["selectors"] = [
            {"column": "A", "pattern": "NOPE*"}]
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Column mapping / transforms ──
    def test_direct_and_constant_mapping(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "E", "value": "0"})
        plan = compile_spec_with(self.workdir, spec)
        writes = {w["col"] for w in plan["writes"]}
        self.assertIn("E", writes)
        self.assertEqual([w["value"] for w in plan["writes"] if w["col"] == "E"][0], "0")

    def test_lookup_missing_key_error(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "C", "target": "C"},
            {"target": "F", "lookup": {"name": "fields",
                                       "field": "compressor", "missing": "error"}},
        ]
        spec["mapping"]["targets"][0]["lookups"] = [
            {"name": "fields", "from": "inheritance.json", "key_column": "C",
             "fields": ["compressor"], "missing": "error"}]
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)  # Z003 missing from index

    def test_lookup_missing_empty_ok(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "F", "lookup": {"name": "fields",
                                       "field": "compressor", "missing": "empty"}})
        spec["mapping"]["targets"][0]["lookups"] = [
            {"name": "fields", "from": "inheritance.json", "key_column": "C",
             "fields": ["compressor"], "missing": "empty"}]
        plan = compile_spec_with(self.workdir, spec)
        f_vals = [w["value"] for w in plan["writes"] if w["col"] == "F"]
        self.assertEqual(f_vals, ["C-1", "C-2", ""])  # Z003 → empty

    # ── Clone anchor rejection ──
    def test_clone_source_is_anchor_rejected(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["clone_roles"][2]["template_row"] = 5  # A5:A6 anchor
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Null residue detection ──
    def test_clone_residue_unhandled(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["nulls"] = []  # template row 3 carries D=F-1
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Duplicate target write ──
    def test_duplicate_target_write_rejected(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"A": "1+{r}"}}  # collides with the A column mapping
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Formula range validation ──
    def test_formula_template_bad_key_rejected(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"E": "{x}+1"}}
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    def test_aggregate_range_out_of_block_rejected(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "G", "rows": "1:9",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Key outputs must be written ──
    def test_key_output_unwritten_rejected(self):
        spec = spec_with(self.workdir)
        spec["validation"]["key_outputs"] = ["Z5"]
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Required coverage ──
    def test_required_coverage_unmatched_rejected(self):
        spec = spec_with(self.workdir)
        spec["validation"]["required_coverage"] = [
            {"source": "source_maoli_flat.csv", "rows": [101, 103, 999]}]
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Base row / columns bounds ──
    def test_base_row_out_of_bounds_rejected(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["base_last_row"] = 999
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    def test_column_out_of_digest_rejected(self):
        spec = spec_with(self.workdir)
        spec["mapping"]["targets"][0]["nulls"] = [{"col": "ZZ", "rows": "all"}]
        with self.assertRaises(SystemExit) as ctx:
            compile_fill.compile_spec(spec, self.workdir["manifest"], self.tmp)
        self.assertEqual(ctx.exception.code, 3)

    # ── Full compile happy path ──
    def test_full_compile_success(self):
        spec = spec_with(self.workdir)
        plan = compile_spec_with(self.workdir, spec)
        self.assertEqual(plan["source_coverage"][0]["matched"], 3)
        self.assertEqual(plan["blocks"][-1]["data_end"], 9)
        self.assertGreater(len(plan["operations"]), 0)
        empties = [rb for rb in plan["readback"] if rb["kind"] == "empty"]
        self.assertEqual(len(empties), 3)  # D5:D7 nulls
        self.assertEqual(plan["key_outputs"], [{"path": "/S/A7", "kind": "value"}])
        self.assertEqual(plan["fill_spec_sha256"], None)  # set by main from the file


class ModNominateTests(unittest.TestCase):
    def test_four_states(self):
        idx_dir = Path(__file__).parent / "_fixtures"
        idx_dir.mkdir(exist_ok=True)
        index = idx_dir / "MOD_INDEX_test.md"
        mods = idx_dir / "MODS_test"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| test_mod | tm | semantic_type::quotation,dimension_set::product_sku |  | MOD_test.md | 1 | private |\n",
            encoding="utf-8")
        (mods / "MOD_test.md").write_text(
            "## Applicability\n- semantic_type: quotation\n\n## 业务逻辑摘要\n- x\n",
            encoding="utf-8")

        # none: no signal
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index), mods, "fill the pipeline sheet", [], [])
        self.assertEqual(r["status"], "none")

        # resolved: single candidate, all signals verified (digest header
        # role facts resolve dimension_set as a verified hit)
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index), mods, "报价汇总 迁移 毛利表",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本"], [])
        self.assertEqual(r["status"], "resolved")

        # ambiguous: digest present but no SKU role fact → dimension_set
        # verifiably missed (fail-closed, no fake hit from digest text)
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index), mods, "报价汇总 迁移 毛利表",
            ["- 表头: 数量 | 报价 | 毛利"], [])
        self.assertEqual(r["status"], "ambiguous")
        self.assertEqual(r["candidates"][0]["missed"],
                         ["dimension_set::product_sku"])

        # ambiguous: structural signal unverified (no digests yet)
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index), mods, "报价汇总 迁移", [], [])
        self.assertEqual(r["status"], "ambiguous")

        # conflict: exclusion fired (24-col fingerprint missing)
        index2 = idx_dir / "MOD_INDEX_test2.md"
        index2.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| t24 | t | semantic_type::quotation | 目标缺少24角色表头指纹 | MOD_test.md | 1 | private |\n",
            encoding="utf-8")
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index2), mods, "报价汇总 迁移 毛利表",
            ["目标 sheet 21行 × 24列"], [])  # 24-col present → no fire
        self.assertEqual(r["status"], "resolved")
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index2), mods, "报价汇总 迁移 毛利表",
            ["目标 sheet 21行 × 3列"], [])  # not 24 → exclusion fires
        self.assertEqual(r["status"], "conflict")
        # evidence-missing conflict carries the re-run hint, not a user demand
        self.assertIn("证据缺失", r["why"])
        r = mod_nominate.resolve(
            mod_nominate.parse_index(index2), mods, "报价汇总 迁移 毛利表",
            [], [])  # no evidence at all → 证据缺失 (not structural mismatch)
        self.assertEqual(r["status"], "conflict")
        self.assertIn("证据缺失", r["why"])

    def test_exclusion_role_fingerprint(self):
        """v2.5: 排除信号带角色参数 — 表头角色多数命中 = 同角色变体放行;
        角色缺失 = 结构不符触发排除; outline JSON 通道可独立提供 24 列证据."""
        idx_dir = Path(__file__).parent / "_fixtures"
        index = idx_dir / "MOD_INDEX_roles.md"
        mods = idx_dir / "MODS_roles"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| t24r | t | semantic_type::quotation | "
            "目标缺少24角色表头指纹(数量,报价,原型机成本,结算价,毛利,系列盈亏) "
            "| MOD_test.md | 1 | private |\n",
            encoding="utf-8")
        (mods / "MOD_test.md").write_text(
            "## Applicability\n- semantic_type: quotation\n\n"
            "## 业务逻辑摘要\n- x\n", encoding="utf-8")
        entries = mod_nominate.parse_index(index)

        # 同角色变体: 表头含全部角色但列数 ≠ 24 → 放行 (TGT-002)
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移 毛利表",
            ["- 表头: 数量 | 报价 | 原型机成本 | 结算价 | 毛利 | 系列盈亏"],
            [])
        self.assertEqual(r["status"], "resolved")

        # 角色缺失: 表头仅含 1/6 角色 (型号清单类表) → 结构不符, 排除触发
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移 毛利表",
            ["- 表头: 型号 | 数量 | 备注 | 日期"], [])
        self.assertEqual(r["status"], "conflict")
        self.assertIn("角色", r["why"])

        # outline JSON 通道: 无 digest 但 outline 有 24 列 sheet → 放行
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移 毛利表", [],
            [{"data": {"sheets": [{"name": "11_FRESH本土", "rows": 21, "cols": 24}]}}])
        self.assertEqual(r["status"], "resolved")

    def test_sheet_marker_signal(self):
        """sheet_marker: outline 含业务标记 sheet (三三三/333) → 命中;
        无标记 → 未命中; outline 未喂 → 待验证."""
        idx_dir = Path(__file__).parent / "_fixtures"
        index = idx_dir / "MOD_INDEX_marker.md"
        mods = idx_dir / "MODS_marker"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| tmark | t | semantic_type::quotation,sheet_marker::三三三\\|333 |  | MOD_test.md | 1 | private |\n",
            encoding="utf-8")
        (mods / "MOD_test.md").write_text(
            "## Applicability\n- semantic_type: quotation\n\n"
            "## 业务逻辑摘要\n- x\n", encoding="utf-8")
        entries = mod_nominate.parse_index(index)
        self.assertEqual(entries[0]["scope"], "semantic_type::quotation,"
                         "sheet_marker::三三三|333")  # \| 反转义

        # outline 含 三三三 sheet → marker 命中 → resolved
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["digest 21行 × 24列"],
            [{"data": {"sheets": [{"name": "三三三", "rows": 29, "cols": 3},
                                  {"name": "11_FRESH本土", "rows": 21, "cols": 24}]}}])
        self.assertEqual(r["status"], "resolved")
        self.assertIn("sheet_marker::三三三|333",
                      r["candidates"][0]["hits"])

        # outline 无标记 sheet → marker missed → fail-closed: 不再自动
        # resolved, 询问用户 (ambiguous)
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["digest 21行 × 24列"],
            [{"data": {"sheets": [{"name": "11_FRESH本土", "rows": 21, "cols": 24}]}}])
        self.assertEqual(r["status"], "ambiguous")
        self.assertEqual(r["candidates"][0]["missed"],
                         ["sheet_marker::三三三|333"])

        # outline 未喂 → marker pending (不阻断提名)
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移", [], [])
        self.assertEqual(r["status"], "ambiguous")
        self.assertIn("sheet_marker::三三三|333",
                      r["candidates"][0]["pending"])

    def _write_simple_mod(self, idx_dir: Path, mods: Path, scope: str,
                          exclusion: str = "") -> list[dict]:
        index = idx_dir / "MOD_INDEX_auto.md"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            f"| tauto | t | {scope} | {exclusion} | MOD_test.md | 1 | private |\n",
            encoding="utf-8")
        (mods / "MOD_test.md").write_text(
            "## Applicability\n- semantic_type: quotation\n\n"
            "## 业务逻辑摘要\n- x\n", encoding="utf-8")
        return mod_nominate.parse_index(index)

    def test_missed_signal_blocks_auto_resolution(self):
        """最小 fixture (issue 02): 单候选 + 存在可验证未命中信号 →
        不自动 resolved (ambiguous, 询问用户)."""
        idx_dir = Path(__file__).parent / "_fixtures"
        mods = idx_dir / "MODS_auto"
        entries = self._write_simple_mod(
            idx_dir, mods,
            "semantic_type::quotation,sheet_marker::三三三\\|333")
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本"],
            [{"data": {"sheets": [{"name": "11_FRESH本土", "rows": 21, "cols": 24}]}}])
        self.assertEqual(r["status"], "ambiguous")
        self.assertEqual(r["candidates"][0]["missed"],
                         ["sheet_marker::三三三|333"])
        self.assertIn("missed", r["why"])

    def test_missed_only_signal_no_candidate(self):
        """仅含 missed 信号 (无任何 hit) 的 MOD 不成候选 → status none —
        fail-closed by absence: 不自动 resolved, 也不冒充命中."""
        idx_dir = Path(__file__).parent / "_fixtures"
        mods = idx_dir / "MODS_auto"
        entries = self._write_simple_mod(
            idx_dir, mods, "sheet_marker::三三三\\|333")
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本"],
            [{"data": {"sheets": [{"name": "11_FRESH本土", "rows": 21, "cols": 24}]}}])
        self.assertEqual(r["status"], "none")
        self.assertEqual(r["candidates"], [])

    def test_unknown_exclusion_fail_closed(self):
        """未知排除条件无 evaluator → fail-closed: 不再默认放行, 记入
        pending_exclusions → ambiguous (询问用户)."""
        idx_dir = Path(__file__).parent / "_fixtures"
        mods = idx_dir / "MODS_auto"
        entries = self._write_simple_mod(
            idx_dir, mods, "semantic_type::quotation",
            exclusion="目标缺少未知指纹")
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本"], [])
        self.assertEqual(r["status"], "ambiguous")
        self.assertEqual(r["candidates"][0]["pending_exclusions"],
                         ["目标缺少未知指纹"])

    def test_dimension_set_verified_from_digest_roles(self):
        """dimension_set 不再「有 digest 文本即命中」(issue 02): 表头含 SKU
        角色 → hit; 表头缺 SKU 角色 → miss; 无 digest → pending."""
        idx_dir = Path(__file__).parent / "_fixtures"
        mods = idx_dir / "MODS_auto"
        entries = self._write_simple_mod(
            idx_dir, mods,
            "semantic_type::quotation,dimension_set::product_sku")
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本"], [])
        self.assertEqual(r["status"], "resolved")
        self.assertIn("dimension_set::product_sku",
                      r["candidates"][0]["hits"])
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: 数量 | 报价 | 毛利"], [])
        self.assertEqual(r["status"], "ambiguous")
        self.assertEqual(r["candidates"][0]["missed"],
                         ["dimension_set::product_sku"])
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移", [], [])
        self.assertEqual(r["status"], "ambiguous")
        self.assertIn("dimension_set::product_sku",
                      r["candidates"][0]["pending"])

    def test_cost_reply_exclusion_evaluator(self):
        """cost_reply 排除「目标缺少客户Sheet重复批次块或Z码和原型机成本角色」
        有 evaluator: 重复块+角色齐 → 放行; 缺角色/缺块 → 结构不符触发排除;
        无 digest → 证据缺失 (conflict, 补证据重跑)."""
        idx_dir = Path(__file__).parent / "_fixtures"
        mods = idx_dir / "MODS_auto"
        entries = self._write_simple_mod(
            idx_dir, mods, "semantic_type::quotation",
            exclusion="目标缺少客户Sheet重复批次块或Z码和原型机成本角色")
        blocks = ("- 数据块:\n"
                  "  - B1 行1-21: \"历史批次1\" (score 60)\n"
                  "  - B2 行22-40: \"历史批次2\" (score 55)")
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本", blocks], [])
        self.assertEqual(r["status"], "resolved")
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: 日期 | 数量 | 备注", blocks], [])
        self.assertEqual(r["status"], "conflict")
        self.assertIn("角色", r["why"])
        r = mod_nominate.resolve(
            entries, mods, "报价汇总 迁移",
            ["- 表头: Z码 | 数量 | 报价 | 原型机成本",
             "- 数据块: 无自动候选 (LLM 依摘要与业务上下文判定)"], [])
        self.assertEqual(r["status"], "conflict")
        self.assertIn("批次块", r["why"])
        r = mod_nominate.resolve(entries, mods, "报价汇总 迁移", [], [])
        self.assertEqual(r["status"], "conflict")
        self.assertIn("证据缺失", r["why"])

    def test_rules_not_in_nomination_two_phase_loading(self):
        """两段加载 (SKILL.md 硬性契约): 提名输出不含完整规则集 (候选只带
        hits/pending/missed/摘要); 用户裁决后经 load_rules_for_selected_mod()
        从 MOD 文件全文加载完整规则, 注入 FillSpec 撰写上下文."""
        idx_dir = Path(__file__).parent / "_fixtures"
        index = idx_dir / "MOD_INDEX_rules.md"
        mods = idx_dir / "MODS_rules"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| trules | t | semantic_type::quotation |  | MOD_rules.md | 1 | private |\n",
            encoding="utf-8")
        (mods / "MOD_rules.md").write_text(
            "## Applicability\n- semantic_type: quotation\n\n"
            "## 业务逻辑摘要\n- 原型机成本 = 源面价 − 铜管成本\n\n"
            "| Rule ID | Group | Gate | Description | Applies to | Notes |\n"
            "|---|---|---|---|---|---|\n"
            "| FLD-006 | business_transformation | mod_gate | 目标原型机成本等于源面价(更新)减源铜管成本。 | 原型机成本、面价(更新)、铜管成本 | 与 FLD-007 共同保证结算价等于面价 |\n"
            "| FRM-002 | business_transformation | mod_gate | 结算价等于原型机成本加铜管成本。 | 结算价 | 与成本拆分规则逐行核对 |\n",
            encoding="utf-8")
        entries = mod_nominate.parse_index(index)
        r = mod_nominate.resolve(entries, mods, "报价汇总 迁移", ["digest"], [])
        self.assertEqual(r["status"], "resolved")
        cand = r["candidates"][0]
        # 提名输出: 摘要可供裁决, 完整规则集不随提名输出
        self.assertIn("summary", cand)
        self.assertNotIn("rules", cand)
        # 裁决后第二段: 从 MOD 文件全文加载完整规则
        full = mod_nominate.load_rules_for_selected_mod(mods, "MOD_rules.md")
        self.assertEqual([x["id"] for x in full], ["FLD-006", "FRM-002"])
        self.assertIn("源面价", full[0]["description"])
        self.assertEqual(full[0]["group"], "business_transformation")
        self.assertEqual(full[0]["notes"], "与 FLD-007 共同保证结算价等于面价")

    def test_ambiguous_candidates_carry_rule_evidence_summary(self):
        """多候选 ambiguous: 裁决选项附规则证据摘要 (id+description, 足够裁决
        判断), 仍不含完整规则集 — 完整规则在选定后加载."""
        idx_dir = Path(__file__).parent / "_fixtures"
        index = idx_dir / "MOD_INDEX_ambig.md"
        mods = idx_dir / "MODS_ambig"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| mod_a | a | semantic_type::quotation |  | MOD_a.md | 1 | private |\n"
            "| mod_b | b | semantic_type::quotation |  | MOD_b.md | 1 | private |\n",
            encoding="utf-8")
        for name, rule_id in (("MOD_a.md", "FLD-006"), ("MOD_b.md", "FRM-002")):
            (mods / name).write_text(
                "## Applicability\n- semantic_type: quotation\n\n"
                "## 业务逻辑摘要\n- x\n\n"
                "| Rule ID | Group | Gate | Description | Applies to | Notes |\n"
                "|---|---|---|---|---|---|\n"
                f"| {rule_id} | business_transformation | mod_gate | 描述{rule_id}。 | X | n |\n",
                encoding="utf-8")
        r = mod_nominate.resolve(mod_nominate.parse_index(index), mods,
                                 "报价汇总 迁移", ["digest"], [])
        self.assertEqual(r["status"], "ambiguous")
        self.assertEqual(len(r["candidates"]), 2)
        for cand in r["candidates"]:
            self.assertNotIn("rules", cand)  # 不随提名输出完整规则集
            self.assertIn("rule_evidence", cand)  # 附裁决用规则证据摘要
            expect = "FLD-006" if cand["name"] == "mod_a" else "FRM-002"
            self.assertEqual(cand["rule_evidence"][0]["id"], expect)
            self.assertIn("description", cand["rule_evidence"][0])

    def test_out_relative_resolves_to_workdir(self):
        """--out 相对路径以 workdir 为基准 — CWD != workdir 时结果不乱落
        (2026-08-12 round2: mod_resolution.json 曾写到 CWD 需手动归位)."""
        import subprocess
        with tempfile.TemporaryDirectory() as tmp, tempfile.TemporaryDirectory() as cwd:
            workdir = Path(tmp)
            cwd = Path(cwd)
            idx = workdir / "MOD_INDEX.md"
            mods = workdir / "MODS"
            mods.mkdir(exist_ok=True)
            idx.write_text(
                "## Registered MODs\n\n"
                "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
                "|---|---|---|---|---|---|---|\n"
                "| t_out | t | semantic_type::quotation |  | MOD_test.md | 1 | private |\n",
                encoding="utf-8")
            (mods / "MOD_test.md").write_text(
                "## Applicability\n- semantic_type: quotation\n\n## 业务逻辑摘要\n- x\n",
                encoding="utf-8")
            r = subprocess.run(
                [sys.executable, str(SKILL_ROOT / "scripts" / "mod_nominate.py"),
                 "--task", "报价汇总 迁移 毛利表", "--files", "a,b",
                 "--workdir", str(workdir), "--index", str(idx),
                 "--mods-dir", str(mods), "--out", "mod_resolution.json"],
                cwd=str(cwd), capture_output=True, text=True, encoding="utf-8")
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertTrue((workdir / "mod_resolution.json").is_file())
            self.assertFalse((cwd / "mod_resolution.json").exists())

    def test_explicit_alias_ignores_unrelated_mod_conflict(self):
        idx_dir = Path(__file__).parent / "_fixtures"
        index = idx_dir / "MOD_INDEX_explicit.md"
        mods = idx_dir / "MODS_explicit"
        mods.mkdir(exist_ok=True)
        index.write_text(
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals | Exclusion Signals | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| old_migration | old | semantic_type::quotation,target_title::报价汇总 | 目标缺少24角色表头指纹 | MOD_old.md | 3 | private |\n"
            "| cost_reply | tcl-email-quote-block | semantic_type::cost_reply_to_quotation_summary_block,source_pattern::*核价邮件*,target_pattern::*报价* |  | MOD_new.md | 1 | private |\n",
            encoding="utf-8")
        for name in ("MOD_old.md", "MOD_new.md"):
            (mods / name).write_text(
                "## Applicability\n- semantic_type: quotation\n\n"
                "## 业务逻辑摘要\n- x\n", encoding="utf-8")

        task = "使用 MOD tcl-email-quote-block，读取核价邮件并写入报价汇总"
        entries = mod_nominate.parse_index(index)
        explicit = mod_nominate.explicit_mod_mentions(entries, task)
        self.assertEqual(explicit, ["cost_reply"])
        r = mod_nominate.resolve(
            entries, mods, task, ["目标 sheet 8行 × 27列"], [],
            explicit_mod=explicit[0])
        self.assertEqual(r["status"], "resolved")
        self.assertEqual([c["name"] for c in r["candidates"]], ["cost_reply"])


class ReceiptHashTests(unittest.TestCase):
    def _promote(self, workdir, expect_code):
        with self.assertRaises(SystemExit) as ctx:
            sys.argv = ["promote_output", "--workdir", str(workdir),
                        "--final", str(workdir / "final.xlsx")]
            promote_output.main()
        self.assertEqual(ctx.exception.code, expect_code)

    def _gate_workdir(self, tmp) -> tuple[Path, dict]:
        """Workdir with real draft/spec/plan + staged inputs + receipt + gate."""
        import zipfile
        workdir = Path(tmp)
        draft = workdir / "validated_draft.xlsx"
        with zipfile.ZipFile(draft, "w") as z:
            z.writestr("xl/workbook.xml", "<workbook/>")  # valid zip for the promote check
        (workdir / "fill_spec.yaml").write_text("task: x", encoding="utf-8")
        (workdir / "source_maoli.xlsx").write_bytes(b"SOURCE-STAGED")
        (workdir / "template.xlsx").write_bytes(b"TEMPLATE-STAGED")
        input_hashes = {
            "source_maoli.xlsx": promote_output.sha256_file(workdir / "source_maoli.xlsx"),
            "template.xlsx": promote_output.sha256_file(workdir / "template.xlsx"),
        }
        (workdir / "execution_plan.json").write_text(json.dumps(
            {"target": "template.xlsx", "input_hashes": input_hashes},
            ensure_ascii=False), encoding="utf-8")
        hashes = {
            "fill_spec_sha256": promote_output.sha256_file(workdir / "fill_spec.yaml"),
            "execution_plan_sha256": promote_output.sha256_file(workdir / "execution_plan.json"),
            "draft_sha256": promote_output.sha256_file(draft),
        }
        receipt = {
            "draft_path": str(draft), **hashes,
            "source_hashes": {"source_maoli.xlsx": input_hashes["source_maoli.xlsx"]},
            "template_sha256": input_hashes["template.xlsx"],
        }
        (workdir / "draft_receipt.json").write_text(
            json.dumps(receipt, ensure_ascii=False), encoding="utf-8")
        return workdir, hashes

    def test_promotion_hash_mismatch_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            workdir, hashes = self._gate_workdir(tmp)
            (workdir / "draft_receipt.json").write_text(json.dumps(
                {"draft_path": str(workdir / "validated_draft.xlsx"),
                 "draft_sha256": "0" * 64, "execution_plan_sha256": "1" * 64,
                 "fill_spec_sha256": "2" * 64}, ensure_ascii=False), encoding="utf-8")
            # gate confirmed with the same (wrong) hashes → receipt still drifts
            (workdir / ".gate3_confirmed").write_text(json.dumps(
                {"hashes": {"fill_spec_sha256": "2" * 64,
                            "execution_plan_sha256": "1" * 64,
                            "draft_sha256": "0" * 64}}), encoding="utf-8")
            self._promote(workdir, 3)
            self.assertFalse((workdir / "final.xlsx").exists())

    def test_gate_confirm_requires_pending(self):
        import execution_gate
        with tempfile.TemporaryDirectory() as tmp:
            workdir = Path(tmp)
            with self.assertRaises(SystemExit) as ctx:
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            self.assertEqual(ctx.exception.code, 3)
            self.assertFalse((workdir / ".gate3_confirmed").exists())

    def test_gate_confirm_refuses_drifted_artifacts(self):
        import execution_gate
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _hashes = self._gate_workdir(tmp)
            with self.assertRaises(SystemExit) as ctx:
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            self.assertEqual(ctx.exception.code, 0)
            # draft changes after presentation → confirm must refuse
            (workdir / "validated_draft.xlsx").write_bytes(b"changed")
            with self.assertRaises(SystemExit) as ctx:
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            self.assertEqual(ctx.exception.code, 3)
            self.assertFalse((workdir / ".gate3_confirmed").exists())

    def test_gate_confirm_then_promote_ok(self):
        import execution_gate
        with tempfile.TemporaryDirectory() as tmp:
            workdir, hashes = self._gate_workdir(tmp)
            with self.assertRaises(SystemExit) as ctx:
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            self.assertEqual(ctx.exception.code, 0)
            with self.assertRaises(SystemExit) as ctx:
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            self.assertEqual(ctx.exception.code, 0)
            self.assertTrue((workdir / ".gate3_confirmed").exists())
            self.assertFalse((workdir / ".gate3_pending").exists())
            self._promote(workdir, 0)
            self.assertTrue((workdir / "final.xlsx").exists())
            self.assertEqual(hashes["draft_sha256"],
                             promote_output.sha256_file(workdir / "final.xlsx"))

    def test_promote_requires_positive_confirmation(self):
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            # pending marker present but NO confirmed record → reject
            (workdir / ".gate3_pending").write_text("pending", encoding="utf-8")
            self._promote(workdir, 3)
            # pending marker GONE but still no confirmed record → reject
            (workdir / ".gate3_pending").unlink()
            self._promote(workdir, 3)

    def test_promote_refuses_confirmed_hash_drift(self):
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            (workdir / ".gate3_confirmed").write_text(json.dumps(
                {"hashes": {"fill_spec_sha256": "a" * 64,
                            "execution_plan_sha256": "b" * 64,
                            "draft_sha256": "c" * 64}}), encoding="utf-8")
            self._promote(workdir, 3)

    def test_promote_preserves_existing_final_on_replace_failure(self):
        """The pre-existing final file must survive a failed atomic replace —
        promotion never deletes the old final before the new one is staged."""
        import execution_gate
        from unittest import mock
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            final = workdir / "final.xlsx"
            final.write_bytes(b"OLD-DELIVERED-FILE")
            with mock.patch("pathlib.Path.replace", side_effect=OSError("locked by Excel")):
                self._promote(workdir, 3)
            # old final intact, staging temp cleaned up
            self.assertEqual(final.read_bytes(), b"OLD-DELIVERED-FILE")
            self.assertFalse((workdir / "final.xlsx.promoting").exists())

    def test_promote_verifies_staged_copy_before_replace(self):
        """A corrupt staged copy must be rejected BEFORE the old final is touched."""
        import execution_gate
        from unittest import mock
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            final = workdir / "final.xlsx"
            final.write_bytes(b"OLD-DELIVERED-FILE")
            real_copy2 = promote_output.shutil.copy2

            def corrupt_copy(src, dst, **kw):
                real_copy2(src, dst, **kw)
                with open(dst, "ab") as f:  # corrupt the staged temp
                    f.write(b"EXTRA-CORRUPTION")

            with mock.patch("promote_output.shutil.copy2", side_effect=corrupt_copy):
                self._promote(workdir, 3)
            self.assertEqual(final.read_bytes(), b"OLD-DELIVERED-FILE")

    def test_promote_rejects_source_input_drift(self):
        """Staged SOURCE changed after the gate → HASH_DRIFT (exit 3), no final."""
        import execution_gate
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            (workdir / "source_maoli.xlsx").write_bytes(b"TAMPERED-SOURCE")
            self._promote(workdir, 3)
            self.assertFalse((workdir / "final.xlsx").exists())

    def test_promote_rejects_template_input_drift(self):
        """Staged TEMPLATE changed after the gate → HASH_DRIFT (exit 3)."""
        import execution_gate
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            (workdir / "template.xlsx").write_bytes(b"TAMPERED-TEMPLATE")
            self._promote(workdir, 3)
            self.assertFalse((workdir / "final.xlsx").exists())

    def test_promote_rejects_missing_input_evidence(self):
        """Receipt without execution-time input hashes → HASH_DRIFT (fail-closed)."""
        import execution_gate
        with tempfile.TemporaryDirectory() as tmp:
            workdir, _ = self._gate_workdir(tmp)
            receipt = json.loads(
                (workdir / "draft_receipt.json").read_text(encoding="utf-8"))
            del receipt["source_hashes"]
            del receipt["template_sha256"]
            (workdir / "draft_receipt.json").write_text(
                json.dumps(receipt, ensure_ascii=False), encoding="utf-8")
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--set", "--workdir", str(workdir)]
                execution_gate.main()
            with self.assertRaises(SystemExit):
                sys.argv = ["execution_gate", "--confirm", "--workdir", str(workdir)]
                execution_gate.main()
            self._promote(workdir, 3)
            self.assertFalse((workdir / "final.xlsx").exists())


class ExecuteInputHashContractTests(unittest.TestCase):
    """Issue 03 contract: execution-time recompute vs compile-time binding."""

    def _make_workdir(self, tmp) -> Path:
        workdir = Path(tmp)
        (workdir / "source_maoli.xlsx").write_bytes(b"SRC-BYTES")
        (workdir / "template.xlsx").write_bytes(b"TPL-BYTES")
        return workdir

    def _plan(self, workdir: Path) -> dict:
        return {
            "target": "template.xlsx",
            "input_hashes": {
                "source_maoli.xlsx":
                    promote_output.sha256_file(workdir / "source_maoli.xlsx"),
                "template.xlsx":
                    promote_output.sha256_file(workdir / "template.xlsx"),
            },
        }

    def test_no_drift_when_untampered(self):
        import execute_batch
        with tempfile.TemporaryDirectory() as tmp:
            workdir = self._make_workdir(tmp)
            plan = self._plan(workdir)
            actual, drifted = execute_batch.input_hash_drift(
                workdir, plan, workdir / "template.xlsx")
            self.assertEqual(drifted, [])
            self.assertEqual(actual["source_maoli.xlsx"],
                             promote_output.sha256_file(workdir / "source_maoli.xlsx"))
            self.assertEqual(actual["template.xlsx"],
                             promote_output.sha256_file(workdir / "template.xlsx"))

    def test_tampered_source_detected(self):
        import execute_batch
        with tempfile.TemporaryDirectory() as tmp:
            workdir = self._make_workdir(tmp)
            plan = self._plan(workdir)
            (workdir / "source_maoli.xlsx").write_bytes(b"TAMPERED")
            _actual, drifted = execute_batch.input_hash_drift(
                workdir, plan, workdir / "template.xlsx")
            self.assertEqual(drifted, ["source_maoli.xlsx"])

    def test_tampered_template_detected(self):
        import execute_batch
        with tempfile.TemporaryDirectory() as tmp:
            workdir = self._make_workdir(tmp)
            plan = self._plan(workdir)
            (workdir / "template.xlsx").write_bytes(b"TAMPERED")
            _actual, drifted = execute_batch.input_hash_drift(
                workdir, plan, workdir / "template.xlsx")
            self.assertEqual(drifted, ["template.xlsx"])

    def test_missing_file_is_drift(self):
        import execute_batch
        with tempfile.TemporaryDirectory() as tmp:
            workdir = self._make_workdir(tmp)
            plan = self._plan(workdir)
            (workdir / "source_maoli.xlsx").unlink()
            _actual, drifted = execute_batch.input_hash_drift(
                workdir, plan, workdir / "template.xlsx")
            self.assertEqual(drifted, ["source_maoli.xlsx"])

    def test_none_binding_is_drift(self):
        """Bound None (unverifiable at compile) must fail closed at execute."""
        import execute_batch
        with tempfile.TemporaryDirectory() as tmp:
            workdir = self._make_workdir(tmp)
            plan = self._plan(workdir)
            plan["input_hashes"]["source_maoli.xlsx"] = None
            _actual, drifted = execute_batch.input_hash_drift(
                workdir, plan, workdir / "template.xlsx")
            self.assertEqual(drifted, ["source_maoli.xlsx"])


class MultiTargetTests(unittest.TestCase):
    def test_multi_target_rejected(self):
        spec = {
            "task": {"intent": "t", "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_maoli.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_maoli.xlsx", "sheets": ["s"]}],
                       "target_sheet": "S"},
            "fingerprints": {"source_structure": "x", "target_structure": "y"},
            "mapping": {"targets": [
                {"sheet": "S1"}, {"sheet": "S2"},
            ]},
            "decisions": [], "gaps": [],
            "lineage": [{"source": "s.csv", "role": "primary", "note": ""}],
            "validation": {"required_coverage": [], "required_empty": [], "key_outputs": []},
        }
        defects = compile_fill.validate_schema(spec, {"files": [], "flattened": []})
        codes = {d["code"] for d in defects}
        self.assertIn("SPEC_TARGETS_TOO_MANY", codes)


class PartialNullsTests(unittest.TestCase):
    def test_partial_nulls_leaves_residue_detected(self):
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)
            # D carried by template row 3; nulls only row 1 → residue on rows 2-3
            spec["mapping"]["targets"][0]["nulls"] = [{"col": "D", "rows": [1]}]
            with self.assertRaises(SystemExit) as ctx:
                compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            self.assertEqual(ctx.exception.code, 3)
            stderr = sys.stderr
            # CLONE_RESIDUE_PARTIAL_NULLS must be among the defects — capture via re-run
            from io import StringIO
            buf = StringIO()
            old = sys.stderr
            sys.stderr = buf
            try:
                with self.assertRaises(SystemExit):
                    compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            finally:
                sys.stderr = old
            self.assertIn("CLONE_RESIDUE_PARTIAL_NULLS", buf.getvalue())

    def test_full_nulls_pass(self):
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)  # base spec already nulls D rows all
            plan = compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            self.assertGreater(len(plan["operations"]), 0)


class LookupNormalizeTests(unittest.TestCase):
    def test_inheritance_index_normalized(self):
        data = {
            "schema": "table-fill-inheritance-index-v1",
            "index": {
                "Z001": {"field_consensus": {
                    "compressor": {"status": "unique", "value": "C-1"},
                    "copper": {"status": "conflict", "value": "X"},
                }},
            },
        }
        out = compile_fill.normalize_lookup_data(data, "x")
        self.assertEqual(out, {"Z001": {"compressor": "C-1"}})

    # ── Q15: lookup 索引完整性 (埃及 FRESH 坑 1) ──
    def test_lookup_table_empty_rejected(self):
        """索引归一化后为空 (field_consensus 被清洗脚本重写丢失) →
        LOOKUP_TABLE_EMPTY (exit 3), 不再静默全空留白."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_empty_lookup_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            codes = compile_fail_codes(wd, lookup_column_spec(wd))
            self.assertIn("LOOKUP_TABLE_EMPTY", codes)

    def test_lookup_table_plain_empty_rejected(self):
        """扁平索引文件本身为空 ({}) → 同样 LOOKUP_TABLE_EMPTY (0 entries)."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            (Path(tmp) / "inheritance.json").write_text("{}", encoding="utf-8")
            self.assertIn("LOOKUP_TABLE_EMPTY",
                          compile_fail_codes(wd, lookup_column_spec(wd)))

    def test_lookup_column_all_missing_warns(self):
        """索引非空但声明 lookup 列全部未命中 → LOOKUP_COLUMN_ALL_MISSING 警告
        (编译继续, Gate 呈现) — 拦截整列静默空, 允许合法缺失."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_all_missing_lookup_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            plan = compile_fill.compile_spec(
                lookup_column_spec(wd), wd["manifest"], Path(tmp))
            codes = [w["code"] for w in plan["warnings"]]
            self.assertIn("LOOKUP_COLUMN_ALL_MISSING", codes)
            self.assertEqual(
                [w["value"] for w in plan["writes"] if w["col"] == "F"],
                ["", "", ""])  # 全列留空但不再静默
            self.assertGreater(len(plan["operations"]), 0)  # 警告不阻断编译

    def test_lookup_all_missing_corrective_action_self_reference(self):
        """整列未命中 corrective_action 提示索引输入自引用排查 (埃及 FRESH 坑 2:
        目标 sheet 误作索引输入 → 同 SKU 多值 → 共识 conflict → 静默缺失)."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_all_missing_lookup_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            plan = compile_fill.compile_spec(
                lookup_column_spec(wd), wd["manifest"], Path(tmp))
            w = next(w for w in plan["warnings"]
                     if w["code"] == "LOOKUP_COLUMN_ALL_MISSING")
            self.assertIn("target sheet", w["corrective_action"])
            self.assertIn("index input", w["corrective_action"])

    def test_lookup_column_partial_miss_no_warning(self):
        """部分命中 (部分行缺失是合法 gaps) → 不触发全列警告."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            (Path(tmp) / "inheritance.json").write_text(
                json.dumps({"Z001": {"compressor": "C-1"}}), encoding="utf-8")
            plan = compile_fill.compile_spec(
                lookup_column_spec(wd), wd["manifest"], Path(tmp))
            self.assertNotIn("LOOKUP_COLUMN_ALL_MISSING",
                             [w["code"] for w in plan["warnings"]])

    def test_lookup_column_all_hit_empty_values_no_warning(self):
        """键全部命中但索引存储值本身为空串 (真命中) → 不触发全列警告
        (命中 ≠ 缺失; 统计在 resolve_lookup 内按命中/缺失语义计数)."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            (Path(tmp) / "inheritance.json").write_text(json.dumps({
                "Z001": {"compressor": ""}, "Z002": {"compressor": ""},
                "Z003": {"compressor": ""},
            }), encoding="utf-8")
            plan = compile_fill.compile_spec(
                lookup_column_spec(wd), wd["manifest"], Path(tmp))
            self.assertNotIn("LOOKUP_COLUMN_ALL_MISSING",
                             [w["code"] for w in plan["warnings"]])

    # ── Q2: lookup key 归一化 (NBSP + strip) ──
    def test_lookup_key_normalized_nbsp(self):
        """源 key 含 NBSP 尾字符 → 归一化后命中索引 (non-empty)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp, src_rows=[
                ["家用", "12K", "Z001\xa0", "F-1", "C-1", "1", "2", "3"],
                ["家用", "18K", "Z002", "F-2", "C-2", "4", "5", "6"],
                ["商用", "24K", "Z003", "F-3", "C-3", "7", "8", "9"],
            ])
            wd["workdir"] = tmp
            (tmp / "inheritance.json").write_text(
                json.dumps({"Z001": {"compressor": "C-1"},
                            "Z002": {"compressor": "C-2"}}),
                encoding="utf-8")
            plan = compile_fill.compile_spec(
                lookup_column_spec(wd), wd["manifest"], tmp)
            f_vals = [w["value"] for w in plan["writes"] if w["col"] == "F"]
            self.assertEqual(f_vals[0], "C-1")  # Z001\xa0 → hit

    def test_lookup_key_normalized_strip(self):
        """源 key 含首尾空格 → 归一化后命中索引 (non-empty)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp, src_rows=[
                ["家用", "12K", " Z001 ", "F-1", "C-1", "1", "2", "3"],
                ["家用", "18K", "Z002", "F-2", "C-2", "4", "5", "6"],
                ["商用", "24K", "Z003", "F-3", "C-3", "7", "8", "9"],
            ])
            wd["workdir"] = tmp
            (tmp / "inheritance.json").write_text(
                json.dumps({"Z001": {"compressor": "C-1"},
                            "Z002": {"compressor": "C-2"}}),
                encoding="utf-8")
            plan = compile_fill.compile_spec(
                lookup_column_spec(wd), wd["manifest"], tmp)
            f_vals = [w["value"] for w in plan["writes"] if w["col"] == "F"]
            self.assertEqual(f_vals[0], "C-1")  # ' Z001 ' → hit


class LookupKeyColumnInvalidContractTests(unittest.TestCase):
    """issue 04: lookup key_column 契约 — 单缺陷码 LOOKUP_KEY_COLUMN_INVALID。

    key_column 必须是当前展平源数据中真实存在的 Excel 列字母; 接受集冻结 =
    col_letter_to_idx 一线语义 (大小写不敏感 1-2 个字母, 不另发明 regex)。
    列级/表级两个入口都在静态验证段拦截 (exit 3, 生成 plan 之前), reason
    区分 invalid_format / out_of_range; range 基准 = 该 lookup 实际 consumer
    source(s) 的展平宽度; resolve_lookup 保留兜底, 绝不裸 IndexError。"""

    def _compile_capture(self, wd: dict, spec: dict):
        """Compile with captured stderr; return (plan, exit_code, stderr)."""
        from io import StringIO
        buf = StringIO()
        old = sys.stderr
        sys.stderr = buf
        try:
            plan = compile_fill.compile_spec(
                spec, wd["manifest"], wd["workdir"])
            return plan, None, buf.getvalue()
        except SystemExit as e:
            return None, e.code, buf.getvalue()
        finally:
            sys.stderr = old

    def _lookup_spec(self, wd: dict, key_column: str = "C",
                     col_key: str | None = None) -> dict:
        spec = lookup_column_spec(wd, key_column=key_column)
        if col_key is not None:
            spec["mapping"]["targets"][0]["columns"][-1]["lookup"][
                "key_column"] = col_key
        return spec

    # ── 用例 1: 列级入口, malformed (invalid_format) ──
    def test_column_level_malformed_key_column_invalid_format(self):
        """columns[].lookup.key_column='sku' (逻辑键名) → exit 3 +
        LOOKUP_KEY_COLUMN_INVALID (invalid_format), plan 之前拦截,
        无裸 traceback。"""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            plan, code, err = self._compile_capture(
                wd, self._lookup_spec(wd, col_key="sku"))
            self.assertIsNone(plan)  # 编译在返回 plan 前失败
            self.assertEqual(code, 3)
            self.assertIn("LOOKUP_KEY_COLUMN_INVALID", err)
            self.assertIn("invalid_format", err)
            self.assertNotIn("Traceback", err)  # 结构化缺陷, 不裸崩
            defect = json.loads(err)["defects"][0]
            self.assertEqual(defect["code"], "LOOKUP_KEY_COLUMN_INVALID")
            self.assertEqual(defect["reason"], "invalid_format")
            self.assertEqual(
                defect["corrective_action"],
                "key_column='sku' is invalid: expected an Excel column letter "
                "present in the flattened source (e.g. G), not a logical "
                "field/key name.")

    # ── 用例 2: 表级入口, out-of-range ──
    def test_table_level_out_of_range_key_column(self):
        """lookups[].key_column='AB' + 默认 8 列源 (A:H) → exit 3 +
        LOOKUP_KEY_COLUMN_INVALID (out_of_range); range 基准 = 实际
        consumer source (source_maoli), 无裸 traceback。"""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            plan, code, err = self._compile_capture(
                wd, self._lookup_spec(wd, key_column="AB"))
            self.assertIsNone(plan)
            self.assertEqual(code, 3)
            self.assertIn("LOOKUP_KEY_COLUMN_INVALID", err)
            self.assertIn("out_of_range", err)
            self.assertNotIn("Traceback", err)
            defect = json.loads(err)["defects"][0]
            self.assertEqual(defect["reason"], "out_of_range")
            self.assertEqual(defect["source"], "source_maoli")
            self.assertIn(
                "key_column='AB' is out of range: flattened source has "
                "columns A:H. Choose an existing source column.",
                defect["corrective_action"])

    # ── 合法输入防回归 (行为零漂移) ──
    def test_lowercase_key_column_still_compiles(self):
        """key_column='c' (小写, 列存在) → compile 通过; guard 不收紧
        col_letter_to_idx 的大小写容忍。"""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            plan, code, err = self._compile_capture(
                wd, self._lookup_spec(wd, key_column="c"))
            self.assertIsNotNone(plan)
            self.assertIsNone(code)
            f_vals = [w["value"] for w in plan["writes"] if w["col"] == "F"]
            self.assertEqual(f_vals[0], "C-1")  # C 列 key 照常命中

    def test_double_letter_key_column_wide_source_compiles(self):
        """key_column='AB' + 28 列源 (A..AB) → compile 通过;
        out_of_range 只按实际 consumer source 宽度判, 不误拒合法 spec。"""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp, src_rows=[
                [f"r{i}" for i in range(28)] for _ in range(3)])
            wd["workdir"] = tmp
            plan, code, err = self._compile_capture(
                wd, self._lookup_spec(wd, key_column="AB"))
            self.assertIsNotNone(plan)
            self.assertIsNone(code)

    def test_unreferenced_table_lookup_not_checked(self):
        """表级 lookup 未被任何 column 引用 → 不消费不 crash, 不检查
        (含非法 key_column='sku' 也放行 — 检查它反而误拒合法 spec)。"""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["lookups"] = [
                {"name": "unused", "from": "inheritance.json",
                 "key_column": "sku", "fields": ["compressor"],
                 "missing": "empty"}]
            plan, code, err = self._compile_capture(wd, spec)
            self.assertIsNotNone(plan)
            self.assertIsNone(code)

    # ── resolve_lookup 兜底 (要求 6: 静态拦截命中后走不到裸 IndexError) ──
    def test_resolve_lookup_backstop_out_of_range(self):
        """静态拦截漏网时 resolve_lookup 也不再裸 IndexError — 同缺陷码 +
        out_of_range, 返回 None (该行跳过, 交给 MATERIALIZE_DEFECTS)。"""
        defects = []
        tbl = {"data": {"Z001": {"compressor": "C-1"}}, "key_column": "D"}
        out = compile_fill.resolve_lookup(
            {"name": "fields", "field": "compressor"}, ["a", "b"],
            {"fields": tbl}, defects)
        self.assertIsNone(out)
        self.assertEqual(defects[0]["code"], "LOOKUP_KEY_COLUMN_INVALID")
        self.assertEqual(defects[0]["reason"], "out_of_range")

    def test_resolve_lookup_backstop_invalid_format(self):
        """key_column='sku' 直接进 resolve_lookup → invalid_format 缺陷 +
        None (原 L1163 裸 values[col_letter_to_idx(kcol)] 已封死)。"""
        defects = []
        tbl = {"data": {}, "key_column": "sku"}
        out = compile_fill.resolve_lookup(
            {"name": "fields", "field": "compressor"}, ["a", "b"],
            {"fields": tbl}, defects)
        self.assertIsNone(out)
        self.assertEqual(defects[0]["code"], "LOOKUP_KEY_COLUMN_INVALID")
        self.assertEqual(defects[0]["reason"], "invalid_format")

    def test_falsy_key_column_path_unchanged(self):
        """key_column 缺失/为空 (falsy) → 不检查不报错 (现状非 crash 路径
        冻结为保持原样: key=None → miss → missing: empty 留空)。"""
        defects = []
        tbl = {"data": {"Z001": {"compressor": "C-1"}}, "key_column": None}
        out = compile_fill.resolve_lookup(
            {"name": "fields", "field": "compressor", "missing": "empty"},
            ["a", "b"], {"fields": tbl}, defects)
        self.assertEqual(out, "")
        self.assertEqual(defects, [])


class NumericPrecisionTests(unittest.TestCase):
    def test_long_precision_value_auto_rounded_with_warning(self):
        """15-digit cost value is AUTO-ROUNDED to 4 decimals with a warning
        (no execute round burned rediscovering the fix)."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)
            # force a long value via a constant mapping on top of the base columns
            spec["mapping"]["targets"][0]["columns"].append(
                {"target": "E", "value": "168.715100569657"})
            plan = compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            vals = [w["value"] for w in plan["writes"] if w["col"] == "E"]
            self.assertEqual(vals, ["168.7151"] * 3)  # auto-rounded in place
            codes = [w["code"] for w in plan["warnings"]]
            self.assertIn("AUTO_ROUND4", codes)

    def test_round4_transform_clears_the_defect(self):
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["columns"].append(
                {"target": "E", "value": "168.715100569657", "transform": "round4"})
            plan = compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            vals = [w["value"] for w in plan["writes"] if w["col"] == "E"]
            self.assertEqual(vals, ["168.7151"] * 3)
            self.assertEqual(plan["warnings"], [])  # explicit transform → no warning

    def test_precision_keep_explicitly_accepts(self):
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["columns"].append(
                {"target": "E", "value": "168.715100569657", "precision": "keep"})
            plan = compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            self.assertEqual(plan["operation_count"] > 0, True)
            self.assertEqual(plan["warnings"], [])  # explicit keep → untouched

    def _set_column_width(self, tmp: Path, widths: dict | None) -> None:
        """重写 fixture 的 target_meta.json 的 column_width (不入指纹, 无需重算)."""
        meta_path = tmp / "target_meta.json"
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        if widths is None:
            meta.pop("column_width", None)
        else:
            meta["column_width"] = widths
        meta_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")

    def _keep_spec(self, wd: dict) -> dict:
        spec = spec_with(wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "E", "value": "168.715100569657", "precision": "keep"})
        return spec

    def test_precision_keep_narrow_column_rejected(self):
        """prepare 已采集列宽 (E=8) 且该列最宽渲染值 (16 字符) 超出列宽 →
        PRECISION_KEEP_NARROW_COLUMN 编译拒绝 (exit 3) — 不再执行期才发现
        text_overflow."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            self._set_column_width(Path(tmp), {"E": 8})
            codes = compile_fail_codes(wd, self._keep_spec(wd))
            self.assertIn("PRECISION_KEEP_NARROW_COLUMN", codes)

    def test_precision_keep_wide_column_ok(self):
        """列宽 20 > 最宽渲染值 16 → keep 编译通过, 无警告."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            self._set_column_width(Path(tmp), {"E": 20})
            plan = compile_spec_with(wd, self._keep_spec(wd))
            self.assertGreater(plan["operation_count"], 0)
            self.assertEqual(plan["warnings"], [])

    def test_precision_keep_width_unknown_warns(self):
        """旧 meta 无 column_width → 保持豁免 (编译通过) + 警告提示无法验证."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            self._set_column_width(Path(tmp), None)
            plan = compile_spec_with(wd, self._keep_spec(wd))
            self.assertGreater(plan["operation_count"], 0)
            codes = [w["code"] for w in plan["warnings"]]
            self.assertIn("PRECISION_KEEP_WIDTH_UNVERIFIED", codes)

    def test_precision_keep_narrow_round4_clears_defect(self):
        """corrective_action 落地: 同一窄列 (E=8) 改用 transform: round4 →
        编译通过, 无警告."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            self._set_column_width(Path(tmp), {"E": 8})
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["columns"].append(
                {"target": "E", "value": "168.715100569657", "transform": "round4"})
            plan = compile_spec_with(wd, spec)
            vals = [w["value"] for w in plan["writes"] if w["col"] == "E"]
            self.assertEqual(vals, ["168.7151"] * 3)

    def test_overlong_integer_still_hard_fails(self):
        """>12-digit integers cannot be shortened by rounding → hard defect."""
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["columns"].append(
                {"target": "E", "value": "12345678901234"})  # 14 digits, no decimals
            from io import StringIO
            buf = StringIO()
            old = sys.stderr
            sys.stderr = buf
            try:
                with self.assertRaises(SystemExit) as ctx:
                    compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
                self.assertEqual(ctx.exception.code, 3)
            finally:
                sys.stderr = old
            self.assertIn("NUMERIC_OVERFLOW_RISK", buf.getvalue())

    def test_short_numbers_pass(self):
        with tempfile.TemporaryDirectory() as tmp:
            wd = make_workdir(Path(tmp))
            wd["workdir"] = Path(tmp)
            spec = spec_with(wd)  # base spec values are short
            plan = compile_fill.compile_spec(spec, wd["manifest"], Path(tmp))
            self.assertGreater(plan["operation_count"], 0)


class PrecisionWidthEstimatorTests(unittest.TestCase):
    """estimate_rendered_width 边界固定 (启发式的机器可验证确定性边界 —
    低估才会放行执行期溢出, 高估只会建议 round4 = 文档首选, 安全方向).

    与 Q7 契约同源: keep 列宽校验的渲染宽度就是这里的估算."""

    def test_general_uses_raw_length(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("168.715100569657"), 16)
        self.assertEqual(est("168.715100569657", "General"), 16)
        # 无 0/# 占位符的格式 (如日期/纯文本) → 按原始字符串长度保守计
        self.assertEqual(est("168.715100569657", "yyyy-mm-dd"), 16)

    def test_format_truncates_decimals(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("168.715100569657", "0.00"), 6)       # 168.72
        self.assertEqual(est("-168.7151", "0.00"), 7)              # -168.72

    def test_thousands_separators(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("1234567.89", "#,##0.00"), 12)        # 1,234,567.89

    def test_zero_pad_integer(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("168.5", "000000.00"), 9)             # 000168.50

    def test_quoted_literals(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("168.5", '0"万元"'), 5)               # 168万元

    def test_parenthesized_negative(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("-168.5", "(0.00)"), 8)               # (168.50), 无负号

    def test_exponent_suffix_counted(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("168.715100569657", "0.00E+00"), 10)  # 保守高估 (安全方向)

    def test_currency_and_percent_symbols(self):
        est = compile_fill.estimate_rendered_width
        self.assertEqual(est("168.5", '"$"#,##0.00'), 7)           # $168.50 (引号内 $ 不双计)
        self.assertEqual(est("0.1685", "0.00%"), 6)                # 16.85% (×100 缩放计入)
        self.assertEqual(est("168.7151", "0.00%"), 9)              # 16871.51%


class OverflowDetectorBackingTests(unittest.TestCase):
    """执行期 text_overflow 检测器仍有单测背书 — keep 窄列前移到编译期拒绝后,
    execute 期检测器是第二道防线, 不得静默回归 (issue 04 契约闭环)."""

    def test_classify_issue_overflow_messages(self):
        from _defect_class import classify_issue
        self.assertEqual(classify_issue("A5", "overflow",
                                        "value overflows column width"), "text_overflow")
        self.assertEqual(classify_issue("A5", "",
                                        "需要 24pt 才能显示"), "text_overflow")
        self.assertEqual(classify_issue("A5", "empty",
                                        "cell value is empty"), "empty_cell")
        self.assertEqual(classify_issue("A5", "", "ok"), "unknown")


class NotePhaseTests(unittest.TestCase):
    def test_note_phase_records_gap(self):
        import note_phase
        with tempfile.TemporaryDirectory() as tmp:
            workdir = Path(tmp)
            (workdir / "run_timing.json").write_text(json.dumps([
                {"kind": "machine", "phase": "prepare_flatten",
                 "started_at": "2026-08-10T10:00:00", "duration_ms": 10000},
            ]), encoding="utf-8")
            with self.assertRaises(SystemExit) as ctx:
                sys.argv = ["note_phase", "--workdir", str(workdir), "--phase", "spec_authoring"]
                note_phase.main()
            self.assertEqual(ctx.exception.code, 0)
            entries = json.loads((workdir / "run_timing.json").read_text(encoding="utf-8"))
            agent = entries[-1]
            self.assertEqual(agent["kind"], "agent")
            self.assertEqual(agent["phase"], "spec_authoring")
            # gap since the machine entry finished (>=0s, measured in wall time)
            self.assertGreaterEqual(agent["duration_ms"], 0)
            self.assertEqual(agent["started_at"], "2026-08-10T10:00:10")

    def test_note_phase_chains_agent_entries(self):
        import note_phase
        with tempfile.TemporaryDirectory() as tmp:
            workdir = Path(tmp)
            (workdir / "run_timing.json").write_text(json.dumps([
                {"kind": "agent", "phase": "spec_authoring",
                 "started_at": "2026-08-10T10:00:00", "duration_ms": 600000},
            ]), encoding="utf-8")
            with self.assertRaises(SystemExit):
                sys.argv = ["note_phase", "--workdir", str(workdir), "--phase", "compile_review"]
                note_phase.main()
            entries = json.loads((workdir / "run_timing.json").read_text(encoding="utf-8"))
            self.assertEqual(entries[-1]["started_at"], "2026-08-10T10:10:00")  # 10min later
            self.assertEqual(entries[-1]["kind"], "agent")


class MultiSourceTests(unittest.TestCase):
    def test_two_sources_merge_in_order(self):
        """rows.sources concatenates matched rows of both sheets in list order."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            # second source sheet: another CSV with 2 rows
            with open(tmp / "source_commercial_flat.csv", "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["商用", "10P", "Z500", "F-5", "C-5", "10", "11", "12", "201"])
                w.writerow(["商用", "20P", "Z501", "F-6", "C-6", "13", "14", "15", "202"])
            wd["manifest"]["flattened"].append(
                {"file": "source_commercial.xlsx", "sheet": "商用", "name": "source_commercial",
                 "csv": "source_commercial_flat.csv", "meta": "m.json",
                 "digest": "d.md", "candidates": "c.yaml"})
            wd["manifest"]["files"].append(
                {"staged": "source_commercial.xlsx", "source": "x", "sha256": "c"})
            wd["workdir"] = tmp
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["rows"] = {
                "sources": [
                    {"source": "source_maoli",
                     "selectors": [{"column": "A", "pattern": "家用*"}]},
                    {"source": "source_commercial"},
                ],
            }
            spec["validation"]["required_coverage"] = [
                {"source": "source_maoli_flat.csv", "rows": [101]},
                {"source": "source_commercial_flat.csv", "rows": [201, 202]},
            ]
            plan = compile_fill.compile_spec(spec, wd["manifest"], tmp)
            # 2 (家用) + 2 (商用) rows → data 7-10 (title 5, header 6)
            self.assertEqual(plan["source_coverage"][0]["matched"], 2)
            self.assertEqual(plan["source_coverage"][1]["matched"], 2)
            self.assertEqual(plan["source_coverage"][1]["source"], "source_commercial_flat.csv")
            rows = [t for _, _, t in plan["row_map"]]
            self.assertEqual(rows, [7, 8, 9, 10])
            self.assertEqual(plan["source_coverage"][0]["required_unmatched"], [])
            self.assertEqual(plan["source_coverage"][1]["required_unmatched"], [])


class PrepareMergeTests(unittest.TestCase):
    def test_merge_flattened_incremental(self):
        from prepare_run import merge_flattened
        first = [
            {"name": "source_maoli_FRESH", "csv": "a.csv"},
            {"name": "target_baojia", "csv": "t.csv"},
        ]
        second = [
            {"name": "source_maoli_COMMERCIAL", "csv": "b.csv"},
        ]
        merged = merge_flattened(first, second)
        self.assertEqual([e["name"] for e in merged],
                         ["source_maoli_FRESH", "target_baojia", "source_maoli_COMMERCIAL"])
        # 同名条目新覆盖旧
        again = merge_flattened(merged, [{"name": "source_maoli_FRESH", "csv": "a2.csv"}])
        by_name = {e["name"]: e for e in again}
        self.assertEqual(by_name["source_maoli_FRESH"]["csv"], "a2.csv")
        self.assertEqual(len(again), 3)


class PrepareStyleGranularityTests(unittest.TestCase):
    """Issue 03: prepare 阶段 B 样式粒度决策事实 (占位行裸行 vs 带样式).

    埃及形态: 占位行 23-52 裸行 (无边框/填充/字体) → clone-append 正确终态;
    MXP 报价单形态: 占位行带样式 (空单元格持边框/填充) → inplace 成立。
    指纹计算不变: style_granularity 不入 structure_facts。
    """

    def _make_template(self, path, content_rows, placeholder_runs, styled_empty):
        """Synthetic xlsx: content rows (borders+fill) + placeholder runs.

        styled_empty=True → 占位行空单元格携带样式 (MXP 报价单形态);
        False → 裸行 (row 元素仅行高, 无单元格, 埃及形态)."""
        from openpyxl import Workbook
        from openpyxl.styles import Border, PatternFill, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="FFF2F2F2")
        for r in range(1, content_rows + 1):
            for c in ("A", "B", "C"):
                cell = ws[f"{c}{r}"]
                cell.value = f"v{r}"
                cell.border = border
                cell.fill = fill
        for (start, end) in placeholder_runs:
            for r in range(start, end + 1):
                if styled_empty:
                    for c in ("A", "B", "C"):
                        ws[f"{c}{r}"].border = border
                else:
                    ws.row_dimensions[r].height = 24
        wb.save(path)

    def _detect(self, path, blocks, flat_rows, num_cols=6):
        from flatten_table import detect_style_granularity
        return detect_style_granularity(str(path), "S", blocks, flat_rows, num_cols)

    def test_egypt_form_placeholder_bare_rows(self):
        """埃及形态: 占位行 23-52 裸行 (无单元格 → 无样式), 克隆源行带样式."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            p = tmp / "egypt.xlsx"
            self._make_template(p, content_rows=21, placeholder_runs=[(23, 52)],
                                styled_empty=False)
            blocks = [{"start": 1, "end": 5, "title": "B1"},
                      {"start": 7, "end": 15, "title": "B2"},
                      {"start": 17, "end": 21, "title": "B3"}]
            flat_rows = [[f"v{i}", str(i)] for i in range(1, 22)]
            sg = self._detect(p, blocks, flat_rows)
            self.assertEqual(
                sg["placeholder_segments"],
                [{"start": 23, "end": 52, "styled": False, "sample": None}])
            # 克隆源行 (title/header/data) 全部带样式 — 克隆携带格式的事实依据
            b1 = sg["clone_source_rows"][0]
            for role in ("title", "header", "data"):
                self.assertTrue(b1[role]["styled"], role)
                self.assertEqual(b1[role]["row"],
                                 {"title": 1, "header": 2, "data": 3}[role])
            self.assertEqual(len(sg["clone_source_rows"]), 3)

    def test_mxp_form_placeholder_styled_rows(self):
        """MXP 报价单形态: 占位行空单元格带样式 → 带样式 (样例: A7)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            p = tmp / "mxp.xlsx"
            self._make_template(p, content_rows=6, placeholder_runs=[(7, 24)],
                                styled_empty=True)
            blocks = [{"start": 1, "end": 6, "title": "报价单"}]
            flat_rows = [[f"v{i}", str(i)] for i in range(1, 7)]
            sg = self._detect(p, blocks, flat_rows)
            self.assertEqual(
                sg["placeholder_segments"],
                [{"start": 7, "end": 24, "styled": True, "sample": "A7"}])

    def test_digest_lines_both_forms(self):
        """digest 输出结论行: 裸行 (带段范围) / 带样式 (样例坐标); 克隆源行行."""
        from structure_digest import build_digest
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            p = tmp / "egypt.xlsx"
            self._make_template(p, content_rows=21, placeholder_runs=[(23, 52)],
                                styled_empty=False)
            blocks = [{"start": 1, "end": 5, "title": "B1"}]
            flat_rows = [[f"v{i}", str(i)] for i in range(1, 22)]
            sg = self._detect(p, blocks, flat_rows)
            meta = {"file": str(p), "sheet": "S",
                    "dimensions": {"rows": 52, "cols": 6},
                    "style_granularity": sg}
            lines = build_digest(meta, None, None, for_target=True)
            self.assertTrue(any("占位行样式: 裸行 (23-52)" in l for l in lines))
            self.assertTrue(any(l.startswith("- 克隆源行样式: B1(") for l in lines))

            p2 = tmp / "mxp.xlsx"
            self._make_template(p2, content_rows=6, placeholder_runs=[(7, 24)],
                                styled_empty=True)
            sg2 = self._detect(p2, [{"start": 1, "end": 6, "title": "报价单"}],
                               [[f"v{i}", str(i)] for i in range(1, 7)])
            meta2 = {"file": str(p2), "sheet": "S",
                     "dimensions": {"rows": 24, "cols": 6},
                     "style_granularity": sg2}
            lines2 = build_digest(meta2, None, None, for_target=True)
            self.assertTrue(any("占位行样式: 带样式 (样例: A7)" in l for l in lines2))

    def test_style_granularity_not_in_fingerprint(self):
        """style_granularity 不入指纹: 带/不带该事实的 meta 指纹相同 (旧 spec
        重编译不触发 fingerprint 失效)."""
        from prepare_run import facts_sha256, structure_facts
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            p = tmp / "egypt.xlsx"
            self._make_template(p, content_rows=21, placeholder_runs=[(23, 52)],
                                styled_empty=False)
            blocks = [{"start": 1, "end": 5, "title": "B1"}]
            flat_rows = [[f"v{i}", str(i)] for i in range(1, 22)]
            sg = self._detect(p, blocks, flat_rows)
            base_meta = {"sheet": "S",
                         "dimensions": {"rows": 52, "cols": 6, "data_rows": 21},
                         "header_band": None,
                         "merged_ranges": [],
                         "blocks": blocks,
                         "columns": [{"col": "A", "nonempty": 21, "numeric_ratio": 0.0}],
                         "formulas": {}, "column_numfmt": {},
                         "merge_anchors": []}
            self.assertEqual(facts_sha256([structure_facts(base_meta)]),
                             facts_sha256([structure_facts(
                                 {**base_meta, "style_granularity": sg})]))

    def test_fill_id_1_gray125_not_styled(self):
        """fillId=1 (gray125 占位填充) 不算带样式: 仅 fillId>=2 的真实填充才
        判带样式 (LibreOffice/默认样式引用 gray125 时不得误报 带样式)."""
        import zipfile
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            p = tmp / "gray125.xlsx"
            # 手写最小 xlsx: styles.xml 定义 3 个样式 (0=默认, 1=gray125,
            # 2=真实填充), sheet 的占位行单元格引用 s=1
            styles_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<fonts count="1"><font><sz val="11"/></font></fonts>'
                '<fills count="3">'
                '<fill><patternFill patternType="none"/></fill>'
                '<fill><patternFill patternType="gray125"/></fill>'
                '<fill><patternFill patternType="solid"><fgColor rgb="FFFFF2F2"/></patternFill></fill>'
                '</fills>'
                '<borders count="1"><border/></borders>'
                '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
                '<cellXfs count="3">'
                '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
                '<xf numFmtId="0" fontId="0" fillId="1" borderId="0" xfId="0" applyFill="1"/>'
                '<xf numFmtId="0" fontId="0" fillId="2" borderId="0" xfId="0" applyFill="1"/>'
                '</cellXfs>'
                '</styleSheet>'
            )
            sheet_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<sheetData>'
                '<row r="1"><c r="A1" t="inlineStr"><is><t>v1</t></is></c></row>'
                '<row r="2"><c r="A2" s="1"/></row>'
                '<row r="3"><c r="A3" s="2"/></row>'
                '</sheetData></worksheet>'
            )
            wb_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                      '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                      'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                      '<sheets><sheet name="S" sheetId="1" r:id="rId1"/></sheets></workbook>')
            rels_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
                        '</Relationships>')
            ct_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                      '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                      '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.'
                      'relationships+xml"/>'
                      '<Default Extension="xml" ContentType="application/xml"/>'
                      '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-'
                      'officedocument.spreadsheetml.sheet.main+xml"/>'
                      '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.'
                      'openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                      '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-'
                      'officedocument.spreadsheetml.styles+xml"/>'
                      '</Types>')
            with zipfile.ZipFile(p, "w") as z:
                z.writestr("[Content_Types].xml", ct_xml)
                z.writestr("_rels/.rels", '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
                            'relationships"><Relationship Id="rId1" Type="http://schemas.'
                            'openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
                            'Target="xl/workbook.xml"/></Relationships>')
                z.writestr("xl/workbook.xml", wb_xml)
                z.writestr("xl/_rels/workbook.xml.rels", rels_xml)
                z.writestr("xl/styles.xml", styles_xml)
                z.writestr("xl/worksheets/sheet1.xml", sheet_xml)
            sg = self._detect(p, [{"start": 1, "end": 1, "title": "B1"}],
                              [["v1", "1"]], num_cols=3)
            # 占位行段: row 2 (gray125 → 裸行) 与 row 3 (solid → 带样式)
            # 连续成一段, 但 gray125 单元格不得使段判带样式 — 样例必须来自
            # row 3 的真实填充 (A3).
            self.assertEqual(sg["placeholder_segments"],
                             [{"start": 2, "end": 3, "styled": True, "sample": "A3"}])

    def test_collect_style_granularity_manifest_wiring(self):
        """prepare_run 把 flatten meta 的样式粒度事实并入 manifest (按条目名)."""
        from prepare_run import collect_style_granularity
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            p = tmp / "egypt.xlsx"
            self._make_template(p, content_rows=21, placeholder_runs=[(23, 52)],
                                styled_empty=False)
            sg = self._detect(p, [{"start": 1, "end": 5, "title": "B1"}],
                              [[f"v{i}", str(i)] for i in range(1, 22)])
            out = collect_style_granularity({
                "src": {"columns": []},
                "target": {"style_granularity": sg},
            })
            self.assertNotIn("src", out)
            self.assertIn("target", out)
            self.assertEqual(out["target"]["placeholder_segments"][0]["start"], 23)


class DetectHeaderRowsContractTests(unittest.TestCase):
    """Case 010: detect_header_rows 跨空行间隙 + 密度不降约束契约.

    修复背景: ATLAS Quotation 模板 row 4 (To Messrs: ATLAS) 被识别为表头带
    起点, row 5 空行导致 row 6 真列头 (Type/Model/...) 因"不连续"被 break,
    header_band=[4]/data_start=5, 六角色 dimension_set 信号 missed。
    修复: 允许表头带跨空行间隙继续, 但新行 nonempty 不得低于带内最大值
    (列头行是表里最满的行, text-dense 数据行有空列 → 密度下降 → 不被吸入)。
    """

    def _cells(self, rows_profile):
        """构造 fake cell 列表: {row: [(col, type, text), ...]}."""
        cells = []
        for r, entries in rows_profile.items():
            for col, ctype, text in entries:
                cells.append({
                    "path": f"/S/{col}{r}",
                    "format": {"type": ctype},
                    "text": text,
                })
        return cells

    def test_blank_gap_spanning_header_band(self):
        """ATLAS 形态: row4 标题(2) + 空行 + row6 列头(6) → band [4,6], data 7."""
        from flatten_table import detect_header_rows
        profile = {
            4: [("A", "SharedString", "To Messrs: ATLAS"),
                ("F", "SharedString", "Date of issue")],
            6: [("A", "SharedString", "Type"),
                ("B", "SharedString", "Model"),
                ("C", "SharedString", "Capacity"),
                ("D", "SharedString", "Pipe"),
                ("E", "SharedString", "Unit Price"),
                ("F", "SharedString", "Panel looking")],
            7: [("A", "SharedString", "Type text"),
                ("B", "SharedString", "9K"),
                ("C", "SharedString", "9000Btu"),
                ("D", "SharedString", "/")],
            8: [("A", "SharedString", "Type text"),
                ("B", "SharedString", "12K"),
                ("C", "SharedString", "12000Btu"),
                ("D", "SharedString", "/")],
        }
        out = detect_header_rows(self._cells(profile), 6)
        self.assertEqual(out, {"header_rows": [4, 6], "data_start_row": 7})

    def test_text_dense_data_rows_not_absorbed(self):
        """数据行 nonempty(4) < 带内最大值(6) → break, 不吸入."""
        from flatten_table import detect_header_rows
        profile = {
            4: [("A", "SharedString", "Title"), ("B", "SharedString", "Info")],
            6: [("A", "SharedString", "H1"), ("B", "SharedString", "H2"),
                ("C", "SharedString", "H3"), ("D", "SharedString", "H4"),
                ("E", "SharedString", "H5"), ("F", "SharedString", "H6")],
            7: [("A", "SharedString", "a"), ("B", "SharedString", "b"),
                ("C", "SharedString", "c"), ("D", "SharedString", "d")],
            8: [("A", "SharedString", "e"), ("B", "SharedString", "f"),
                ("C", "SharedString", "g"), ("D", "SharedString", "h")],
        }
        out = detect_header_rows(self._cells(profile), 6)
        self.assertEqual(out, {"header_rows": [4, 6], "data_start_row": 7})

    def test_no_gap_consecutive_band_unchanged(self):
        """无空行间隙的连续表头带 (报价汇总 24 列形态) 行为不变: [2]/3.

        数据行是 text 少数派 (2 text + 4 number → text < nonempty//2)
        → 非 header → break, 与修复前一致."""
        from flatten_table import detect_header_rows
        profile = {
            1: [("A", "SharedString", "TITLE")],
            2: [("A", "SharedString", "类别"), ("B", "SharedString", "产品类别"),
                ("C", "SharedString", "订单明细"), ("D", "SharedString", "工厂型号"),
                ("E", "SharedString", "配置描述"), ("F", "SharedString", "压缩机")],
            3: [("A", "SharedString", "Pioneer"), ("B", "SharedString", "9KCH"),
                ("C", "Number", "500"), ("D", "Number", "236.53"),
                ("E", "Number", "0"), ("F", "Number", "0")],
        }
        out = detect_header_rows(self._cells(profile), 6)
        self.assertEqual(out, {"header_rows": [2], "data_start_row": 3})

    def test_numeric_data_row_breaks_band(self):
        """数据行 majority-numeric (text < nonempty//2) → 非 header → break."""
        from flatten_table import detect_header_rows
        profile = {
            2: [("A", "SharedString", "类别"), ("B", "SharedString", "产品类别"),
                ("C", "SharedString", "订单明细"), ("D", "SharedString", "工厂型号"),
                ("E", "SharedString", "配置描述"), ("F", "SharedString", "压缩机")],
            3: [("A", "SharedString", "Pioneer"), ("B", "SharedString", "9KCH"),
                ("C", "Number", "500"), ("D", "Number", "236.53"),
                ("E", "Number", "0"), ("F", "Number", "0")],
        }
        out = detect_header_rows(self._cells(profile), 6)
        self.assertEqual(out, {"header_rows": [2], "data_start_row": 3})


class DecisionStringTests(unittest.TestCase):
    def test_decisions_with_colon_parsed_as_dict_rejected(self):
        """decisions 条目含 ': ' 未加引号 → YAML 解析成 dict → 编译期报错."""
        spec = {
            "task": {"intent": "t", "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_maoli.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_maoli.xlsx", "sheets": ["s"]}],
                       "target_sheet": "S"},
            "fingerprints": {"source_structure": "x", "target_structure": "y"},
            "mapping": {"targets": [{"sheet": "S"}]},
            "decisions": [{"标题确认": "由用户在 Gate 裁决"}],  # dict 而非 str
            "gaps": [],
            "lineage": [{"source": "s.csv", "role": "primary", "note": ""}],
            "validation": {"required_coverage": [], "required_empty": [], "key_outputs": []},
        }
        defects = compile_fill.validate_schema(spec, {"files": [], "flattened": []})
        codes = {d["code"] for d in defects}
        self.assertIn("SPEC_NON_STRING_ITEM", codes)
        msg = next(d["message"] for d in defects if d["code"] == "SPEC_NON_STRING_ITEM")
        self.assertIn("': '", msg)

    def test_spec_non_string_item_msg(self):
        """SPEC_NON_STRING_ITEM corrective_action 直接教整行双引号写法 (含
        冒号整行加引号), 示例用原文案 (k: v 重构) 而非 dict repr — 埃及 FRESH
        复盘: 旧建议 `- "{'追加新历史块': '...'}"` 把整个 dict 包成字符串,
        语法不可行, Agent 得自己领悟."""
        spec = {
            "task": {"intent": "t", "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_maoli.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_maoli.xlsx", "sheets": ["s"]}],
                       "target_sheet": "S"},
            "fingerprints": {"source_structure": "x", "target_structure": "y"},
            "mapping": {"targets": [{"sheet": "S"}]},
            "decisions": [{"追加新历史块": "源文件 11_FRESH本土 的毛利数据"},
                          {"备注": '源文件 "11" 的毛利'}],
            "gaps": [],
            "lineage": [{"source": "s.csv", "role": "primary", "note": ""}],
            "validation": {"required_coverage": [], "required_empty": [], "key_outputs": []},
        }
        defects = compile_fill.validate_schema(spec, {"files": [], "flattened": []})
        cas = [d["corrective_action"] for d in defects
               if d["code"] == "SPEC_NON_STRING_ITEM"]
        self.assertEqual(len(cas), 2)
        ca = cas[0]
        self.assertIn("用双引号包裹整行", ca)
        self.assertIn('- "追加新历史块: 源文件 11_FRESH本土 的毛利数据"', ca)
        self.assertNotIn("'追加新历史块':", ca)
        self.assertIn('- "备注: 源文件 \\"11\\" 的毛利"', cas[1],
                      "值内含双引号时 corrective 示例必须转义, 保持照抄即用")

    def test_quoted_decisions_pass(self):
        spec = {
            "task": {"intent": "t", "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_maoli.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_maoli.xlsx", "sheets": ["s"]}],
                       "target_sheet": "S"},
            "fingerprints": {"source_structure": "x", "target_structure": "y"},
            "mapping": {"targets": [{"sheet": "S"}]},
            "decisions": ["标题确认: 由用户在 Gate 裁决"],
            "gaps": [],
            "lineage": [{"source": "s.csv", "role": "primary", "note": ""}],
            "validation": {"required_coverage": [], "required_empty": [], "key_outputs": []},
        }
        defects = compile_fill.validate_schema(spec, {"files": [], "flattened": []})
        self.assertEqual([d for d in defects if d["code"] == "SPEC_NON_STRING_ITEM"], [])


class MultiBlockTests(unittest.TestCase):
    def test_two_blocks_sequential_layout(self):
        """blocks[] lays out sequentially: block1 spacer/title/header/data then
        block2 spacer/title/header/data; each block has its own aggregates."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            with open(tmp / "source_commercial_flat.csv", "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["商用", "10P", "Z500", "F-5", "C-5", "10", "11", "12", "201"])
                w.writerow(["商用", "20P", "Z501", "F-6", "C-6", "13", "14", "15", "202"])
            wd["manifest"]["flattened"].append(
                {"file": "source_commercial.xlsx", "sheet": "商用", "name": "source_commercial",
                 "csv": "source_commercial_flat.csv", "meta": "m.json",
                 "digest": "d.md", "candidates": "c.yaml"})
            wd["manifest"]["files"].append(
                {"staged": "source_commercial.xlsx", "source": "x", "sha256": "c"})
            wd["workdir"] = tmp
            spec = spec_with(wd)
            block_template = {
                "clone_roles": [
                    {"role": "spacer"},
                    {"role": "title", "template_row": 1, "value": "块标题"},
                    {"role": "header", "template_row": 2},
                    {"role": "data", "template_row": 3},
                ],
                "formulas": {
                    "per_row": {"E": "A{r}*2"},
                    "aggregates": [{"col": "F", "rows": "1:{n}",
                                    "formula": "SUM(E{r1}:E{r2})", "style": "anchor"}],
                },
            }
            b1 = dict(block_template, rows={"source": "source_maoli",
                                            "selectors": [{"column": "A", "pattern": "家用*"}]})
            b2 = dict(block_template, rows={"source": "source_commercial"})
            spec["mapping"]["targets"][0]["blocks"] = [b1, b2]
            spec["validation"]["required_coverage"] = [
                {"source": "source_maoli_flat.csv", "rows": [101]},
                {"source": "source_commercial_flat.csv", "rows": [201, 202]},
            ]
            spec["validation"]["key_outputs"] = ["A8", "F8", "A13"]
            spec["validation"]["key_outputs"] = ["A8", "F8", "A13"]
            plan = compile_fill.compile_spec(spec, wd["manifest"], tmp)
            # block1: spacer5 title6 header7 data8-9; block2: spacer10 title11 header12 data13-14
            rows = [t for _, _, t in plan["row_map"]]
            self.assertEqual(rows, [8, 9, 13, 14])
            self.assertEqual(plan["source_coverage"][0]["block"], 0)
            self.assertEqual(plan["source_coverage"][1]["block"], 1)
            # aggregates: F8 (block1) and F13 (block2), each SUM over its own block
            agg_paths = [op["path"] for op in plan["operations"]
                         if "formula" in op.get("props", {}) and "SUM" in op["props"]["formula"]]
            self.assertEqual(agg_paths, ["/S/F8", "/S/F13"])
            # titles: A6 and A11
            title_ops = [op for op in plan["operations"]
                         if op.get("path", "").endswith(("A6", "A11"))
                         and "value" in op.get("props", {})]
            self.assertEqual(len(title_ops), 2)
            # per-row formula E with block-local {n}: block1 rows 8-9 → n=2, block2 rows 13-14 → n=2
            e_formulas = [op["props"]["formula"] for op in plan["operations"]
                          if op.get("path", "").endswith("/E8") or op.get("path", "").endswith("/E13")]
            self.assertEqual(e_formulas, ["A8*2", "A13*2"])


class SourceConsumptionUniquenessTests(unittest.TestCase):
    """Issue 05: (source, original_row) 全局恰好一次 — 跨块/跨源条目重复消费
    编译期拒绝 (fail-closed); 单消费 (MultiSource/MultiBlock 既有场景) 不回归."""

    def _two_blocks_same_source_spec(self, wd, sel_a, sel_b) -> dict:
        spec = spec_with(wd)
        block_template = {
            "clone_roles": [
                {"role": "spacer"},
                {"role": "title", "template_row": 1, "value": "块标题"},
                {"role": "header", "template_row": 2},
                {"role": "data", "template_row": 3},
            ],
        }
        b1 = dict(block_template, rows={"source": "source_maoli",
                                        "selectors": sel_a})
        b2 = dict(block_template, rows={"source": "source_maoli",
                                        "selectors": sel_b})
        spec["mapping"]["targets"][0]["blocks"] = [b1, b2]
        spec["validation"]["key_outputs"] = ["A8", "A13"]
        return spec

    def test_two_blocks_overlapping_selectors_rejected(self):
        """两个 block 的 selectors 命中同一源行 → SOURCE_ROW_CONSUMED_TWICE
        (缺陷码 + corrective_action), 不再静默重复消费."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            spec = self._two_blocks_same_source_spec(
                wd, [{"column": "A", "pattern": "家用*"}],
                [{"column": "A", "pattern": "家用*"}])
            from io import StringIO
            buf = StringIO()
            old = sys.stderr
            sys.stderr = buf
            try:
                with self.assertRaises(SystemExit) as ctx:
                    compile_fill.compile_spec(spec, wd["manifest"], tmp)
                self.assertEqual(ctx.exception.code, 3)
            finally:
                sys.stderr = old
            payload = json.loads(buf.getvalue())
            self.assertEqual(payload["code"], "SOURCE_ROW_CONSUMED_TWICE")
            self.assertIn("source_maoli_flat.csv", payload["message"])
            self.assertIn("101", payload["message"])
            self.assertTrue(payload.get("corrective_action"))

    def test_same_block_two_sources_entries_overlapping_rejected(self):
        """同一块 rows.sources 两个条目选中同一源行 → SOURCE_ROW_CONSUMED_TWICE."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["rows"] = {
                "sources": [
                    {"source": "source_maoli",
                     "selectors": [{"column": "A", "pattern": "家用*"}]},
                    {"source": "source_maoli",
                     "selectors": [{"column": "A", "pattern": "家用*"}]},
                ],
            }
            codes = compile_fail_codes(wd, spec)
            self.assertIn("SOURCE_ROW_CONSUMED_TWICE", codes)

    def test_disjoint_selectors_across_blocks_still_compiles(self):
        """块间 selectors 不相交 (家用* vs 商用*) → 编译通过, row_map 每行唯一."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            spec = self._two_blocks_same_source_spec(
                wd, [{"column": "A", "pattern": "家用*"}],
                [{"column": "A", "pattern": "商用*"}])
            plan = compile_fill.compile_spec(spec, wd["manifest"], tmp)
            pairs = [(s, o) for s, o, _ in plan["row_map"]]
            self.assertEqual(len(pairs), len(set(pairs)))
            self.assertEqual(len(pairs), 3)  # 家用×2 (block1) + 商用×1 (block2)


class InplacePositionModelTests(unittest.TestCase):
    """ADR 0007 position-model boundary: inplace reuse / trim / hybrid overflow /
    group_merges / sets / props."""

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_inplace_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _spec(self, **mutations):
        spec = copy.deepcopy(INPLACE_BASE_SPEC)
        spec["fingerprints"] = {
            "source_structure": self.wd["manifest"]["fingerprints"]["source_structure"],
            "target_structure": self.wd["manifest"]["fingerprints"]["target_structure"],
        }
        for path, value in mutations.items():
            node = spec
            for p in path.split(".")[:-1]:
                node = node[p] if isinstance(node, dict) else node[int(p)]
            if isinstance(node, list):
                node[int(path.split(".")[-1])] = value
            else:
                node[path.split(".")[-1]] = value
        return spec

    def _compile(self, spec):
        return compile_fill.compile_spec(spec, self.wd["manifest"], self.wd["workdir"])

    def _compile_fails_with(self, spec, code):
        from io import StringIO
        buf = StringIO()
        old = sys.stderr
        sys.stderr = buf
        try:
            with self.assertRaises(SystemExit) as ctx:
                self._compile(spec)
            self.assertEqual(ctx.exception.code, 3)
        finally:
            sys.stderr = old
        self.assertIn(code, buf.getvalue())
        return buf.getvalue()

    # ── Row layout: inplace region, trim, hybrid overflow ──
    def test_inplace_region_no_adds_trim_tail(self):
        plan = self._compile(self._spec())
        # region rows exist — no add ops; trim removes the surplus tail
        self.assertEqual([op for op in plan["operations"] if op["command"] == "add"], [])
        removes = [op["path"] for op in plan["operations"] if op["command"] == "remove"]
        self.assertEqual(removes, ["/S/row[10]"])
        # fills land on the retained region rows 7-9
        filled = {op["path"] for op in plan["operations"]
                  if op.get("props", {}).get("value") is not None}
        self.assertIn("/S/A7", filled)
        self.assertIn("/S/C9", filled)
        self.assertEqual(plan["expected_final_row_count"], 13)
        self.assertEqual(plan["structural_deltas"], {"adds": 0, "removes": 1,
                                                     "inplace_trim": 1,
                                                     "inplace_overflow": 0})
        self.assertEqual(plan["schema_version"], "2.5")
        self.assertEqual(plan["blocks"][-1]["mode"], "inplace")

    def test_inplace_hybrid_overflow_clones_after_region(self):
        wd = make_inplace_workdir(self.tmp, n_source_rows=5)
        wd["workdir"] = self.tmp
        spec = self._spec()
        spec["fingerprints"] = wd["manifest"]["fingerprints"]
        plan = compile_fill.compile_spec(spec, wd["manifest"], self.tmp)
        adds = [op for op in plan["operations"] if op["command"] == "add"]
        self.assertEqual(len(adds), 1)  # N=5 > capacity=4 → 1 overflow clone
        self.assertEqual([a["after"] for a in adds], ["/S/row[10]"])
        self.assertEqual([op for op in plan["operations"] if op["command"] == "remove"], [])
        self.assertEqual(plan["expected_final_row_count"], 15)
        rows = [t for _, _, t in plan["row_map"]]
        self.assertEqual(rows, [7, 8, 9, 10, 11])
        filled = {op["path"] for op in plan["operations"]
                  if op.get("props", {}).get("value") is not None}
        self.assertIn("/S/B11", filled)

    # ── Declaration invariants ──
    def test_inplace_multiple_blocks_rejected(self):
        spec = self._spec()
        blk = {"clone_roles": [{"role": "data", "mode": "inplace", "start_row": 7,
                                "capacity": 4, "template_row": 8}],
               "rows": {"source": "source_maoli"}}
        spec["mapping"]["targets"][0]["blocks"] = [blk, blk]
        self._compile_fails_with(spec, "INPLACE_MULTIPLE_BLOCKS")

    def test_inplace_not_last_block_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "mode": "inplace", "start_row": 7,
                              "capacity": 4, "template_row": 8}],
             "rows": {"source": "source_maoli"}},
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"}},
        ]
        self._compile_fails_with(spec, "INPLACE_NOT_LAST_BLOCK")

    def test_inplace_region_out_of_bounds_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["clone_roles"][0]["capacity"] = 99
        self._compile_fails_with(spec, "INPLACE_REGION_OUT_OF_BOUNDS")

    def test_inplace_no_clone_source_rejected(self):
        spec = self._spec()
        del spec["mapping"]["targets"][0]["clone_roles"][0]["template_row"]
        self._compile_fails_with(spec, "INPLACE_NO_CLONE_SOURCE")

    def test_inplace_role_not_last_in_block_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "data", "mode": "inplace", "start_row": 7, "capacity": 4,
             "template_row": 8},
            {"role": "spacer"},
        ]
        self._compile_fails_with(spec, "INPLACE_NOT_LAST_BLOCK")

    # ── Geometry invariants ──
    def test_inplace_region_overlap_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"}, "remove_rows": [8]},
            {"clone_roles": [{"role": "data", "mode": "inplace", "start_row": 7,
                              "capacity": 4, "template_row": 8}],
             "rows": {"source": "source_maoli"}},
        ]
        self._compile_fails_with(spec, "INPLACE_REGION_OVERLAP")

    def test_structural_op_out_of_zone_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"}, "remove_rows": [6]},
            {"clone_roles": [{"role": "data", "mode": "inplace", "start_row": 7,
                              "capacity": 4, "template_row": 8}],
             "rows": {"source": "source_maoli"}},
        ]
        self._compile_fails_with(spec, "STRUCTURAL_OP_OUT_OF_ZONE")

    def test_per_group_total_trigger_minimal_mutation(self):
        """每组合计被拒 fixture 的触发因素最小变异实证 (2026-08-13):
        被拒形态 (聚合列 F 进 nulls D/F) → 只把聚合列移出 nulls 列
        (F→G, nulls 与其余完全不动) → 编译通过 — 触发因素就是
        "聚合列进 nulls", 与"硬编码范围"无关."""
        import _probe_fixtures as pf
        spec = pf._per_group_total_hardcoded_ranges(
            copy.deepcopy(BASE_SPEC), self.wd)
        spec["fingerprints"] = {
            "source_structure": self.wd["manifest"]["fingerprints"]["source_structure"],
            "target_structure": self.wd["manifest"]["fingerprints"]["target_structure"],
        }
        self._compile_fails_with(spec, "DUPLICATE_TARGET_WRITE")
        spec["mapping"]["targets"][0]["formulas"]["aggregates"] = [
            {"col": "G", "rows": "1:2",
             "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
            {"col": "G", "rows": "3:3",
             "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
        ]
        self._compile(spec)  # 不抛 SystemExit = 编译通过

    def test_sets_in_region_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["sets"] = [{"path": "A8", "value": "x"}]
        self._compile_fails_with(spec, "INPLACE_REGION_OVERLAP")

    # ── Placeholder residue double baseline ──
    def test_placeholder_residue_unhandled(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["nulls"] = [{"col": "D", "rows": "all"}]  # F uncovered
        self._compile_fails_with(spec, "PLACEHOLDER_RESIDUE_UNHANDLED")

    def test_placeholder_residue_partial_nulls(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["nulls"] = [
            {"col": "D", "rows": "all"}, {"col": "F", "rows": [1]}]
        self._compile_fails_with(spec, "PLACEHOLDER_RESIDUE_PARTIAL_NULLS")

    # ── group_merges lowering ──
    def test_group_merges_inplace_rebuild(self):
        plan = self._compile(self._spec())
        merges = [op["props"]["merge"] for op in plan["operations"]
                  if op.get("props", {}).get("merge")]
        # 家用×2 group merges A7:A8; 商用 singleton never merges
        self.assertEqual(merges, ["A7:A8"])
        gb = plan["group_boundaries"][0]
        self.assertEqual(gb["col"], "A")
        self.assertEqual(gb["region_start"], 7)
        self.assertEqual(gb["region_end"], 9)
        self.assertEqual(gb["expected_merges"], ["A7:A8"])
        # anchors written, non-anchors explicitly cleared
        writes = {op["path"]: op["props"].get("value") for op in plan["operations"]
                  if op["command"] == "set" and "value" in op["props"]
                  and op["path"].startswith("/S/A")}
        self.assertEqual(writes["/S/A7"], "家用")
        self.assertEqual(writes["/S/A9"], "商用")
        self.assertIsNone(writes["/S/A8"])
        self.assertNotIn("/S/A7", [rb["path"] for rb in plan["readback"]
                                   if rb["kind"] == "empty"])

    def test_group_merges_append_block_general_capability(self):
        """Q8.6: group_merges is a general block capability (append included)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "A"}]
            plan = compile_fill.compile_spec(spec, wd["manifest"], tmp)
            merges = [op["props"]["merge"] for op in plan["operations"]
                      if op.get("props", {}).get("merge")]
            self.assertEqual(merges, ["A7:A8"])  # rows 7-8 家用; row 9 singleton
            gb = plan["group_boundaries"][0]
            self.assertEqual(gb["region_start"], 7)
            self.assertEqual(gb["region_end"], 9)

    def test_group_merge_anchor_uncovered(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "F", "group_by": "A"}]
        spec["mapping"]["targets"][0]["nulls"] = [{"col": "D", "rows": "all"}]
        self._compile_fails_with(spec, "GROUP_MERGE_ANCHOR_UNCOVERED")

    def test_group_by_column_unmapped(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "F"}]
        self._compile_fails_with(spec, "GROUP_BY_COLUMN_UNMAPPED")

    def test_merge_mode_conflict(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["merges"] = [{"col": "A", "rows": "1:{n}"}]
        self._compile_fails_with(spec, "MERGE_MODE_CONFLICT")

    # ── sets ──
    def test_sets_write_clear_and_final_translation(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["sets"] = [
            {"path": "A4", "value": "To Messrs: MXP"},
            {"path": "F13", "value": None},
            {"path": "A14", "value": "* ship to Algeria"},
        ]
        spec["validation"]["key_outputs"] = ["A7", "A14"]
        plan = self._compile(spec)
        ops_by_path = {op["path"]: op for op in plan["operations"]
                       if op["command"] == "set"}
        # ops execute at TEMPLATE coordinates (before the trim shift)
        self.assertEqual(ops_by_path["/S/A4"]["props"]["value"], "To Messrs: MXP")
        self.assertEqual(ops_by_path["/S/F13"]["props"]["value"], None)
        self.assertEqual(ops_by_path["/S/A14"]["props"]["value"], "* ship to Algeria")
        # readback/registry use FINAL coordinates (A14→13, F13→12)
        rb = {rb["path"]: rb for rb in plan["readback"]}
        self.assertIn("/S/A4", rb)
        self.assertIn("/S/A13", rb)
        self.assertEqual(rb["/S/A13"]["expect"], "* ship to Algeria")
        self.assertIn("/S/F12", rb)
        self.assertEqual(rb["/S/F12"]["kind"], "empty")
        # key_outputs translated: A14 → final A13
        self.assertIn({"path": "/S/A13", "kind": "value"}, plan["key_outputs"])
        self.assertEqual([s["path"] for s in plan["sets"]],
                         ["/S/A4", "/S/F12", "/S/A13"])

    def test_sets_out_of_bounds_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["sets"] = [{"path": "A99", "value": "x"}]
        self._compile_fails_with(spec, "SET_OUT_OF_BOUNDS")

    def test_sets_duplicate_write_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["sets"] = [
            {"path": "A4", "value": "x"}, {"path": "A4", "value": "y"}]
        self._compile_fails_with(spec, "DUPLICATE_TARGET_WRITE")

    # ── props whitelist ──
    def test_column_numberformat_props_applied(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E", "props": {"numberformat": "$#,##0.00"}},
        ]
        plan = self._compile(spec)
        e_ops = [op for op in plan["operations"] if op.get("path", "").endswith(("/E7", "/E8", "/E9"))]
        self.assertEqual(len(e_ops), 3)
        self.assertTrue(all(op["props"].get("numberformat") == "$#,##0.00" for op in e_ops))

    def test_props_whitelist_violation_rejected(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["columns"].append(
            {"source": "D", "target": "F", "props": {"font.bold": True}})
        self._compile_fails_with(spec, "PROPS_WHITELIST_VIOLATION")

    def test_set_props_numberformat_legal_with_null(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["sets"] = [
            {"path": "E13", "value": None, "props": {"numberformat": "$#,##0.00"}}]
        plan = self._compile(spec)
        op = next(op for op in plan["operations"] if op["command"] == "set"
                  and op["path"] == "/S/E13")
        self.assertEqual(op["props"]["value"], None)
        self.assertIn("/S/E12", [rb["path"] for rb in plan["readback"] if rb["kind"] == "empty"])

    # ── mapping features added for the MXP scenario ──
    def test_fallback_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            with open(tmp / "source_maoli_flat.csv", "w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerow(["家用", "12K", "Z001", "", "C-1", "1", "2", "3", "101"])
            wd["workdir"] = tmp
            spec = spec_with(wd)
            spec["mapping"]["targets"][0]["columns"] = [
                {"source": "A", "target": "A"},
                {"source": "B", "target": "B"},
                {"source": "C", "target": "C"},
                {"source": "D", "target": "E", "fallback": "B"},
            ]
            plan = compile_fill.compile_spec(spec, wd["manifest"], tmp)
            e_vals = [w["value"] for w in plan["writes"] if w["col"] == "E"]
            self.assertEqual(e_vals, ["12K"])  # D empty → falls back to B

    def test_transforms_chain(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            spec = spec_with(wd)
            spec["mapping"]["transforms"] = [
                {"name": "rename_z", "function": "regex_replace",
                 "pattern": "^Z001$", "replacement": "ZED"},
                {"name": "strip", "function": "strip"},
            ]
            spec["mapping"]["targets"][0]["columns"] = [
                {"source": "A", "target": "A"},
                {"source": "B", "target": "B"},
                {"source": "C", "target": "E", "transforms": ["rename_z", "strip"]},
            ]
            spec["mapping"]["targets"][0]["nulls"] = [{"col": "C", "rows": "all"},
                                                      {"col": "D", "rows": "all"}]
            plan = compile_fill.compile_spec(spec, wd["manifest"], tmp)
            e_vals = [w["value"] for w in plan["writes"] if w["col"] == "E"]
            self.assertEqual(e_vals, ["ZED", "Z002", "Z003"])

    # ── pptx DOM-path sets ──
    def test_pptx_dom_path_sets(self):
        ops = []
        records = []
        compile_fill._emit_sets(
            {"sheet": "slide[1]/table[@id=1]",
             "sets": [{"path": "/slide[1]/table[@id=1]/tr[2]/tc[3]", "value": "x"}]},
            None, None, None, ops, lambda p, k, v: records.append((p, k, v)),
            [], 6, 3)
        self.assertEqual(ops[0]["props"], {"text": "x"})
        self.assertEqual(records, [("/slide[1]/table[@id=1]/tr[2]/tc[3]", "value", "x")])


class AnchorStyleInheritanceTests(unittest.TestCase):
    """Case 010 盲区修复契约 (2026-08-18): inplace 组锚点落在旧合并区非锚点
    格时 (该格无字体样式), 编译器继承占位区内同列既有锚点样式 (font/alignment)。
    优先级: spec 显式 `styles` > 继承值 > STYLE_DEFAULTS (逐键, spec 声明过的
    键不被继承覆盖)。`meta.merge_anchor_styles` 由 prepare 采集
    (flatten_table.collect_anchor_styles), 不入结构指纹。"""

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_styled_anchor_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _spec(self) -> dict:
        return {
            "task": {"intent": "锚点样式继承契约", "selected_mod": "NONE",
                     "selected_mod_revision": None},
            "inputs": {"sources": ["source_quotation.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_quotation.xlsx",
                                          "sheets": ["报价"]}],
                       "target_sheet": "S"},
            "fingerprints": {
                "source_structure": self.wd["manifest"]["fingerprints"]["source_structure"],
                "target_structure": self.wd["manifest"]["fingerprints"]["target_structure"],
            },
            "mapping": {"targets": [{
                "sheet": "S", "base_last_row": 40,
                "clone_roles": [{"role": "data", "mode": "inplace", "start_row": 7,
                                 "capacity": 18, "template_row": 8}],
                "rows": {"source": "source_quotation"},
                "columns": [
                    {"source": "A", "target": "A"},
                    {"source": "B", "target": "B"},
                    {"source": "C", "target": "C"},
                    {"source": "D", "target": "D"},
                    {"source": "E", "target": "E"},
                ],
                "group_merges": [
                    {"col": "A", "group_by": "A", "style": "label"},
                    {"col": "F", "group_by": "A", "label": ""},
                ],
                "sets": [{"path": "A4", "value": "To Messrs: MXP"}],
            }]},
            "validation": {"required_coverage": [], "required_empty": [],
                           "key_outputs": ["A4", "A7"]},
            "decisions": [], "gaps": [], "lineage": [],
        }

    def _merge_ops(self, plan: dict, col: str) -> list:
        """真正的组/范围合并 op (merge 值为范围串) — 排除 merge-clear
        (`{"merge": false}`) 与单格残留清理。"""
        return [o for o in plan["operations"]
                if o.get("command") == "set"
                and isinstance(o.get("props", {}).get("merge"), str)
                and o.get("path", "").startswith(f"/S/{col}")]

    def test_inplace_group_merges_inherit_anchor_font(self):
        plan = compile_fill.compile_spec(self._spec(), self.wd["manifest"], self.tmp)
        a_ops = self._merge_ops(plan, "A")
        self.assertTrue(a_ops, "A 列应有组锚点 merge op")
        for o in a_ops:
            self.assertEqual(o["props"].get("font.name"), "微软雅黑",
                             "A 列锚点应继承模板锚点字体")
            self.assertIs(o["props"].get("font.bold"), True)
            self.assertEqual(o["props"].get("font.size"), "12pt")
        f_ops = self._merge_ops(plan, "F")
        self.assertTrue(f_ops, "F 列应有组锚点 merge op")
        for o in f_ops:
            self.assertNotIn("font.name", o["props"],
                             "F 列锚点无字体样式 → 不得注入字体")
            self.assertEqual(o["props"].get("alignment.horizontal"), "center")

    def test_spec_styles_override_inherited_font(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["styles"] = {"label": {"font.size": "10pt"}}
        plan = compile_fill.compile_spec(spec, self.wd["manifest"], self.tmp)
        for o in self._merge_ops(plan, "A"):
            self.assertEqual(o["props"].get("font.size"), "10pt",
                             "spec 显式 styles 优先于继承值")
            self.assertEqual(o["props"].get("font.name"), "微软雅黑",
                             "继承值补 spec 未声明的键")

    def test_plain_merges_inplace_inherit_too(self):
        spec = self._spec()
        spec["mapping"]["targets"][0]["group_merges"] = [
            {"col": "F", "group_by": "A", "label": ""}]  # F 残留仍由 label 覆盖
        spec["mapping"]["targets"][0]["merges"] = [
            {"col": "A", "rows": "1:{n}", "style": "label"}]
        plan = compile_fill.compile_spec(spec, self.wd["manifest"], self.tmp)
        for o in self._merge_ops(plan, "A"):
            self.assertEqual(o["props"].get("font.name"), "微软雅黑")

    def test_no_anchor_styles_meta_falls_back_to_defaults(self):
        wd = make_preformatted_quotation_workdir(self.tmp)  # 无 merge_anchor_styles
        wd["workdir"] = self.tmp
        spec = self._spec()
        spec["fingerprints"] = {
            "source_structure": wd["manifest"]["fingerprints"]["source_structure"],
            "target_structure": wd["manifest"]["fingerprints"]["target_structure"],
        }
        plan = compile_fill.compile_spec(spec, wd["manifest"], self.tmp)
        for o in self._merge_ops(plan, "A"):
            self.assertNotIn("font.name", o["props"],
                             "无样式元数据 → 不得注入字体, 保持默认 label 样式")
            self.assertEqual(o["props"].get("alignment.horizontal"), "center")

    def test_inplace_anchor_pins_font_scheme_none(self):
        """残余盲区 (2026-08-19): 继承字面字体 (font.name) 落到原无字体格时,
        officecli 会注入 scheme=minor 主题引用使字体渲染回主题 minor 字体 (宋体)。
        编译器必须以 font.scheme=none 钉住字面字体, 防注入覆盖继承字体名;
        同时 plan 声明 strip_scheme_none, 执行期把 val='none' 元素整体剥离
        (WPS 对任何 scheme 元素 — 含 none — 仍按主题字体渲染)。"""
        plan = compile_fill.compile_spec(self._spec(), self.wd["manifest"], self.tmp)
        a_ops = self._merge_ops(plan, "A")
        self.assertTrue(a_ops, "A 列应有组锚点 merge op")
        for o in a_ops:
            self.assertEqual(o["props"].get("font.name"), "微软雅黑")
            self.assertEqual(o["props"].get("font.scheme"), "none",
                             "继承字面字体必须钉 font.scheme=none, 不得漏留 "
                             "officecli 注入的 minor (否则渲染为主题 minor 字体=宋体)")
        self.assertIs(plan.get("strip_scheme_none"), True,
                      "写入 font.scheme=none 的 plan 必须声明执行期 scheme 剥离")

    def test_no_font_pin_no_strip_declaration(self):
        """无字体样式元数据 → 无 pin → plan 不声明 strip_scheme_none
        (执行器不跑 styles 后处理)。"""
        wd = make_preformatted_quotation_workdir(self.tmp)
        wd["workdir"] = self.tmp
        spec = self._spec()
        spec["fingerprints"] = {
            "source_structure": wd["manifest"]["fingerprints"]["source_structure"],
            "target_structure": wd["manifest"]["fingerprints"]["target_structure"],
        }
        plan = compile_fill.compile_spec(spec, wd["manifest"], self.tmp)
        self.assertNotIn("font.name", self._merge_ops(plan, "A")[0]["props"])
        self.assertFalse(plan.get("strip_scheme_none"),
                         "无 font.scheme=none 写入时不得声明 scheme 剥离")

    def test_spec_explicit_font_scheme_not_overridden(self):
        """spec 显式 font.scheme 逐键优先, pin_font_scheme 不得覆盖用户有意使用的
        主题字体 (如 minor)。"""
        spec = self._spec()
        spec["mapping"]["targets"][0]["styles"] = {"label": {"font.scheme": "minor"}}
        plan = compile_fill.compile_spec(spec, self.wd["manifest"], self.tmp)
        for o in self._merge_ops(plan, "A"):
            self.assertEqual(o["props"].get("font.scheme"), "minor",
                             "spec 显式 font.scheme 优先, pin 不覆盖")


class FillSpecContractTests(unittest.TestCase):
    """组合行为契约 (FILLSPEC「组合行为契约」章节) 的编译用例背书.

    只测外部行为: 照契约声明写的 spec, 编译器接受 (编译通过) 或按文档
    声明的错误码拒绝. 每条契约声明都对应一个用例, 防止文档与编译器行为漂移.
    """

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _fail_codes(self, spec) -> list[str]:
        return compile_fail_codes(self.wd, spec)

    def _fail_payload(self, spec) -> dict:
        """Compile-fail payload (parsed stderr JSON); asserts exit 3."""
        from io import StringIO
        buf = StringIO()
        old = sys.stderr
        sys.stderr = buf
        try:
            with self.assertRaises(SystemExit) as ctx:
                compile_fill.compile_spec(spec, self.wd["manifest"],
                                          self.wd["workdir"])
            self.assertEqual(ctx.exception.code, 3)
        finally:
            sys.stderr = old
        return json.loads(buf.getvalue())

    # ── Q1: group_merges × formulas / aggregates ──
    def test_group_merges_plus_aggregate_different_column(self):
        """聚合在独立列: 一等支持 — 编译通过, 锚点聚合公式写在块首行."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "A"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "G", "rows": "1:{n}",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        agg = [op for op in plan["operations"] if op["command"] == "set"
               and "formula" in op.get("props", {}) and "SUM" in op["props"]["formula"]]
        self.assertEqual([op["path"] for op in agg], ["/S/G7"])
        self.assertEqual(agg[0]["props"]["formula"], "SUM(A7:A9)")

    def test_group_merges_and_aggregate_same_column_duplicate_write(self):
        """同列: 组锚点写与聚合锚点写都落在块首行 → DUPLICATE_TARGET_WRITE."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "A"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "A", "rows": "1:{n}",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))
    def test_group_merges_and_per_row_formula_same_column_duplicate_write(self):
        """同列: group lowering 拥有该列每一行 (锚点写/非锚点清空), 与 per_row
        公式冲突 → DUPLICATE_TARGET_WRITE."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "A"}]
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"A": "{r}*2"}}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))

    def test_aggregate_and_per_row_formula_same_column_duplicate_write(self):
        """聚合与逐行公式同列: 首行锚点格双写 → DUPLICATE_TARGET_WRITE."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"E": "A{r}*2"},
            "aggregates": [{"col": "E", "rows": "1:{n}",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))

    def test_nulls_rows_invalid_list_of_ranges_rejected(self):
        """nulls rows 用 ['1:2','3:4'] 混合列表 → NULLS_ROWS_INVALID 结构化
        拒绝 (曾以 Python traceback 崩溃)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["nulls"] = [
            {"col": "D", "rows": ["1:2", "3:4"]}]
        codes = self._fail_codes(spec)
        self.assertIn("NULLS_ROWS_INVALID", codes)

    def test_null_specs_valid_forms_pass(self):
        """nulls rows 合法形态 ('all' / int 列表 / 'a:b' 字符串) 不被
        NULLS_ROWS_INVALID 拒绝 (残留/部分覆盖仍按既有规则报错)."""
        for rows in ("all", [1, 3], "2:4"):
            spec = spec_with(self.wd)
            spec["mapping"]["targets"][0]["nulls"] = [{"col": "D", "rows": rows}]
            codes = self._fail_codes(spec)
            self.assertNotIn("NULLS_ROWS_INVALID", codes, f"rows={rows!r}")

    def test_plain_merges_register_no_readback(self):
        """Q10: plain merges (1:{n}) 只写 merge 属性, 不产生 readback 条目."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["merges"] = [
            {"col": "E", "rows": "1:{n}", "style": "label"}]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        e_readback = [rb for rb in plan["readback"] if rb["path"].startswith("/S/E")]
        self.assertEqual(e_readback, [])
        merges = [op["props"]["merge"] for op in plan["operations"]
                  if op.get("props", {}).get("merge")]
        self.assertEqual(merges, ["E7:E9"])

    def test_title_clone_from_anchor_row_ok(self):
        """Q8: title/header 的 template_row 选锚点行无编译检查 (编译通过),
        data 的 template_row 选锚点行 → CLONE_SOURCE_IS_ANCHOR."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "title", "template_row": 5, "value": "块标题"},  # A5:A6 锚点
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        self.assertEqual(self._fail_codes(spec), [])
        spec2 = spec_with(self.wd)
        spec2["mapping"]["targets"][0]["clone_roles"][2]["template_row"] = 5
        self.assertIn("CLONE_SOURCE_IS_ANCHOR", self._fail_codes(spec2))

    def test_title_value_deferred_to_fill_phase(self):
        """Q9: 标题 value 延迟到 adds 之后写入 (op 顺序恒为 add→...→set)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "title", "template_row": 1, "value": "块标题"},
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        kinds = [op["command"] for op in plan["operations"]]
        first_set = kinds.index("set")
        self.assertNotIn("set", kinds[:first_set])
        self.assertGreater(kinds.count("add"), 0)

    def test_merges_and_aggregates_same_column_ok(self):
        """Q12: merges 1:{n} + aggregates 1:{n} 同列 — 编译通过
        (聚合锚点 = 块首行 = 合并锚点)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["merges"] = [
            {"col": "E", "rows": "1:{n}", "style": "label"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "E", "rows": "1:{n}",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertEqual(self._fail_codes(spec), [])

    def test_multi_range_aggregates_same_column(self):
        """Q12: 同列多条显式范围聚合 — 编译通过, 每条落在各自显式行
        (块内多组小计行写法)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [
                {"col": "E", "rows": "1:2",
                 "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
                {"col": "E", "rows": "3:3",
                 "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
            ]}
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        agg_paths = [op["path"] for op in plan["operations"]
                     if "formula" in op.get("props", {}) and "SUM" in op["props"]["formula"]]
        self.assertEqual(agg_paths, ["/S/E7", "/S/E9"])

    def test_merges_and_per_row_formula_same_column_ok(self):
        """merges (1:{n}) 只写 merge 属性不写值 → 无映射的列上可与 per_row
        公式共存 — 编译通过."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["merges"] = [
            {"col": "E", "rows": "1:{n}", "style": "label"}]
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"E": "A{r}*2"}}
        self.assertEqual(self._fail_codes(spec), [])

    # ── Q13: group_aggregates 一等能力 (组聚合写组锚点行) ──
    def test_group_aggregates_egypt_3_groups_anchors(self):
        """埃及等价: 3 产品组 (家用×2 / 商用×2 / 工程×1) + V 列组聚合 — 编译
        通过, 聚合公式落各组锚点行 ({r1}:{r2} 按组起止展开), readback 自动
        登记各锚点 nonempty."""
        wd = make_egypt_workdir(self.tmp)
        wd["workdir"] = self.tmp
        spec = spec_with(wd)
        spec["fingerprints"] = wd["manifest"]["fingerprints"]
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [
                {"group_by": "A", "col": "V",
                 "formula": "IFERROR(ROUND(SUM(T{r1}:T{r2})/SUM(S{r1}:S{r2}),4),0)",
                 "style": "anchor"}]}
        self.assertEqual(compile_fail_codes(wd, spec), [])
        plan = compile_spec_with(wd, spec)
        agg = [op for op in plan["operations"] if op["command"] == "set"
               and op["path"].startswith("/S/V")
               and "formula" in op.get("props", {})]
        # 数据行 7-11, 组 (1,2)/(3,4)/(5,5) → 锚点 V7/V9/V11
        self.assertEqual([op["path"] for op in agg], ["/S/V7", "/S/V9", "/S/V11"])
        self.assertEqual(agg[0]["props"]["formula"],
                         "IFERROR(ROUND(SUM(T7:T8)/SUM(S7:S8),4),0)")
        self.assertEqual(agg[1]["props"]["formula"],
                         "IFERROR(ROUND(SUM(T9:T10)/SUM(S9:S10),4),0)")
        self.assertEqual(agg[2]["props"]["formula"],
                         "IFERROR(ROUND(SUM(T11:T11)/SUM(S11:S11),4),0)")
        ga_readback = [(rb["path"], rb["kind"]) for rb in plan["readback"]
                       if rb["path"].startswith("/S/V")]
        self.assertEqual(ga_readback,
                         [("/S/V7", "nonempty"), ("/S/V9", "nonempty"),
                          ("/S/V11", "nonempty")])
        # 验收 4 (观测形态): 每个公式的展开范围恒在数据块 7-11 内 —
        # 组范围由数据派生, 越块由编译器以 AGG_RANGE_INVALID 内部守卫拒绝
        for op in agg:
            rows = re.findall(r"(?:T|S)(\d+):(?:T|S)(\d+)", op["props"]["formula"])
            self.assertTrue(rows,
                            f"公式缺 {r'{r1}:{r2}'} 范围: {op['props']['formula']}")
            for lo, hi in rows:
                self.assertTrue(7 <= int(lo) <= int(hi) <= 11,
                                f"组范围越块: {op['props']['formula']}")

    def test_group_aggregates_missing_group_by_rejected(self):
        """group_aggregates 条目缺 group_by → GROUP_BY_COLUMN_UNMAPPED."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [{"col": "G",
                                  "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("GROUP_BY_COLUMN_UNMAPPED", self._fail_codes(spec))

    def test_group_aggregates_malformed_shape_rejected(self):
        """group_aggregates 声明形态非法 (条目非 mapping / per_group 非列表) →
        GROUP_AGGREGATES_INVALID 结构化拒绝 (曾会静默吞掉或 AttributeError
        崩溃)."""
        for formulas in (
                {"group_aggregates": ["SUM(A{r1}:A{r2})"]},
                {"group_aggregates": {"per_group": {"group_by": "A",
                                                    "col": "G",
                                                    "formula": "SUM(A{r1}:A{r2})"},
                                      "whole_run": None}},
                {"group_aggregates": "SUM(A{r1}:A{r2})"}):
            spec = spec_with(self.wd)
            spec["mapping"]["targets"][0]["formulas"] = formulas
            self.assertIn("GROUP_AGGREGATES_INVALID",
                          self._fail_codes(spec),
                          f"formulas={formulas!r}")

    def test_group_aggregates_with_group_merges_same_col_duplicate(self):
        """组聚合与 group_merges 同列: 组锚点双写 (块首行两组锚点重合) →
        DUPLICATE_TARGET_WRITE."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["group_merges"] = [
            {"col": "G", "group_by": "A", "label": "X"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [{"group_by": "A", "col": "G",
                                  "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))

    def test_group_aggregates_with_per_row_same_col_duplicate(self):
        """组聚合与 per_row 公式同列: 锚点格双写 → DUPLICATE_TARGET_WRITE."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"G": "A{r}*2"},
            "group_aggregates": [{"group_by": "A", "col": "G",
                                  "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))

    def test_group_aggregates_col_in_nulls_duplicate(self):
        """组聚合列进 nulls → 锚点先被 nulls 清空再写公式 (特征 "first as
        empty") → DUPLICATE_TARGET_WRITE (组聚合列必须独立于 nulls)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["nulls"] = [
            {"col": "D", "rows": "all"}, {"col": "G", "rows": "all"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [{"group_by": "A", "col": "G",
                                  "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))

    def test_group_aggregates_whole_run_gated_pre_spike(self):
        """whole_run (跨块总计) 落点语义 (末块尾部 vs 独立行) 需 spike 锁定 —
        spike 前声明 (dict 形态) → CAPABILITY_NOT_ROLLED_OUT."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": {
                "per_group": [{"group_by": "A", "col": "G",
                               "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}],
                "whole_run": {"col": "G", "formula": "SUM(A5:A9)",
                              "rows": "last_block_tail"}}}
        self.assertIn("CAPABILITY_NOT_ROLLED_OUT", self._fail_codes(spec))

    def test_group_aggregates_whole_run_list_entry_gated(self):
        """列表条目形态的 whole_run 声明同样被门拒绝 (spec 草案形态)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [
                {"group_by": "A", "col": "G",
                 "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
                {"whole_run": {"col": "G", "formula": "SUM(A5:A9)",
                               "rows": "last_block_tail"}},
            ]}
        self.assertIn("CAPABILITY_NOT_ROLLED_OUT", self._fail_codes(spec))

    def test_group_aggregates_group_by_unmapped_rejected(self):
        """group_by 列无列映射 → GROUP_BY_COLUMN_UNMAPPED."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [{"group_by": "F", "col": "G",
                                  "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        self.assertIn("GROUP_BY_COLUMN_UNMAPPED", self._fail_codes(spec))

    def test_group_aggregates_formula_template_invalid(self):
        """公式模板未知键 → FORMULA_TEMPLATE_INVALID (静态期拒绝)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [{"group_by": "A", "col": "G",
                                  "formula": "SUM(A{r9}:A{r2})", "style": "anchor"}]}
        self.assertIn("FORMULA_TEMPLATE_INVALID", self._fail_codes(spec))

    # ── Q2: 算术派生列 (FLD-006 减法) 标准模式 ──
    def test_derived_column_subtraction_per_row_formula(self):
        """减法派生列标准模式: per_row formula + ROUND(...,2) 写在独立未映射列 —
        编译通过且公式按数据行展开写入 plan."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"G": "IFERROR(ROUND(A{r}-B{r},2),0)"}}
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        formulas = {op["path"]: op["props"]["formula"] for op in plan["operations"]
                    if op["command"] == "set" and op.get("path", "").startswith("/S/G")}
        self.assertEqual(formulas, {
            "/S/G7": "IFERROR(ROUND(A7-B7,2),0)",
            "/S/G8": "IFERROR(ROUND(A8-B8,2),0)",
            "/S/G9": "IFERROR(ROUND(A9-B9,2),0)",
        })

    def test_derived_column_cannot_share_a_mapped_column(self):
        """派生列必须独立: 同列既有列映射又有 per_row 公式 → 双写
        DUPLICATE_TARGET_WRITE."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"source": "B", "target": "E"})
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"E": "A{r}*2"}}
        self.assertIn("DUPLICATE_TARGET_WRITE", self._fail_codes(spec))

    # ── Q3: 映射列 × 合并列落值 (锚点写 / 非锚点抑制+清空) ──
    def test_mapped_group_column_anchor_writes_non_anchor_cleared(self):
        """映射列 + group_merges 同列: 一等支持 — 锚点写物化值, 非锚点抑制
        并显式清空 (EMPTY readback), singleton 组永不合并."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "A"}]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        writes = {op["path"]: op["props"].get("value") for op in plan["operations"]
                  if op["command"] == "set" and "value" in op["props"]
                  and op["path"].startswith("/S/A")}
        self.assertEqual(writes["/S/A7"], "家用")   # 锚点写物化值
        self.assertEqual(writes["/S/A9"], "商用")
        self.assertIsNone(writes["/S/A8"])           # 非锚点抑制 (显式清空)
        empties = [rb["path"] for rb in plan["readback"] if rb["kind"] == "empty"]
        self.assertIn("/S/A8", empties)
        merges = [op["props"]["merge"] for op in plan["operations"]
                  if op.get("props", {}).get("merge")]
        self.assertEqual(merges, ["A7:A8"])

    # ── Q5: 空值 / 0-口径 ──
    def test_empty_source_value_left_blank(self):
        """空源值 → 缺失格留空 (空串写入, readback 期待空串) — 不是 EMPTY 清空."""
        with open(self.tmp / "source_maoli_flat.csv", "w", newline="",
                  encoding="utf-8-sig") as f:
            csv.writer(f).writerow(["家用", "12K", "Z001", "", "C-1", "1", "2", "3", "101"])
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E"}]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        e_op = next(op for op in plan["operations"] if op.get("path") == "/S/E7")
        self.assertEqual(e_op["props"]["value"], "")
        rb = next(rb for rb in plan["readback"] if rb["path"] == "/S/E7")
        self.assertEqual(rb["kind"], "value")
        self.assertEqual(rb["expect"], "")

    def test_zero_source_value_is_not_missing(self):
        """数值 0 不是缺失 — 0-口径原样写入 '0'."""
        with open(self.tmp / "source_maoli_flat.csv", "w", newline="",
                  encoding="utf-8-sig") as f:
            csv.writer(f).writerow(["家用", "12K", "Z001", "F-1", "C-1", "0", "2", "3", "101"])
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "F", "target": "G"}]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        g_op = next(op for op in plan["operations"] if op.get("path") == "/S/G7")
        self.assertEqual(g_op["props"]["value"], "0")

    def test_formula_referenced_column_missing_writes_zero(self):
        """0-口径二分 (issue 01): **入公式链列缺失 → 数值 0** (常量
        value: "0"), readback 期待数值 "0" — 空串会令公式求值链按非空文本
        判错 → #VALUE! → IFERROR 兜底 0 (Case 07 §8)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"target": "F", "value": "0"},      # 被 H 公式引用
        ]
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"H": "IFERROR(F{r}-A{r},0)"}}
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        f7 = next(op for op in plan["operations"] if op.get("path") == "/S/F7")
        self.assertEqual(f7["props"]["value"], "0")
        rb = next(rb for rb in plan["readback"] if rb["path"] == "/S/F7")
        self.assertEqual(rb["kind"], "value")
        self.assertEqual(rb["expect"], "0")

    def test_standalone_display_column_blank_still_accepted(self):
        """0-口径二分 (issue 01): **独立展示、不入任何公式链的字段缺失 →
        才可留空** (空串 readback) — 与「入公式链列 → 数值 0」二分判据是
        公式是否引用, 独立列留空仍被接受."""
        with open(self.tmp / "source_maoli_flat.csv", "w", newline="",
                  encoding="utf-8-sig") as f:
            csv.writer(f).writerow(["家用", "12K", "Z001", "", "C-1", "1", "2", "3", "101"])
            csv.writer(f).writerow(["家用", "18K", "Z002", "", "C-2", "4", "5", "6", "102"])
            csv.writer(f).writerow(["商用", "24K", "Z003", "", "C-3", "7", "8", "9", "103"])
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E"},     # 独立展示列 (源 D 空)
            {"target": "F", "value": "0"},      # 入公式链列 (H 公式引用)
        ]
        spec["mapping"]["targets"][0]["nulls"] = [{"col": "D", "rows": "all"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"H": "IFERROR(F{r}-A{r},0)"}}
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        e7 = next(op for op in plan["operations"] if op.get("path") == "/S/E7")
        self.assertEqual(e7["props"]["value"], "")
        rb = {r["path"]: r for r in plan["readback"]}
        self.assertEqual(rb["/S/E7"]["kind"], "value")
        self.assertEqual(rb["/S/E7"]["expect"], "")
        self.assertEqual(rb["/S/F7"]["expect"], "0")

    # ── Q6: lookup missing 语义 ──
    def test_lookup_missing_empty_leaves_blank_cell(self):
        """missing: empty → 缺失 key 的格留空 (空串 readback); 命中行正常取值."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "F", "lookup": {"name": "fields",
                                       "field": "compressor", "missing": "empty"}})
        spec["mapping"]["targets"][0]["lookups"] = [
            {"name": "fields", "from": "inheritance.json", "key_column": "C",
             "fields": ["compressor"], "missing": "empty"}]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        rb = {rb["path"]: rb for rb in plan["readback"]}
        self.assertEqual(rb["/S/F7"]["expect"], "C-1")   # Z001 命中
        self.assertEqual(rb["/S/F9"]["expect"], "")       # Z003 缺失 → 留空

    def test_lookup_field_missing_schema_absent(self):
        """字段不在索引 schema → LOOKUP_FIELD_MISSING (missing 策略不豁免)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "F", "lookup": {"name": "fields",
                                       "field": "voltage", "missing": "empty"}})
        spec["mapping"]["targets"][0]["lookups"] = [
            {"name": "fields", "from": "inheritance.json", "key_column": "C",
             "fields": ["voltage"], "missing": "empty"}]
        self.assertIn("LOOKUP_FIELD_MISSING", self._fail_codes(spec))

    # ── Q7: precision: keep vs round4 ──
    def test_precision_keep_compiles(self):
        """precision: keep 显式接受长精度 — 编译通过且无警告 (round4 仍为推荐序,
        见 doc-coverage 守卫 test_fillspec_precision_recommendation_order)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "E", "value": "168.715100569657", "precision": "keep"})
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        self.assertEqual(plan["warnings"], [])

    # ── Q8: CLONE_SOURCE_IS_ANCHOR 作用域 (只查 data role) ──
    def test_anchor_check_applies_to_data_role_only(self):
        """fixture 锚点 = A5. title/header 克隆源选锚点行无编译检查 (边界已文档化,
        FILLSPEC Q8 — 公式残留风险标注); data role 克隆源选锚点行 → 拒绝."""
        for role in ("title", "header"):
            spec = spec_with(self.wd)
            spec["mapping"]["targets"][0]["clone_roles"] = [
                {"role": role, "template_row": 5, "value": "X"},
                {"role": "data", "template_row": 7},
            ]
            self.assertEqual(self._fail_codes(spec), [],
                             f"{role} 克隆源=锚点行应无编译检查 (文档化边界)")
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "title", "template_row": 2, "value": "X"},
            {"role": "data", "template_row": 5},
        ]
        self.assertIn("CLONE_SOURCE_IS_ANCHOR", self._fail_codes(spec))

    # ── Q9: title/header value 延迟到 fills 阶段 (adds 之后) ──
    def test_title_value_written_after_all_adds(self):
        """标题值 op 恒在所有 add/remove 之后 (deferred_values — 防 duplicate_row:
        add 之间穿插 cell 写入破坏 officecli 行簿记)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "spacer"},
            {"role": "title", "template_row": 1, "value": "HDR"},
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        spec["validation"]["key_outputs"] = ["A8"]  # spacer 后移: 数据行 8-10
        plan = compile_spec_with(self.wd, spec)
        ops = plan["operations"]
        last_add = max(i for i, o in enumerate(ops) if o["command"] == "add")
        title_idx = next(i for i, o in enumerate(ops)
                         if o["command"] == "set"
                         and o.get("props", {}).get("value") == "HDR")
        self.assertGreater(title_idx, last_add)
        # add 之间不得穿插任何 cell 写入
        add_positions = [i for i, o in enumerate(ops) if o["command"] == "add"]
        for a, b in zip(add_positions, add_positions[1:]):
            self.assertTrue(all(ops[i]["command"] == "add"
                                for i in range(a + 1, b)),
                            "add 之间穿插非 add op")

    # ── Q10: readback 断言种类 (register 语义) ──
    def test_readback_kinds_by_write_source(self):
        """value→值断言; nulls/清空→EMPTY; 公式/合并锚点→nonempty; 一格一 kind."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"F": "A{r}*2"}}
        plan = compile_spec_with(self.wd, spec)
        by_kind = {}
        for rb in plan["readback"]:
            by_kind.setdefault(rb["kind"], []).append(rb["path"])
        self.assertIn("/S/D7", by_kind["empty"])     # nulls (BASE_SPEC D 列) → EMPTY
        self.assertIn("/S/A7", by_kind["value"])     # 列映射 → 值断言
        self.assertIn("/S/F7", by_kind["nonempty"])  # per_row 公式 → nonempty
        self.assertEqual(len({rb["path"] for rb in plan["readback"]}),
                         len(plan["readback"]), "一格一断言")

    # ── 布局决策树: append 块 remove_rows 越界 → REMOVE_TARGETS_APPEND_ZONE ──
    def test_remove_rows_beyond_base_append_zone_rejected(self):
        """埃及等价 (probe 2026-08-13): append 克隆 + remove_rows > base_last_row
        → 先行的 add 全部插在 base 之下推移行号, remove 用裸模板坐标命中刚插入的
        新数据行 (自毁 plan, 行数断言恒等抓不住) → 拒绝, 缺陷携带行号、块标签与
        拦截理由, 且不无条件指向 inplace."""
        spec = spec_with(self.wd)  # base_last_row=4, 数据行 5-7
        spec["mapping"]["targets"][0]["remove_rows"] = [5, 6, 7]
        payload = self._fail_payload(spec)
        self.assertEqual(payload["code"], "STATIC_VALIDATION_FAILED")
        defects = [d for d in payload["defects"]
                   if d["code"] == "REMOVE_TARGETS_APPEND_ZONE"]
        self.assertEqual([d["row"] for d in defects], [5, 6, 7])
        for d in defects:
            self.assertEqual(d["block"], "block[0]")
            self.assertIn("base_last_row 4", d["message"])
            self.assertIn("自毁", d["message"])
            ca = d["corrective_action"]
            self.assertIn("append-only", ca)
            self.assertIn("占位行自然下沉", ca)
            self.assertIn("inplace", ca)
            self.assertIn("样式", ca)
            self.assertLess(ca.index("append-only"), ca.index("inplace"),
                            "append-only 是首选, inplace 只能是条件选项")

    def test_remove_rows_within_base_classic_shrink_compiles(self):
        """经典场景 (源行数 < 模板行数): remove_rows ≤ base_last_row 在 add 区
        之外不被推移 → 编译通过, plan 与无 remove_rows 的基线相比仅多该 remove op
        (既有运行不受影响)."""
        baseline = compile_spec_with(self.wd, spec_with(self.wd))
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["remove_rows"] = [3]  # ≤ base 4
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        diff = [op for op in plan["operations"] if op not in baseline["operations"]]
        self.assertEqual(len(diff), 1)
        self.assertEqual(diff[0]["command"], "remove")
        self.assertEqual(diff[0]["path"], "/S/row[3]")
        self.assertEqual(plan["structural_deltas"]["removes"], 1)
        # 执行顺序: remove 恒在全部 add 之后
        ops = plan["operations"]
        last_add = max(i for i, o in enumerate(ops) if o["command"] == "add")
        remove_idx = next(i for i, o in enumerate(ops) if o["command"] == "remove")
        self.assertGreater(remove_idx, last_add)

    def test_remove_zone_multi_block_any_out_of_bounds_rejected(self):
        """blocks[] 多块: 任一非 inplace 块 remove_rows 越界 → 拒绝, 缺陷携带
        越界块的标签."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli",
                      "selectors": [{"column": "A", "pattern": "家用*"}]},
             "remove_rows": [2]},   # ≤ base 4 — 合法
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli",
                      "selectors": [{"column": "A", "pattern": "商用*"}]},
             "remove_rows": [7]},   # > base 4 — 自毁
        ]
        payload = self._fail_payload(spec)
        bad = [d for d in payload["defects"]
               if d["code"] == "REMOVE_TARGETS_APPEND_ZONE"]
        self.assertEqual(len(bad), 1)
        self.assertEqual(bad[0]["block"], "block[1]")
        self.assertEqual(bad[0]["row"], 7)

    def test_remove_zone_multi_block_all_legal_compiles(self):
        """blocks[] 多块: 全部 ≤ base_last_row → 编译通过, 每块 remove op 依块序
        自底向上生成."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli",
                      "selectors": [{"column": "A", "pattern": "家用*"}]},
             "remove_rows": [2]},
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli",
                      "selectors": [{"column": "A", "pattern": "商用*"}]},
             "remove_rows": [3]},
        ]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        removes = [op["path"] for op in plan["operations"]
                   if op["command"] == "remove"]
        self.assertEqual(removes, ["/S/row[2]", "/S/row[3]"])

    def test_remove_zone_skips_inplace_block(self):
        """inplace 块消费编译器推导的 Trim, 不消费 remove_rows — 不在检查范围
        (remove_rows 越界值也不触发 REMOVE_TARGETS_APPEND_ZONE)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_inplace_workdir(tmp)
            wd["workdir"] = tmp
            spec = copy.deepcopy(INPLACE_BASE_SPEC)
            spec["fingerprints"] = wd["manifest"]["fingerprints"]
            spec["mapping"]["targets"][0]["remove_rows"] = [15]  # > base 14
            codes = compile_fail_codes(wd, spec)
            self.assertNotIn("REMOVE_TARGETS_APPEND_ZONE", codes)
            compile_fill.compile_spec(spec, wd["manifest"], tmp)  # 编译通过

    # ── ID-1/ID-2: block 顶层键静态校验 + 块级 formulas 取代契约 ──

    def test_block_top_level_aggregates_rejected(self):
        """block 顶层 `aggregates:` 曾静默丢弃 (resolve_blocks 透传 +
        _emit_block_ops 只读 formulas) — 现编译期 BLOCK_KEY_STRUCTURE_INVALID
        拒绝 (exit 3), corrective_action 点名应写 formulas: {aggregates: [...]}."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"},
             "aggregates": [{"col": "G", "rows": "1:{n}",
                             "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]},
        ]
        payload = self._fail_payload(spec)
        d = next(x for x in payload["defects"]
                 if x["code"] == "BLOCK_KEY_STRUCTURE_INVALID")
        self.assertEqual(d["block"], "block[0]")
        self.assertEqual(d["key"], "aggregates")
        self.assertIn("formulas.aggregates", d["corrective_action"])
        self.assertIn("静默忽略", d["message"])

    def test_block_top_level_per_row_and_group_aggregates_rejected(self):
        """block 顶层其它错位键 (per_row / group_aggregates) 同码拒绝, 各自
        点名正确嵌套 (黑名单内明确点名)."""
        for key, nesting in (("per_row", "formulas.per_row"),
                             ("group_aggregates", "formulas.group_aggregates")):
            spec = spec_with(self.wd)
            spec["mapping"]["targets"][0]["blocks"] = [
                {"clone_roles": [{"role": "data", "template_row": 3}],
                 "rows": {"source": "source_maoli"},
                 key: {"G": "A{r}*2"}},
            ]
            payload = self._fail_payload(spec)
            d = next(x for x in payload["defects"]
                     if x["code"] == "BLOCK_KEY_STRUCTURE_INVALID")
            self.assertEqual(d["key"], key, f"block 顶层 {key} 应被点名")
            self.assertIn(nesting, d["corrective_action"], f"{key} → {nesting}")

    def test_block_top_level_unknown_key_rejected(self):
        """block 顶层未知键 (typo 如单数 `formula`) → 同一缺陷码
        BLOCK_KEY_STRUCTURE_INVALID, corrective_action 点名合法顶层键."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"},
             "formula": {"per_row": {"G": "A{r}*2"}}},
        ]
        payload = self._fail_payload(spec)
        d = next(x for x in payload["defects"]
                 if x["code"] == "BLOCK_KEY_STRUCTURE_INVALID")
        self.assertEqual(d["key"], "formula")
        self.assertIn("clone_roles", d["corrective_action"])
        self.assertIn("rows", d["corrective_action"])

    def test_block_formulas_replaces_target_per_row(self):
        """块级声明 `formulas` 即整体取代 target 级 per_row (不合并): 块声明
        aggregates 而不含 per_row → 块内无任何 target 级 per_row 公式."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"G": "A{r}*2"}}
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"},
             "formulas": {"aggregates": [{"col": "H", "rows": "1:{n}",
                                          "formula": "SUM(A{r1}:A{r2})",
                                          "style": "anchor"}]}},
        ]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        g_ops = [op for op in plan["operations"]
                 if op["command"] == "set" and op["path"].startswith("/S/G")]
        self.assertEqual(g_ops, [],
                         "块级 formulas 取代 target 级 per_row: 块内不得出现 "
                         "target 级 G 公式 (不合并)")
        agg = [op for op in plan["operations"]
               if "formula" in op.get("props", {})
               and "SUM" in op["props"]["formula"]]
        self.assertEqual([op["path"] for op in agg], ["/S/H5"],
                         "块声明自身的 aggregates 仍生效")

    def test_block_no_formulas_inherits_target_per_row(self):
        """块级不声明 `formulas` → 缺省继承 target 级 per_row: 块数据行出现
        target 级 per_row 公式 (只写差异 = 继承)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {"per_row": {"G": "A{r}*2"}}
        spec["mapping"]["targets"][0]["blocks"] = [
            {"clone_roles": [{"role": "data", "template_row": 3}],
             "rows": {"source": "source_maoli"}},
        ]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        formulas = {op["path"]: op["props"]["formula"] for op in plan["operations"]
                    if op["command"] == "set" and op["path"].startswith("/S/G")
                    and "formula" in op.get("props", {})}
        self.assertEqual(formulas, {
            "/S/G5": "A5*2", "/S/G6": "A6*2", "/S/G7": "A7*2"},
            "块级不声明 formulas → 继承 target 级 per_row (数据行 5-7)")


class HeaderRowGuardContractTests(unittest.TestCase):
    """issue 02 / Case 08 U1 — HEADER_ROW_CONSIDERED_DATA 表头行守卫契约:

    展平 CSV 首行（表头文本行）是候选数据行; rows 无 selector 或 selector
    未排除首行且首行是表头文本行 → 编译警告 HEADER_ROW_CONSIDERED_DATA
    (失败语义不变, 记 warnings); 加 pattern/not_pattern 排除表头行 → 无警告。
    探测第一个触发条件 (机械事实以 probe 固化): 首行 = 全文本标签 (>= 2 个
    非空 cell 且无数值 cell) 才判定为表头文本行 — 真实数据行几乎必带数值, 因此
    普通 fill (首行是数据行) 不误报。
    """

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_header_row_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _warn_codes(self, spec) -> list[str]:
        plan = compile_fill.compile_spec(spec, self.wd["manifest"], self.wd["workdir"])
        return [w["code"] for w in plan["warnings"]]

    def test_header_row_considered_data_without_selector(self):
        """rows 无 selector + 源 flat 首行为表头文本行 → HEADER_ROW_CONSIDERED_DATA
        警告 (compile 仍 accept, 失败语义不变)."""
        spec = spec_with(self.wd)
        self.assertEqual(self._warn_codes(spec), ["HEADER_ROW_CONSIDERED_DATA"])
        codes = compile_fail_codes(self.wd, spec)
        self.assertEqual(codes, [], "警告不改变失败语义 (不 exit 3)")

    def test_header_row_guard_message_and_corrective_action(self):
        """警告携带 source 与 corrective_action: 「加 pattern/not_pattern 排除
        表头行」并给出 `column A pattern 业务类别*` 示例词."""
        spec = spec_with(self.wd)
        plan = compile_fill.compile_spec(spec, self.wd["manifest"], self.wd["workdir"])
        guard = next(w for w in plan["warnings"]
                     if w["code"] == "HEADER_ROW_CONSIDERED_DATA")
        self.assertEqual(guard["source"], "source_maoli")
        for word in ("pattern", "not_pattern", "排除表头行", "候选数据行"):
            self.assertIn(word, guard["message"] + guard["corrective_action"],
                          f"守卫缺词 {word!r}")
        self.assertIn("业务类别", guard["corrective_action"])

    def test_header_row_excluded_by_selector_no_warning(self):
        """加 pattern 排除表头行 (行 A=类别 不匹配 家用*) → 无
        HEADER_ROW_CONSIDERED_DATA 警告."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["rows"]["selectors"] = [
            {"column": "A", "pattern": "家用*"}]
        self.assertEqual(self._warn_codes(spec), [],
                         "selector 排除表头行后不应再警告")

    def test_data_first_source_no_warning(self):
        """首行是普通数据行 (含数值 cell) → 不误报 (既有 fixture 全绿的前提)."""
        with tempfile.TemporaryDirectory() as td:
            wd = make_workdir(Path(td))
            wd["workdir"] = Path(td)
            spec = spec_with(wd)
            plan = compile_fill.compile_spec(spec, wd["manifest"], Path(td))
            codes = [w["code"] for w in plan["warnings"]]
            self.assertNotIn("HEADER_ROW_CONSIDERED_DATA", codes)


class MergeModeConflictContractTests(unittest.TestCase):
    """issue 03 / Case 07 改进 1/3 (E3 删除) — MERGE_MODE_CONFLICT corrective_action
    指向「正确组合」而非只写「每列一种合并模式」:

    - 该列承载聚合 (聚合锚点 / 合并覆盖残留) → 删除其 group_merges 条目, 改用
      同范围 merges + aggregates 对 (聚合锚点=合并锚点);
    - 普通标签列 → 每列保留一种合并模式 (group_merges 或 merges 二选一)。

    不绑定 pattern 名、不绑定任务 (通用层指引)。"""

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _conflict_spec(self, aggregate: bool) -> dict:
        """复用 PROBE_CASES 的构建器 (单一真实来源, probe-first): 聚合列 /
        普通标签列 两个 MERGE_MODE_CONFLICT 触发形态."""
        import _probe_fixtures as pf
        spec = spec_with(self.wd)
        builder = (pf._merge_conflict_aggregate_col if aggregate
                   else pf._merge_conflict_label_col)
        builder(spec, self.wd)
        return spec

    def _conflict_defect(self, spec) -> dict:
        r = compile_fill.probe_spec(spec, self.wd["manifest"], self.wd["workdir"])
        self.assertFalse(r["accepted"], "同列混用 merges + group_merges 应被拒绝")
        self.assertEqual(r["exit_code"], 3,
                         "MERGE_MODE_CONFLICT 是编译拒绝 (exit 3)")
        codes = [d.get("code") for d in r["defects"]]
        self.assertIn("MERGE_MODE_CONFLICT", codes)
        return next(d for d in r["defects"] if d.get("code") == "MERGE_MODE_CONFLICT")

    def test_aggregate_column_in_group_merges_conflict(self):
        """聚合列误进 group_merges → exit 3 + code=MERGE_MODE_CONFLICT."""
        d = self._conflict_defect(self._conflict_spec(aggregate=True))
        self.assertEqual(d["col"], "G")

    def test_aggregate_column_corrective_action_wording(self):
        """聚合列冲突 corrective_action 含「merges+aggregates」「聚合锚点=合并锚点」
        「group_merges」措辞."""
        d = self._conflict_defect(self._conflict_spec(aggregate=True))
        text = d["message"] + d["corrective_action"]
        for word in ("group_merges", "merges + aggregates", "聚合锚点=合并锚点",
                     "聚合", "删除其 group_merges 条目"):
            self.assertIn(word, text, f"聚合列冲突指引缺词 {word!r}")

    def test_label_column_corrective_action_single_mode(self):
        """普通标签列冲突 corrective_action 指向「保留一种合并模式」, 不误导向
        聚合组合."""
        d = self._conflict_defect(self._conflict_spec(aggregate=False))
        for word in ("普通标签列", "保留一种"):
            self.assertIn(word, d["corrective_action"],
                          f"标签列冲突指引缺词 {word!r}")
        self.assertNotIn("merges + aggregates", d["corrective_action"],
                         "普通标签列不应被导向聚合组合")

    def test_corrective_action_not_bound_to_pattern_or_task(self):
        """指引是通用层: 不引用 pattern 名、不绑定任务 (用户裁决)."""
        for aggregate in (True, False):
            d = self._conflict_defect(self._conflict_spec(aggregate))
            for banned in ("multiproduct", "single_quotation", "MXP", "Case 08"):
                self.assertNotIn(banned, d["message"] + d["corrective_action"],
                                 f"指引泄漏 pattern/任务词 {banned!r}")


class ExecutorRenderDefaultContractTests(unittest.TestCase):
    """issue 03 / Case 07 改进 4 — execute_batch.py `--render` 默认 none → html:

    - 省略 `--render` 时按 html 执行 (纯文本模型结构渲染检查, 不宣称视觉验证);
    - 产物 (render_qa) 计入 execute 返回的机器证据 (receipt.render_qa);
    - 显式 `--render none` 仍可用 (跳过).
    """

    def test_render_mode_default_is_html(self):
        self.assertEqual(execute_batch.RENDER_MODE_DEFAULT, "html",
                         "execute 省略 --render 时应默认 html (Case 07 改进 4)")

    def test_render_qa_none_skips_in_machine_evidence_shape(self):
        """--render none 的 render_qa 结果仍是机器证据一部分 (mode/status), 但跳过
        渲染产物."""
        r = execute_batch.render_qa(Path("."), Path("x.xlsx"), "", "none")
        self.assertEqual(r["mode"], "none")
        self.assertEqual(r["status"], "skipped")
        self.assertIn("status", r)

    def test_render_arg_parses_html_default_without_flag(self):
        """真实 CLI 契约: 省略 --render 时 argparse 默认 = html (进出接口契约: 文档
        与执行器默认值一致, 无法漂移)."""
        ns = execute_batch.build_arg_parser().parse_args(
            ["--plan", "p.json", "--template", "t.xlsx", "--workdir", "."])
        self.assertEqual(ns.render, "html")
        ns2 = execute_batch.build_arg_parser().parse_args(
            ["--plan", "p.json", "--template", "t.xlsx", "--workdir", ".",
             "--render", "none"])
        self.assertEqual(ns2.render, "none", "显式 --render none 仍可用")


class ExecutionOrderContractTests(unittest.TestCase):
    """「执行顺序保证」契约 (FILLSPEC 章节) 的编译用例背书 — 文档声称与编译器
    行为 lockstep: op 全局顺序不变量 (E1) / remove 目标身份 (E2) / 自底向上
    (E3) / 坐标翻译边界 (E4) / 机械事实派生 (mechanical_facts + mapping.md)."""

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _compile(self, spec):
        return compile_fill.compile_spec(spec, self.wd["manifest"], self.wd["workdir"])

    def _remove_rows_of(self, plan) -> list[int]:
        return [int(re.search(r"/row\[(\d+)\]$", op["path"]).group(1))
                for op in plan["operations"] if op["command"] == "remove"]

    # ── E1: op 全局顺序不变量 (clear → add → remove → merge → fill) ──
    def test_global_op_order_invariant_append_only(self):
        """E1 (append-only 形态): 全局序列 = clear → add → remove → merge → fill —
        add 序列连续 (之间零 cell 写入 — 值写入穿插 add 破坏行簿记 →
        duplicate_row); 全部 remove 在最后一个 add 之后; merge 属性写入与值写入
        全部在全部结构操作之后."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "title", "template_row": 1, "value": "HDR"},
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        spec["mapping"]["targets"][0]["remove_rows"] = [3]
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"E": "A{r}*2"},
            "aggregates": [{"col": "F", "rows": "1:{n}",
                            "formula": "SUM(E{r1}:E{r2})", "style": "anchor"}]}
        spec["mapping"]["targets"][0]["merges"] = [
            {"col": "D", "rows": "1:{n}", "style": "label"}]
        plan = self._compile(spec)
        ops = plan["operations"]
        add_idx = [i for i, o in enumerate(ops) if o["command"] == "add"]
        remove_idx = [i for i, o in enumerate(ops) if o["command"] == "remove"]
        set_idx = [i for i, o in enumerate(ops) if o["command"] == "set"]
        # 任意两个 add 之间零 cell 写入
        for a, b in zip(add_idx, add_idx[1:]):
            self.assertTrue(all(ops[i]["command"] == "add" for i in range(a + 1, b)),
                            "add 之间穿插非 add op — 破坏行簿记 (duplicate_row)")
        # 全部 remove 在最后一个 add 之后
        self.assertTrue(remove_idx, "本 spec 应有 remove")
        self.assertGreater(max(remove_idx), max(add_idx))
        # merge 属性写入与值写入全部在最后一个结构操作之后 (append-only 成立)
        merge_set_idx = [i for i in set_idx if "merge" in ops[i].get("props", {})]
        value_set_idx = [i for i in set_idx
                         if "value" in ops[i].get("props", {})
                         or "formula" in ops[i].get("props", {})]
        last_struct = max(add_idx + remove_idx)
        self.assertTrue(merge_set_idx and value_set_idx, "本 spec 应有 merge 与值写入")
        self.assertTrue(all(i > last_struct for i in merge_set_idx + value_set_idx),
                        "append-only: merge/值写入必须全部在 add/remove 之后")

    def test_inplace_hybrid_op_order(self):
        """E1 (inplace 混合形态): 全局序列 = append 块全部操作 → sets → 终末
        inplace 结构操作 (overflow add / trim remove) → inplace 值操作. sets 是
        值写却先于 inplace 结构操作 — 位置由 Excel 行移位搬移 (契约精确化,
        FILLSPEC「执行顺序保证」E1 与「v2.5: Row Layout Mode」同源)."""
        # trim 形态 (N=3 < capacity 4): sets → trim removes → 值写, 无 add
        wd2 = make_inplace_workdir(self.tmp, n_source_rows=3)
        wd2["workdir"] = self.tmp
        spec2 = copy.deepcopy(INPLACE_BASE_SPEC)
        spec2["fingerprints"] = wd2["manifest"]["fingerprints"]
        spec2["mapping"]["targets"][0]["sets"] = [
            {"path": "A4", "value": "To Messrs: MXP"}]
        ops2 = compile_fill.compile_spec(spec2, wd2["manifest"], self.tmp)["operations"]
        set_a4 = next(i for i, o in enumerate(ops2) if o.get("path") == "/S/A4")
        remove_idx = [i for i, o in enumerate(ops2) if o["command"] == "remove"]
        self.assertEqual([o["command"] for o in ops2[:set_a4]], [],
                         "sets 先于 inplace 结构操作 (trim 前执行, 由移位搬移)")
        self.assertGreater(min(remove_idx), set_a4, "trim removes 在 sets 之后")
        self.assertEqual([o["command"] for o in ops2[max(remove_idx) + 1:]], ["set"] * len(ops2[max(remove_idx) + 1:]),
                         "trim removes 之后只剩 inplace 值写, 无结构 op")
        # overflow 形态 (N=5 > capacity 4): sets → overflow add → 值写, 无 remove
        wd3 = make_inplace_workdir(self.tmp, n_source_rows=5)
        wd3["workdir"] = self.tmp
        spec3 = copy.deepcopy(INPLACE_BASE_SPEC)
        spec3["fingerprints"] = wd3["manifest"]["fingerprints"]
        spec3["mapping"]["targets"][0]["sets"] = [
            {"path": "A4", "value": "To Messrs: MXP"}]
        ops3 = compile_fill.compile_spec(spec3, wd3["manifest"], self.tmp)["operations"]
        set_a4 = next(i for i, o in enumerate(ops3) if o.get("path") == "/S/A4")
        add_idx = [i for i, o in enumerate(ops3) if o["command"] == "add"]
        self.assertEqual(len(add_idx), 1, "N=5 → 1 overflow 克隆")
        self.assertGreater(min(add_idx), set_a4, "overflow add 在 sets 之后")
        self.assertEqual([o["command"] for o in ops3[max(add_idx) + 1:]],
                         ["set"] * len(ops3[max(add_idx) + 1:]),
                         "overflow add 之后只剩 inplace 值写, 无结构 op")

    # ── E2: add 之后 remove 的目标身份 ──
    def test_remove_targets_template_coordinates_not_shifted(self):
        """E2: remove_rows 是模板坐标, 不随 add 推移 — remove op 的 path 精确
        等于 spec 声明的行号; 被删行全部 ≤ base_last_row; add 插入行全部 >
        base_last_row (两者无交集, remove 永不命中刚插入的新数据行)."""
        spec = spec_with(self.wd)  # base_last_row=4, 数据行 5-7
        spec["mapping"]["targets"][0]["remove_rows"] = [2, 3]
        plan = self._compile(spec)
        base = spec["mapping"]["targets"][0]["base_last_row"]
        removes = self._remove_rows_of(plan)
        self.assertEqual(sorted(removes), [2, 3],
                         "remove op 行号必须 == spec 声明的模板坐标")
        self.assertTrue(all(r <= base for r in removes))
        insert_rows = []
        for op in plan["operations"]:
            if op["command"] == "add" and "after" in op:
                anchor = int(re.search(r"/row\[(\d+)\]$", op["after"]).group(1))
                insert_rows.append(anchor + 1)
        self.assertTrue(all(r > base for r in insert_rows),
                        "add 全部插在 base_last_row 之下 (append 区)")
        self.assertEqual(set(removes) & set(insert_rows), set(),
                         "remove 目标与 add 插入行无交集")
        mf = plan["mechanical_facts"]
        self.assertTrue(mf["removes"]["all_within_base"])
        self.assertIn("不被 add 推移", mf["removes"]["conclusion"])
        self.assertTrue(all(r > base for r in mf["add_zone"]["append_insert_rows"]),
                        "机械事实: append 插入行全部在 base_last_row 之下")
        self.assertEqual(mf["add_zone"]["overflow_insert_rows"], [],
                         "append-only plan 无 overflow 克隆")

    # ── E3: remove 底上序 ──
    def test_removes_bottom_up_order(self):
        """E3: remove 自底向上 — 每个 append 块内 remove_rows 按行号降序生成
        (先删上行会让按位置解析的执行器下行坐标失效); inplace Trim 同理由
        自底向上 (尾部裁剪)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["remove_rows"] = [1, 2, 3, 4]
        plan = self._compile(spec)
        self.assertEqual(self._remove_rows_of(plan), [4, 3, 2, 1])
        self.assertTrue(plan["mechanical_facts"]["removes"]["bottom_up"])
        # inplace Trim: region 7-10, N=2 → 尾部裁剪 10, 9 (自底向上)
        wd2 = make_inplace_workdir(self.tmp, n_source_rows=2)
        wd2["workdir"] = self.tmp
        spec2 = copy.deepcopy(INPLACE_BASE_SPEC)
        spec2["fingerprints"] = wd2["manifest"]["fingerprints"]
        plan2 = compile_fill.compile_spec(spec2, wd2["manifest"], self.tmp)
        self.assertEqual(self._remove_rows_of(plan2), [10, 9])
        self.assertEqual(plan2["mechanical_facts"]["trim"]["rows"], [10, 9])
        self.assertIn("自底向上", plan2["mechanical_facts"]["trim"]["conclusion"])

    # ── E4: 坐标翻译边界 ──
    def test_ops_template_readback_final_coordinates(self):
        """E4: ops 用模板坐标, readback 用最终坐标. append-only 无移位 →
        readback 坐标 == op 坐标; inplace trim → set op 保持模板坐标, readback /
        sets 记录翻译为最终坐标 (区下所有行 −1)."""
        # append-only: 无 inplace → 无行移位
        plan = self._compile(spec_with(self.wd))
        mf = plan["mechanical_facts"]
        self.assertFalse(mf["shift"]["present"])
        op_rows = {int(re.search(r"/([A-Z]+)(\d+)$", op["path"]).group(2))
                   for op in plan["operations"] if op["command"] == "set"}
        rb_rows = {int(re.search(r"/([A-Z]+)(\d+)$", rb["path"]).group(2))
                   for rb in plan["readback"]}
        self.assertEqual(op_rows, rb_rows, "无移位时 readback 坐标应等于模板坐标")
        # inplace trim: 3 数据行, capacity 4 → trim 1, shift −1
        wd2 = make_inplace_workdir(self.tmp, n_source_rows=3)
        wd2["workdir"] = self.tmp
        spec2 = copy.deepcopy(INPLACE_BASE_SPEC)
        spec2["fingerprints"] = wd2["manifest"]["fingerprints"]
        spec2["mapping"]["targets"][0]["sets"] = [
            {"path": "A4", "value": "To Messrs: MXP"},
            {"path": "A14", "value": "* ship to Algeria"},
        ]
        plan2 = compile_fill.compile_spec(spec2, wd2["manifest"], self.tmp)
        mf2 = plan2["mechanical_facts"]
        self.assertTrue(mf2["shift"]["present"])
        self.assertEqual(mf2["shift"]["region"], "7-10")
        self.assertEqual(mf2["shift"]["value"], -1)
        self.assertTrue(mf2["shift"]["readback_translated"])
        # op path 保持模板坐标
        set_paths = {op["path"] for op in plan2["operations"]
                     if op["command"] == "set" and op["path"].endswith(("A4", "A14"))}
        self.assertEqual(set_paths, {"/S/A4", "/S/A14"})
        # readback 与 sets 记录翻译为最终坐标 (A14 → A13)
        rb = {rb["path"] for rb in plan2["readback"]}
        self.assertIn("/S/A13", rb)
        self.assertNotIn("/S/A14", rb)
        self.assertEqual([s["path"] for s in plan2["sets"]], ["/S/A4", "/S/A13"])

    # ── 机械事实栏 (从契约派生, 非自由文本) ──
    def test_mechanical_facts_derived_not_free_text(self):
        """mechanical_facts 与 ops/布局机械一致: 对同一 plan 重算 removes、
        锚点链与 add 插入行, 断言事实栏 == 重算结果 (机械派生, 非自由文本)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["remove_rows"] = [3]
        plan = self._compile(spec)
        mf = plan["mechanical_facts"]
        self.assertEqual(mf["removes"]["rows"], sorted(self._remove_rows_of(plan)))
        self.assertEqual(mf["base_last_row"],
                         spec["mapping"]["targets"][0]["base_last_row"])
        after_rows = sorted({int(re.search(r"/row\[(\d+)\]$", op["after"]).group(1))
                             for op in plan["operations"]
                             if op["command"] == "add" and "after" in op})
        self.assertEqual(mf["anchor_chain"]["after_rows"], after_rows)
        insert_rows = [a + 1 for a in after_rows]
        self.assertEqual(mf["add_zone"]["append_insert_rows"], sorted(set(insert_rows)))
        self.assertEqual(mf["add_zone"]["overflow_insert_rows"], [])
        # 契约声明的顺序不变量与 plan 一致
        self.assertEqual(mf["op_order_invariant"], "clear → add → remove → merge → fill")

    def test_mapping_md_mechanical_facts_rendered(self):
        """mapping.md「执行机械事实」栏由 mechanical_facts 派生渲染 — 对 append
        运行, removes ≤ base 的结论显式可见 (验收标准: 埃及案例 Agent 不读源码
        即可回答 remove/add 交互问题)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["remove_rows"] = [3]
        plan = self._compile(spec)
        md = compile_fill.render_mapping(spec, plan, self.wd["manifest"])
        self.assertIn("## 执行机械事实", md)
        self.assertIn("clear → add → remove → merge → fill", md)
        self.assertIn("removes: [3]", md)
        self.assertIn("不被 add 推移", md)
        self.assertIn("readback 坐标 == 模板坐标", md)
        self.assertIn("TEMPLATE_ROW_GAP", md)

    def test_op_order_invariant_two_forms(self):
        """E1 两形态 (2026-08-13 修正): mechanical_facts.op_order_invariant 按
        plan 形态派生 — append-only → clear→add→remove→merge→fill; inplace
        混合 → append 块全部操作→sets→inplace 结构→inplace 值写. 契约字符串
        与 FILLSPEC E1 同源 (防止回退到单一恒式)."""
        # append-only 形态
        plan = self._compile(spec_with(self.wd))
        self.assertEqual(plan["mechanical_facts"]["op_order_invariant"],
                         "clear → add → remove → merge → fill")
        # inplace 混合形态 (INPLACE_BASE_SPEC: region 7-10 + sets)
        wd2 = make_inplace_workdir(self.tmp, n_source_rows=3)
        wd2["workdir"] = self.tmp
        spec2 = copy.deepcopy(INPLACE_BASE_SPEC)
        spec2["fingerprints"] = wd2["manifest"]["fingerprints"]
        spec2["mapping"]["targets"][0]["sets"] = [
            {"path": "A4", "value": "To Messrs: MXP"},
        ]
        plan2 = compile_fill.compile_spec(spec2, wd2["manifest"], self.tmp)
        self.assertEqual(
            plan2["mechanical_facts"]["op_order_invariant"],
            "append 块全部操作 → sets → 终末 inplace 块结构操作 "
            "(overflow 克隆 add → trim remove) → inplace 值操作")
        self.assertNotEqual(plan["mechanical_facts"]["op_order_invariant"],
                            plan2["mechanical_facts"]["op_order_invariant"],
                            "两形态的顺序不变量必须不同 (E1 双表述)")

    def test_anchor_chain_gap_rejected_before_plan(self):
        """锚点链事实的背书: add 引用行若落在目标行号空洞 → TEMPLATE_ROW_GAP 拒绝
        (exit 3) — plan 能产出 ⟹ 锚点链引用行均存在. mapping.md 的
        gap_checked 是 plan 存在性不变量, 非自由文本."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            wd["workdir"] = tmp
            meta = json.loads((tmp / "target_meta.json").read_text(encoding="utf-8"))
            meta["row_gaps"] = [4]  # title add 的 after 锚点 /row[4] (BASE_SPEC)
            (tmp / "target_meta.json").write_text(json.dumps(meta), encoding="utf-8")
            spec = spec_with(wd)
            codes = compile_fail_codes(wd, spec)
            self.assertIn("TEMPLATE_ROW_GAP", codes)


class CapabilityMappingContractTests(unittest.TestCase):
    """能力映射表 (FILLSPEC「能力映射表」章节) 的编译用例背书.

    每条"一等/变通"表达模式 → 一个编译用例, 保证新 MOD 规则入库时可对照
    验证"这条业务规则能否表达、用什么模式表达".
    """

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _fail_codes(self, spec) -> list[str]:
        return compile_fail_codes(self.wd, spec)

    def test_arith_derived_expression(self):
        """算术派生 (减法) → per_row formula — 一等."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"G": "IFERROR(ROUND(A{r}-B{r},2),0)"}}
        self.assertEqual(self._fail_codes(spec), [])

    def test_field_inheritance_expression(self):
        """字段继承 → columns + lookups (missing: empty) — 一等."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "F", "lookup": {"name": "fields",
                                       "field": "copper", "missing": "empty"}})
        spec["mapping"]["targets"][0]["lookups"] = [
            {"name": "fields", "from": "inheritance.json", "key_column": "C",
             "fields": ["copper"], "missing": "empty"}]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        self.assertEqual([w["value"] for w in plan["writes"] if w["col"] == "F"],
                         ["P-1", "P-2", ""])

    def test_routing_expression(self):
        """路由 (条件取列) → selectors 行过滤 + fallback 列回退 — 一等."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["rows"]["selectors"] = [
            {"column": "A", "pattern": "家用*"}]
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E", "fallback": "B"}]
        self.assertEqual(self._fail_codes(spec), [])

    def test_zero_policy_expression(self):
        """0-口径 → 常量 value "0" + 多列求和缺失输入按 0 — 一等."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].extend([
            {"target": "F", "value": "0"},
            {"source": ["F", "G"], "target": "G"}])
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        self.assertEqual(sorted({w["value"] for w in plan["writes"] if w["col"] == "G"}),
                         ["15", "3", "9"])
        self.assertIn("F", {w["col"] for w in plan["writes"]})

    def test_constant_expression(self):
        """常量 → columns.value — 一等."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["columns"].append(
            {"target": "F", "value": "DP AT SIGHT"})
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        self.assertEqual(sorted({w["value"] for w in plan["writes"] if w["col"] == "F"}),
                         ["DP AT SIGHT"])

    def test_per_group_total_workaround(self):
        """每组合计 → 变通: blocks[] 每组合一块 + 块级 aggregates 1:{n}
        (组合边界由数据决定, spec 无法表达动态组内范围)."""
        spec = spec_with(self.wd)
        block_cfg = {
            "clone_roles": [
                {"role": "spacer"},
                {"role": "title", "template_row": 1, "value": "块标题"},
                {"role": "header", "template_row": 2},
                {"role": "data", "template_row": 3},
            ],
            "formulas": {"aggregates": [{"col": "G", "rows": "1:{n}",
                                          "formula": "SUM(A{r1}:A{r2})",
                                          "style": "anchor"}]},
        }
        spec["mapping"]["targets"][0]["blocks"] = [
            dict(block_cfg, rows={"source": "source_maoli",
                                  "selectors": [{"column": "A", "pattern": "家用*"}]}),
            dict(block_cfg, rows={"source": "source_maoli",
                                  "selectors": [{"column": "A", "pattern": "商用*"}]}),
        ]
        spec["validation"]["key_outputs"] = ["A6", "G8", "G13"]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        agg = [op["path"] for op in plan["operations"]
               if "formula" in op.get("props", {}) and "SUM" in op["props"]["formula"]]
        # block1: spacer5 title6 header7 data8-9 → 聚合 G8; block2: spacer10 title11 header12 data13
        self.assertEqual(agg, ["/S/G8", "/S/G13"])

    def test_per_group_total_first_class(self):
        """每组合计 → 一等: group_aggregates (group_by + col + formula,
        组锚点行落公式, 组边界由数据决定)."""
        wd = make_egypt_workdir(self.tmp)
        wd["workdir"] = self.tmp
        spec = spec_with(wd)
        spec["fingerprints"] = wd["manifest"]["fingerprints"]
        spec["mapping"]["targets"][0]["formulas"] = {
            "group_aggregates": [
                {"group_by": "A", "col": "V",
                 "formula": "SUM(T{r1}:T{r2})", "style": "anchor"}]}
        self.assertEqual(compile_fail_codes(wd, spec), [])
        plan = compile_spec_with(wd, spec)
        self.assertEqual(
            [op["path"] for op in plan["operations"]
             if op["command"] == "set" and op["path"].startswith("/S/V")
             and "formula" in op.get("props", {})],
            ["/S/V7", "/S/V9", "/S/V11"])

    def test_per_group_total_explicit_ranges_same_block(self):
        """每组合计 → 一等: 单块 + 多条显式范围聚合 (V/W 同块, 聚合列不进
        nulls) — 埃及案例最终方案的同形脱敏, 编译通过."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "spacer"},
            {"role": "title", "template_row": 1, "value": "块标题"},
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E"},
        ]
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [
                {"col": "V", "rows": "1:2",
                 "formula": "IFERROR(ROUND(SUM(T{r1}:T{r2})/SUM(S{r1}:S{r2}),4),0)",
                 "style": "anchor"},
                {"col": "W", "rows": "1:2",
                 "formula": "IFERROR(ROUND(SUM(U{r1}:U{r2})/SUM(S{r1}:S{r2}),4),0)",
                 "style": "anchor"},
                {"col": "V", "rows": "3:3",
                 "formula": "IFERROR(ROUND(SUM(T{r1}:T{r2})/SUM(S{r1}:S{r2}),4),0)",
                 "style": "anchor"},
                {"col": "W", "rows": "3:3",
                 "formula": "IFERROR(ROUND(SUM(U{r1}:U{r2})/SUM(S{r1}:S{r2}),4),0)",
                 "style": "anchor"},
            ]}
        spec["validation"]["key_outputs"] = ["A8", "V8", "W8", "V10", "W10"]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        formulas = {op["path"]: op["props"]["formula"] for op in plan["operations"]
                    if "formula" in op.get("props", {})}
        # spacer5 title6 header7 data8-10: 组1 = 行 8-9, 组2 = 行 10
        self.assertIn("/S/V8", formulas)
        self.assertIn("/S/W8", formulas)
        self.assertIn("/S/V10", formulas)
        self.assertIn("/S/W10", formulas)
        self.assertEqual(formulas["/S/V8"], "IFERROR(ROUND(SUM(T8:T9)/SUM(S8:S9),4),0)")
        self.assertEqual(formulas["/S/V10"], "IFERROR(ROUND(SUM(T10:T10)/SUM(S10:S10),4),0)")

    def test_per_group_total_aggregate_col_in_nulls_rejected(self):
        """每组合计负面表达: 同形 spec 聚合列 (F) 进 nulls → 锚点双写
        ("first as empty" — nulls 先清空锚点格, 聚合再写公式)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "spacer"},
            {"role": "title", "template_row": 1, "value": "块标题"},
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        spec["mapping"]["targets"][0]["columns"] = [
            {"source": "A", "target": "A"},
            {"source": "B", "target": "B"},
            {"source": "C", "target": "C"},
            {"source": "D", "target": "E"},
        ]
        spec["mapping"]["targets"][0]["nulls"] = [
            {"col": "D", "rows": "all"}, {"col": "F", "rows": "all"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [
                {"col": "F", "rows": "1:2",
                 "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
                {"col": "F", "rows": "3:3",
                 "formula": "SUM(A{r1}:A{r2})", "style": "anchor"},
            ]}
        spec["validation"]["key_outputs"] = ["A8"]
        codes = self._fail_codes(spec)
        self.assertIn("DUPLICATE_TARGET_WRITE", codes)
        self.assertNotIn("AGG_RANGE_INVALID", codes)

    def test_pptx_group_merges_not_rolled_out(self):
        """暂无表达 (pptx group_merges) → PPTX_CAPABILITY_NOT_ROLLED_OUT."""
        self.wd["manifest"]["target"]["sheet"] = "slide[1]/table[@id=1]"
        spec = spec_with(self.wd)
        spec["inputs"]["platform"] = "pptx"
        spec["inputs"]["target_sheet"] = "slide[1]/table[@id=1]"
        t = spec["mapping"]["targets"][0]
        t["sheet"] = "slide[1]/table[@id=1]"
        t["first_data_row"] = 2
        t["group_merges"] = [{"col": "A", "group_by": "A"}]
        self.assertIn("PPTX_CAPABILITY_NOT_ROLLED_OUT", self._fail_codes(spec))

    def test_pptx_inplace_not_rolled_out(self):
        """暂无表达 (pptx inplace) → PPTX_CAPABILITY_NOT_ROLLED_OUT."""
        self.wd["manifest"]["target"]["sheet"] = "slide[1]/table[@id=1]"
        spec = spec_with(self.wd)
        spec["inputs"]["platform"] = "pptx"
        spec["inputs"]["target_sheet"] = "slide[1]/table[@id=1]"
        t = spec["mapping"]["targets"][0]
        t["sheet"] = "slide[1]/table[@id=1]"
        t["first_data_row"] = 2
        t["clone_roles"] = [{"role": "data", "mode": "inplace",
                             "start_row": 2, "capacity": 4, "template_row": 2}]
        t["rows"] = {"source": "source_maoli"}
        t.pop("base_last_row", None)
        self.assertIn("PPTX_CAPABILITY_NOT_ROLLED_OUT", self._fail_codes(spec))

    # ── issue 06: pptx 未支持声明 fail-closed (不再静默丢弃) ──

    def _pptx_spec(self) -> dict:
        """Base spec flipped to a pptx table target (slide[1]/table[@id=1]),
        with the pptx-unsupported BASE_SPEC nulls dropped and a DOM-path
        key_output (the supported pptx shape)."""
        self.wd["manifest"]["target"]["sheet"] = "slide[1]/table[@id=1]"
        spec = spec_with(self.wd)
        spec["inputs"]["platform"] = "pptx"
        spec["inputs"]["target_sheet"] = "slide[1]/table[@id=1]"
        t = spec["mapping"]["targets"][0]
        t["sheet"] = "slide[1]/table[@id=1]"
        t["first_data_row"] = 2
        t.pop("nulls", None)
        spec["validation"]["key_outputs"] = ["/slide[1]/table[@id=1]/tr[2]/tc[1]"]
        return spec

    def _fail_defects(self, spec) -> list[dict]:
        """Compile-fail defect payload (parsed stderr); asserts exit 3."""
        from io import StringIO
        buf = StringIO()
        old = sys.stderr
        sys.stderr = buf
        try:
            with self.assertRaises(SystemExit) as ctx:
                compile_fill.compile_spec(spec, self.wd["manifest"],
                                          self.wd["workdir"])
            self.assertEqual(ctx.exception.code, 3)
        finally:
            sys.stderr = old
        return json.loads(buf.getvalue()).get("defects", [])

    def test_pptx_per_row_formula_rejected(self):
        """pptx formulas.per_row → PPTX_CAPABILITY_NOT_ROLLED_OUT (曾静默丢弃)."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"G": "IFERROR(ROUND(A{r}-B{r},2),0)"}}
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_CAPABILITY_NOT_ROLLED_OUT"
                 and "per_row" in x.get("message", ""))
        self.assertTrue(d["corrective_action"])

    def test_pptx_aggregates_rejected(self):
        """pptx formulas.aggregates → PPTX_CAPABILITY_NOT_ROLLED_OUT."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "G", "rows": "1:{n}",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_CAPABILITY_NOT_ROLLED_OUT"
                 and "aggregates" in x.get("message", ""))
        self.assertTrue(d["corrective_action"])

    def test_pptx_merges_rejected(self):
        """pptx merges → PPTX_CAPABILITY_NOT_ROLLED_OUT."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["merges"] = [
            {"col": "E", "rows": "1:{n}", "style": "label"}]
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_CAPABILITY_NOT_ROLLED_OUT"
                 and "merges" in x.get("message", ""))
        self.assertTrue(d["corrective_action"])

    def test_pptx_nulls_rejected(self):
        """pptx nulls (克隆残留置空, pptx 无克隆) → PPTX_CAPABILITY_NOT_ROLLED_OUT."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["nulls"] = [{"col": "D", "rows": "all"}]
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_CAPABILITY_NOT_ROLLED_OUT"
                 and "nulls" in x.get("message", ""))
        self.assertTrue(d["corrective_action"])

    def test_pptx_remove_rows_rejected(self):
        """pptx remove_rows (无结构行操作) → PPTX_CAPABILITY_NOT_ROLLED_OUT."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["remove_rows"] = [3]
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_CAPABILITY_NOT_ROLLED_OUT"
                 and "remove_rows" in x.get("message", ""))
        self.assertTrue(d["corrective_action"])

    def test_pptx_columns_props_rejected(self):
        """pptx columns[].props (numberformat 白名单) → 拒绝 — pptx 文本格
        无数字格式, props 曾静默不应用."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["columns"].append(
            {"source": "B", "target": "E", "props": {"numberformat": "0.00"}})
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_CAPABILITY_NOT_ROLLED_OUT"
                 and "props" in x.get("message", ""))
        self.assertTrue(d["corrective_action"])

    def test_pptx_rows_out_of_bounds_rejected(self):
        """first_data_row + 匹配行数越过表格实际行数 → 编译期拒绝
        (PPTX_TARGET_ROWS_OUT_OF_BOUNDS), 而非执行期才失败."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["first_data_row"] = 19  # 19+3−1=21 > 20
        defects = self._fail_defects(spec)
        d = next(x for x in defects
                 if x["code"] == "PPTX_TARGET_ROWS_OUT_OF_BOUNDS")
        self.assertIn("first_data_row", d["message"])
        self.assertIn("tr[21]", d["message"])
        self.assertTrue(d["corrective_action"])

    def test_pptx_rows_exactly_at_last_row_accepted(self):
        """边界: 填充恰好止于表格最后一行 (first + n − 1 == dims.rows)
        → 接受 (越界指 tr 超过实际行数, 不是首行+行数越过)."""
        spec = self._pptx_spec()
        spec["mapping"]["targets"][0]["first_data_row"] = 18  # 18+3−1=20 == 20
        spec["validation"]["key_outputs"] = ["/slide[1]/table[@id=1]/tr[18]/tc[1]"]
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        self.assertEqual(plan["operations"][-1]["path"],
                         "/slide[1]/table[@id=1]/tr[20]/tc[3]")

    def test_pptx_basic_fill_compiles(self):
        """pptx 列值填充 (唯一已 rollout 能力) → 编译通过; ops 全为
        tr/tc DOM 路径的 set + text 属性; readback 逐格登记."""
        spec = self._pptx_spec()
        self.assertEqual(self._fail_codes(spec), [])
        plan = compile_spec_with(self.wd, spec)
        ops = plan["operations"]
        self.assertEqual(len(ops), 9)  # 3 源行 × 3 列映射
        for op in ops:
            self.assertEqual(op["command"], "set")
            self.assertTrue(re.fullmatch(
                r"/slide\[1\]/table\[@id=1\]/tr\[[234]\]/tc\[[123]\]", op["path"]),
                op["path"])
            self.assertEqual(set(op["props"]), {"text"})
        rb = plan["readback"]
        self.assertEqual(len(rb), 9)
        self.assertTrue(all(r["kind"] == "value" for r in rb))
        self.assertTrue(all(
            r["path"].startswith("/slide[1]/table[@id=1]/tr[") for r in rb))


class PreformattedQuotationPatternContractTests(unittest.TestCase):
    """issue 02 — 完整 Canonical Pattern `preformatted_quotation_inplace` 的
    机械契约测试: 从 catalog entry 本身 (文本参数替换, 而非测试里手写一份等价
    skeleton) 实例化 spec, 并走 PUBLIC Compiler CLI (MxpEndToEndTests 同款 seam)
    编译出 MXP 同形 plan:

    target extent 40 → 占位区 18 (rows 7-24) → 匹配 13 → Trim 5 → 最终 35;
    group boundaries 数据驱动 (三连续组 + 一个永不合并的 singleton);
    block-external sets 中区下写随 Trim 位移; numberformat 落在 mapped value
    column 的 value-owner operations; key outputs/readback 覆盖组锚点、格式化值
    与绝对写; 零 static defects。"""

    PATTERN_ID = "preformatted_quotation_inplace"

    # One explicit instantiation: catalog fragment placeholders → concrete
    # neutral values (the roles are documented inside the fragment itself).
    SUBSTITUTIONS = {
        "<TARGET_SHEET>": "S", "<BASE_LAST_ROW>": "40",
        "<REGION_START>": "7", "<REGION_CAPACITY>": "18", "<TEMPLATE_ROW>": "8",
        "<SOURCE_NAME>": "source_quotation",
        "<GROUP_SRC>": "A", "<MODEL_SRC>": "B", "<CAP_SRC>": "C",
        "<DESC_SRC>": "D", "<PRICE_SRC>": "E",
        "<GROUP_COL>": "A", "<MODEL_COL>": "B", "<CAP_COL>": "C",
        "<DESC_COL>": "D", "<PRICE_COL>": "E", "<LABEL_COL>": "F",
        "<NUMFMT>": "$#,##0.00",
        "<HDR_ADDR>": "A4", "<HDR_VALUE>": "To Messrs: Example Co.",
        "<HDR2_ADDR>": "F4", "<HDR2_VALUE>": "Date of issue: 2026-01-01",
        "<FOOT_ADDR>": "A36", "<FOOT_VALUE>": "* Valid for 30 days from issue",
        "<KO1>": "A4", "<KO2>": "A7", "<KO3>": "A19",
        "<KO4>": "A36", "<KO5>": "E7", "<KO6>": "E19",
    }

    def _pattern_entry(self) -> dict:
        import yaml
        text = (SKILL_ROOT / "assets" / "combination_patterns.yaml").read_text(
            encoding="utf-8")
        entry = next(p for p in yaml.safe_load(text)["patterns"]
                     if p["id"] == self.PATTERN_ID)
        for key in ("question", "answer", "fragment", "note"):
            self.assertIn(key, entry, f"catalog entry 缺字段 {key!r}")
        return entry

    def _instantiate(self, entry: dict) -> dict:
        """Textual substitution on the catalog fragment (comments included);
        leftover placeholders fail loudly instead of silently passing."""
        import yaml
        frag = entry["fragment"]
        for token, value in self.SUBSTITUTIONS.items():
            self.assertIn(token, frag, f"fragment 缺占位符 {token}")
            frag = frag.replace(token, value)
        leftovers = re.findall(r"<[A-Z0-9_]+>", frag)
        self.assertEqual(leftovers, [],
                         f"fragment 存在未替换占位符: {leftovers}")
        return yaml.safe_load(frag)

    def _spec(self, wd: dict, instantiated: dict) -> dict:
        return {
            "task": {"intent": "预格式报价模板 inplace 填充 (pattern contract)",
                     "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_quotation.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_quotation.xlsx",
                                          "sheets": ["报价"]}],
                       "target_sheet": "S"},
            "fingerprints": {
                "source_structure": wd["manifest"]["fingerprints"]["source_structure"],
                "target_structure": wd["manifest"]["fingerprints"]["target_structure"],
            },
            **instantiated,
            "decisions": ["占位区消费: 13 行物化, 尾部 Trim 5 行由编译器推导 (不写 remove_rows)"],
            "gaps": [],
            "lineage": [{"source": "source_quotation_flat.csv", "role": "primary",
                         "note": "每个匹配源行恰好写入一个占位行"}],
        }

    def _compile_cli(self, tmp: Path) -> dict:
        """The public Compiler CLI seam (the same one MxpEndToEndTests uses)."""
        import subprocess
        proc = subprocess.run(
            [sys.executable, "-X", "utf8",
             str(SKILL_ROOT / "scripts" / "compile_fill.py"),
             "--spec", "fill_spec.yaml", "--workdir", "."],
            cwd=str(tmp), capture_output=True, text=True, encoding="utf-8",
            errors="replace", timeout=300)
        self.assertEqual(proc.returncode, 0, proc.stderr[-2000:])
        return json.loads((tmp / "execution_plan.json").read_text(encoding="utf-8"))

    def test_pattern_instantiation_compiles_to_mxp_homomorphic_plan(self):
        """catalog entry 本身 → 参数替换 → 合法 FillSpec YAML → 公开 CLI 编译:
        plan 断言 inplace 物化 / Trim 5 / 最终 35 / 组边界数据驱动 (singleton
        不合并) / sets 模板坐标执行与最终位移记录 / numberformat 落 ops /
        key outputs 与 readback 覆盖组锚点、格式化值、绝对写; 零 defects."""
        import yaml
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            wd = make_preformatted_quotation_workdir(tmp)
            entry = self._pattern_entry()
            instantiated = self._instantiate(entry)
            spec = self._spec(wd, instantiated)
            (tmp / "fill_spec.yaml").write_text(
                yaml.safe_dump(spec, allow_unicode=True, sort_keys=False),
                encoding="utf-8")
            plan = self._compile_cli(tmp)

            # ── 确定性形状 (MXP 同形关键计数): 40 → 18 占位区 → 13 匹配 →
            #    Trim 5 → 最终 35 ──
            self.assertEqual(plan["schema_version"], "2.5")
            self.assertEqual(plan["expected_final_row_count"], 35)
            self.assertEqual(plan["structural_deltas"],
                             {"adds": 0, "removes": 5, "inplace_trim": 5,
                              "inplace_overflow": 0})
            # terminal inplace data role 被物化
            self.assertEqual(plan["blocks"][-1]["mode"], "inplace")
            self.assertEqual((plan["blocks"][-1]["data_start"],
                              plan["blocks"][-1]["data_end"]), (7, 19))

            # ── group boundaries 数据驱动: 三连续组 + singleton 永不合并 ──
            gb = {g["col"]: g for g in plan["group_boundaries"]}
            self.assertEqual(sorted(gb), ["A", "F"])
            for col in ("A", "F"):
                self.assertEqual((gb[col]["region_start"], gb[col]["region_end"]),
                                 (7, 19), col)
                self.assertEqual(gb[col]["expected_merges"],
                                 [f"{col}7:{col}10", f"{col}11:{col}15",
                                  f"{col}16:{col}18"], col)
            merge_props = [op["props"]["merge"] for op in plan["operations"]
                           if isinstance(op.get("props", {}).get("merge"), str)]
            self.assertNotIn("A19:A19", merge_props)  # Delta singleton 不产生 merge

            # ── sets: op 保持模板坐标执行, 记录翻译为最终坐标 (A36 → A31) ──
            set_ops = {op["path"]: op for op in plan["operations"]
                       if op["command"] == "set"
                       and op["path"] in ("/S/A4", "/S/F4", "/S/A36")}
            self.assertEqual(set(set_ops), {"/S/A4", "/S/F4", "/S/A36"})
            self.assertEqual(set_ops["/S/A36"]["props"]["value"],
                             "* Valid for 30 days from issue")
            self.assertEqual([s["path"] for s in plan["sets"]],
                             ["/S/A4", "/S/F4", "/S/A31"])

            # ── numberformat 进入对应 value-owner operations ──
            e_ops = [op for op in plan["operations"]
                     if op["command"] == "set"
                     and re.fullmatch(r"/S/E\d+", op["path"])]
            self.assertEqual(len(e_ops), 13)
            self.assertTrue(all(op["props"].get("numberformat") == "$#,##0.00"
                                for op in e_ops))

            # ── key outputs / readback 覆盖组锚点、格式化值、绝对写 ──
            self.assertEqual(plan["key_outputs"], [
                {"path": "/S/A4", "kind": "value"},
                {"path": "/S/A7", "kind": "value"},
                {"path": "/S/A19", "kind": "value"},
                {"path": "/S/A31", "kind": "value"},
                {"path": "/S/E7", "kind": "value"},
                {"path": "/S/E19", "kind": "value"},
            ])
            rb = {rb["path"]: rb for rb in plan["readback"]}
            self.assertEqual(rb["/S/A7"]["expect"], "Alpha")     # 组锚点 (映射值)
            self.assertEqual(rb["/S/A19"]["expect"], "Delta")    # singleton 锚点
            self.assertEqual(rb["/S/E7"]["expect"], "100")       # numberformat 值格
            self.assertEqual(rb["/S/A31"]["expect"],
                             "* Valid for 30 days from issue")   # 区下绝对写最终坐标
            self.assertEqual(rb["/S/F7"]["kind"], "empty")       # label-only 锚点清空
            self.assertEqual(plan["warnings"], [])               # 无 static defects


class MultiproductBlockPatternContractTests(unittest.TestCase):
    """issue 03 — 完整 Canonical Pattern `multiproduct_block_append` 的机械契约
    测试: 从 catalog entry 本身 (文本参数替换, 而非测试里手写一份等价 skeleton)
    实例化 spec, 并走 PUBLIC Compiler CLI (MxpEndToEndTests 同款 seam) 编译出
    家用/商用双块 append 的同形 plan:

    base_last_row 4 → 家用块 (5 行数据, 悦风 1:2 + 清爽星 3:5 两组) data 8-12 →
    商用块 (3 行数据, 单组) data 16-18; 每源分组 V 显式范围 merges+aggregates
    (V8/V10 + V16); 总盈亏 W 一条 1:{n} merges+aggregates (W8/W16); 类别列
    group_merges (A) + label-only 列 (E); 克隆残留 D/F/X nulls (每块逐行清空);
    key_outputs (块首 + 聚合组锚点 + 总盈亏锚点) 全 written; 零 static defects。"""

    PATTERN_ID = "multiproduct_block_append"

    # One explicit instantiation: catalog fragment placeholders → concrete
    # neutral values (the roles are documented inside the fragment itself).
    SUBSTITUTIONS = {
        "<TARGET_SHEET>": "S", "<BASE_LAST_ROW>": "4",
        "<SOURCE_HOUSE>": "source_house",
        "<SOURCE_COMMERCIAL>": "source_commercial",
        "<TITLE_HOUSE>": "家用块标题", "<TITLE_COMMERCIAL>": "商用块标题",
        "<GROUP_SRC>": "A", "<GROUP_COL>": "A",
        "<MODEL_SRC>": "B", "<MODEL_COL>": "B",
        "<MODEL2_SRC>": "C", "<MODEL2_COL>": "C",
        "<LABEL_COL>": "E",
        "<NULL_COL_1>": "D", "<NULL_COL_2>": "F", "<NULL_COL_3>": "X",
        "<AGG_COL_V>": "V", "<TOTAL_COL_W>": "W",
        "<V_ROWS_H1>": "1:2", "<V_ROWS_H2>": "3:5", "<V_ROWS_C1>": "1:3",
        "<KO1>": "A8", "<KO2>": "V8", "<KO3>": "V10", "<KO4>": "W8",
        "<KO5>": "A16", "<KO6>": "V16", "<KO7>": "W16",
    }

    def _pattern_entry(self) -> dict:
        import yaml
        text = (SKILL_ROOT / "assets" / "combination_patterns.yaml").read_text(
            encoding="utf-8")
        entry = next(p for p in yaml.safe_load(text)["patterns"]
                     if p["id"] == self.PATTERN_ID)
        for key in ("question", "answer", "fragment", "note"):
            self.assertIn(key, entry, f"catalog entry 缺字段 {key!r}")
        return entry

    def _instantiate(self, entry: dict) -> dict:
        """Textual substitution on the catalog fragment (comments included);
        leftover placeholders fail loudly instead of silently passing."""
        import yaml
        frag = entry["fragment"]
        for token, value in self.SUBSTITUTIONS.items():
            self.assertIn(token, frag, f"fragment 缺占位符 {token}")
            frag = frag.replace(token, value)
        leftovers = re.findall(r"<[A-Z0-9_]+>", frag)
        self.assertEqual(leftovers, [],
                         f"fragment 存在未替换占位符: {leftovers}")
        return yaml.safe_load(frag)

    def _spec(self, wd: dict, instantiated: dict) -> dict:
        return {
            "task": {"intent": "家用+商用双数据块追加 (pattern contract)",
                     "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_maoli.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_maoli.xlsx",
                                          "sheets": ["FRESH订家用机型毛利情况",
                                                     "商用机型毛利情况"]}],
                       "target_sheet": "S"},
            "fingerprints": {
                "source_structure": wd["manifest"]["fingerprints"]["source_structure"],
                "target_structure": wd["manifest"]["fingerprints"]["target_structure"],
            },
            **instantiated,
            "decisions": ["双数据块按家用/商用各一块追加; 每块总盈亏 W 一条 1:{n}"],
            "gaps": [],
            "lineage": [{"source": "source_house_flat.csv", "role": "primary",
                         "note": ""},
                        {"source": "source_commercial_flat.csv", "role": "primary",
                         "note": ""}],
        }

    def _compile_cli(self, tmp: Path) -> dict:
        """The public Compiler CLI seam (the same one MxpEndToEndTests uses)."""
        import subprocess
        proc = subprocess.run(
            [sys.executable, "-X", "utf8",
             str(SKILL_ROOT / "scripts" / "compile_fill.py"),
             "--spec", "fill_spec.yaml", "--workdir", "."],
            cwd=str(tmp), capture_output=True, text=True, encoding="utf-8",
            errors="replace", timeout=300)
        self.assertEqual(proc.returncode, 0, proc.stderr[-2000:])
        return json.loads((tmp / "execution_plan.json").read_text(encoding="utf-8"))

    def _agg_paths(self, plan) -> list[str]:
        return [op["path"] for op in plan["operations"]
                if "formula" in op.get("props", {})
                and "SUM" in op["props"]["formula"]]

    def _merge_ranges(self, plan) -> list[str]:
        return [op["props"]["merge"] for op in plan["operations"]
                if isinstance(op.get("props", {}).get("merge"), str)]

    def test_pattern_instantiation_compiles_to_homomorphic_plan(self):
        """catalog entry 本身 → 参数替换 → 合法 FillSpec YAML → 公开 CLI 编译:
        两块布局 (家用 data 8-12 / 商用 data 16-18) + 每源分组 V 的
        merges+aggregates (V8/V10/V16) + 总盈亏 W 1:{n} (W8/W16) + 类别列
        group_merges (A) + label-only (E) + 克隆残留 D/F/X nulls 逐行清空 +
        key_outputs 全 written; 零 defects。"""
        import yaml
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            wd = make_multiproduct_block_workdir(tmp)
            entry = self._pattern_entry()
            instantiated = self._instantiate(entry)
            spec = self._spec(wd, instantiated)
            (tmp / "fill_spec.yaml").write_text(
                yaml.safe_dump(spec, allow_unicode=True, sort_keys=False),
                encoding="utf-8")
            plan = self._compile_cli(tmp)
            self.assertEqual(plan["schema_version"], "2.5")

            # ── 两块布局: 家用 5 行 (两组) data 8-12, 商用 3 行 data 16-18 ──
            data_blocks = [b for b in plan["blocks"] if b.get("data_start")]
            self.assertEqual(len(data_blocks), 2)
            self.assertEqual([(b["data_start"], b["data_end"]) for b in data_blocks],
                             [(8, 12), (16, 18)])

            # ── 每源分组 V 的 merges+aggregates + 总盈亏 W 1:{n} ──
            self.assertEqual(self._agg_paths(plan),
                             ["/S/V8", "/S/V10", "/S/W8", "/S/V16", "/S/W16"],
                             "家用两组 V + 总盈亏 W + 商用一组 V")
            merges = set(self._merge_ranges(plan))
            self.assertEqual(merges, {
                # 类别列 group_merges (A) + label-only 列 (E)
                "A8:A9", "A10:A12", "A16:A18",
                "E8:E9", "E10:E12", "E16:E18",
                # 每源分组 V 显式范围 merges
                "V8:V9", "V10:V12", "V16:V18",
                # 总盈亏 W 一条 1:{n}
                "W8:W12", "W16:W18",
            }, "块合并形态: 类别/label-only/每源分组 V/总盈亏 W")
            self.assertNotIn("V8:V10", merges)  # V 按组范围, 不整块

            # ── group_boundaries 记录类别列 (A) 与 label-only 列 (E) ──
            g_a_house = next(g for g in plan["group_boundaries"]
                             if g["col"] == "A" and g["region_start"] == 8)
            self.assertEqual((g_a_house["region_start"], g_a_house["region_end"]),
                             (8, 12), "家用块类别列 group boundary 8-12")
            self.assertEqual(g_a_house["expected_merges"], ["A8:A9", "A10:A12"])
            self.assertEqual(
                sorted({g["col"] for g in plan["group_boundaries"]}), ["A", "E"],
                "group_boundaries 覆盖类别列 A 与 label-only 列 E")

            # ── 克隆残留 D/F/X: 每块每数据行 explicit empty readback ──
            empty = {rb["path"] for rb in plan["readback"] if rb["kind"] == "empty"}
            for col in ("D", "F", "X"):
                self.assertEqual(
                    {p for p in empty if re.fullmatch(rf"/S/{col}\d+", p)},
                    {f"/S/{col}{r}" for r in (8, 9, 10, 11, 12, 16, 17, 18)},
                    f"nulls {col} 应逐行清空 (家用 5 + 商用 3)")

            # ── key_outputs 全 written: 块首 + 聚合组锚点 + 总盈亏锚点 ──
            self.assertEqual(plan["key_outputs"], [
                {"path": "/S/A8", "kind": "value"},
                {"path": "/S/V8", "kind": "nonempty"},
                {"path": "/S/V10", "kind": "nonempty"},
                {"path": "/S/W8", "kind": "nonempty"},
                {"path": "/S/A16", "kind": "value"},
                {"path": "/S/V16", "kind": "nonempty"},
                {"path": "/S/W16", "kind": "nonempty"},
            ])
            self.assertEqual(plan["warnings"], [])  # 无 static defects


class SingleBlockQuotationPatternContractTests(unittest.TestCase):
    """issue 02 — 完整 Canonical Pattern `single_quotation_block_append` 的机械
    契约测试: 从 catalog entry 本身 (文本参数替换, 而非测试里手写一份等价 skeleton)
    实例化 spec, 并走 PUBLIC Compiler CLI (MxpEndToEndTests 同款 seam) 编译出
    MXP 17_MXP 单块 append 的同形 plan:

    base_last_row 4 → 单块 spacer5/title6/header7/data8-12 (5 数据行);
    净价 per_row Q/T/U/V/W (ROUND 精准: 减/乘/除 ROUND2、纯加法 T 不加、
    比率 W ROUND4); 总盈亏 Y 一条 1:{n} merges+aggregates (聚合锚点=块首行=
    合并锚点); 商业/费用列 S 0-口径常量; 克隆残留 + [1] 外部引用列 nulls
    (X/Z/H 逐行清空); key_outputs (块首 + 公式格 + 总盈亏锚点) 全 written;
    零 static defects, 且因 selector 排除表头行 → 无 HEADER_ROW_CONSIDERED_DATA。
    """

    PATTERN_ID = "single_quotation_block_append"

    # One explicit instantiation: catalog fragment placeholders → concrete
    # neutral values (the roles are documented inside the fragment itself).
    SUBSTITUTIONS = {
        "<TARGET_SHEET>": "S", "<BASE_LAST_ROW>": "4", "<TITLE>": "单块报价标题",
        "<SOURCE_NAME>": "source_quote",
        "<GROUP_SRC>": "A", "<GROUP_COL>": "A",
        "<MODEL_SRC>": "B", "<MODEL_COL>": "B",
        "<COST_SRC>": "D", "<COST_COL>": "R", "<CMD_COL>": "S",
        "<SELECTOR_COL>": "A", "<SELECTOR_PATTERN>": "核心*",
        "<TOTAL_COL>": "Y",
        "<NULL_COL_1>": "X", "<NULL_COL_2>": "Z", "<NULL_COL_3>": "H",
        "<KO1>": "A8", "<KO2>": "Q8", "<KO3>": "W8", "<KO4>": "Y8",
    }

    def _pattern_entry(self) -> dict:
        import yaml
        text = (SKILL_ROOT / "assets" / "combination_patterns.yaml").read_text(
            encoding="utf-8")
        entry = next(p for p in yaml.safe_load(text)["patterns"]
                     if p["id"] == self.PATTERN_ID)
        for key in ("question", "answer", "fragment", "note"):
            self.assertIn(key, entry, f"catalog entry 缺字段 {key!r}")
        return entry

    def _instantiate(self, entry: dict) -> dict:
        """Textual substitution on the catalog fragment (comments included);
        leftover placeholders fail loudly instead of silently passing."""
        import yaml
        frag = entry["fragment"]
        for token, value in self.SUBSTITUTIONS.items():
            self.assertIn(token, frag, f"fragment 缺占位符 {token}")
            frag = frag.replace(token, value)
        leftovers = re.findall(r"<[A-Z0-9_]+>", frag)
        self.assertEqual(leftovers, [],
                         f"fragment 存在未替换占位符: {leftovers}")
        return yaml.safe_load(frag)

    def _spec(self, wd: dict, instantiated: dict) -> dict:
        return {
            "task": {"intent": "核价邮件 → 报价汇总单块追加 (pattern contract)",
                     "selected_mod": "NONE", "selected_mod_revision": None},
            "inputs": {"sources": ["source_quote.xlsx"], "target": "target.xlsx",
                       "source_sheets": [{"source": "source_quote.xlsx",
                                          "sheets": ["核价"]}],
                       "target_sheet": "S"},
            "fingerprints": {
                "source_structure": wd["manifest"]["fingerprints"]["source_structure"],
                "target_structure": wd["manifest"]["fingerprints"]["target_structure"],
            },
            **instantiated,
            "decisions": ["单块 append: 净价公式链按 ROUND 精准 (减/乘/除 2、比率 4、"
                          "纯加法不加)"],
            "gaps": [],
            "lineage": [{"source": "source_quote_flat.csv", "role": "primary",
                         "note": "每个匹配源行恰好写入一个数据行"}],
        }

    def _compile_cli(self, tmp: Path) -> dict:
        """The public Compiler CLI seam (the same one MxpEndToEndTests uses)."""
        import subprocess
        proc = subprocess.run(
            [sys.executable, "-X", "utf8",
             str(SKILL_ROOT / "scripts" / "compile_fill.py"),
             "--spec", "fill_spec.yaml", "--workdir", "."],
            cwd=str(tmp), capture_output=True, text=True, encoding="utf-8",
            errors="replace", timeout=300)
        self.assertEqual(proc.returncode, 0, proc.stderr[-2000:])
        return json.loads((tmp / "execution_plan.json").read_text(encoding="utf-8"))

    def _per_row_formulas(self, plan) -> dict:
        return {op["path"]: op["props"]["formula"] for op in plan["operations"]
                if op["command"] == "set" and "formula" in op.get("props", {})
                and "SUM" not in op["props"]["formula"]}

    def test_pattern_instantiation_compiles_to_homomorphic_plan(self):
        """catalog entry 本身 → 参数替换 → 合法 FillSpec YAML → 公开 CLI 编译:
        单块布局 (data 8-12) + 净价 per_row Q/T/U/V/W 的 ROUND 精准公式 + 总盈亏
        Y 1:{n} merges+aggregates + 0-口径常量 S + 克隆残留/外部引用 X/Z/H nulls
        逐行清空 + key_outputs 全 written; 零 defects; selector 排除表头行 →
        无 HEADER_ROW_CONSIDERED_DATA 警告."""
        import yaml
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            wd = make_single_block_workdir(tmp)
            entry = self._pattern_entry()
            instantiated = self._instantiate(entry)
            spec = self._spec(wd, instantiated)
            (tmp / "fill_spec.yaml").write_text(
                yaml.safe_dump(spec, allow_unicode=True, sort_keys=False),
                encoding="utf-8")
            plan = self._compile_cli(tmp)
            self.assertEqual(plan["schema_version"], "2.5")

            # ── 单块布局: 5 数据行 data 8-12 ──
            data_blocks = [b for b in plan["blocks"] if b.get("data_start")]
            self.assertEqual(len(data_blocks), 1)
            self.assertEqual([(b["data_start"], b["data_end"]) for b in data_blocks],
                             [(8, 12)])

            # ── 净价 per_row ROUND 精准: 减/乘/除 ROUND2, 纯加法 T 不加, 比率 W ROUND4 ──
            formulas = {p[len("/S/"):]: f for p, f in
                        self._per_row_formulas(plan).items()}
            row8 = {f"{col}8": formulas.get(f"{col}8") for col in "QTUVW"}
            self.assertEqual(list(row8), ["Q8", "T8", "U8", "V8", "W8"])
            q, t = row8["Q8"], row8["T8"]
            self.assertIn("ROUND(", q)
            self.assertLess(q.index("ROUND"), q.index(",2)"))
            self.assertNotIn("ROUND", t)          # 纯加法 T = S+R 不加 ROUND
            for col in ("U8", "V8"):
                self.assertIn("ROUND(", row8[col])
            self.assertIn("ROUND(V8/U8,4)", row8["W8"])  # 比率 ROUND4

            # ── 总盈亏 Y: 一条 1:{n} merges+aggregates (聚合锚点=块首行=合并锚点) ──
            agg_paths = [op["path"] for op in plan["operations"]
                         if "formula" in op.get("props", {})
                         and "SUM" in op["props"]["formula"]]
            self.assertEqual(agg_paths, ["/S/Y8"])
            merges = [op["props"]["merge"] for op in plan["operations"]
                      if isinstance(op.get("props", {}).get("merge"), str)]
            self.assertIn("Y8:Y12", merges, "总盈亏 Y 一条 1:{n} merge")

            # ── 0-口径常量 S 每行写出数值 0 (入公式链费用列无数值 → 0) ──
            s_writes = [w for w in plan["writes"] if w["col"] == "S"]
            self.assertEqual([w["value"] for w in s_writes], ["0"] * 5)

            # ── 克隆残留 + [1] 外部引用列 X/Z/H nulls 逐行清空 ──
            empty = {rb["path"] for rb in plan["readback"] if rb["kind"] == "empty"}
            for col in ("H", "X", "Z"):
                self.assertEqual(
                    {p for p in empty if re.fullmatch(rf"/S/{col}\d+", p)},
                    {f"/S/{col}{r}" for r in range(8, 13)},
                    f"nulls {col} 应逐行清空 (数据行 8-12)")

            # ── key_outputs 全 written: 块首 + 净价公式格 + 比率公式格 + 总盈亏锚点 ──
            self.assertEqual(plan["key_outputs"], [
                {"path": "/S/A8", "kind": "value"},
                {"path": "/S/Q8", "kind": "nonempty"},
                {"path": "/S/W8", "kind": "nonempty"},
                {"path": "/S/Y8", "kind": "nonempty"},
            ])
            # selector 排除表头行 → 无 HEADER_ROW_CONSIDERED_DATA; 零 static defects
            self.assertEqual(plan["warnings"], [])


class ProbeTests(unittest.TestCase):
    """compile_fill.py --probe: compile-only verification, zero side effects.

    The probe runs the exact same pipeline as a real compile — its answer is
    authoritative and always matches what a full compile would do."""

    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_ctx.name)
        self.wd = make_workdir(self.tmp)
        self.wd["workdir"] = self.tmp

    def tearDown(self):
        self.tmp_ctx.cleanup()

    def _probe(self, spec) -> dict:
        return compile_fill.probe_spec(spec, self.wd["manifest"], self.tmp)

    def test_probe_accepted_spec(self):
        spec = spec_with(self.wd)
        r = self._probe(spec)
        self.assertTrue(r["accepted"])
        self.assertEqual(r["exit_code"], 0)
        self.assertGreater(r["operations"], 0)
        self.assertEqual(r["defects"], [])

    def test_probe_rejected_with_code(self):
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["group_merges"] = [{"col": "A", "group_by": "A"}]
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "A", "rows": "1:{n}",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        r = self._probe(spec)
        self.assertFalse(r["accepted"])
        self.assertEqual(r["exit_code"], 3)
        self.assertEqual(r["code"], "STATIC_VALIDATION_FAILED")
        codes = {d["code"] for d in r["defects"]}
        self.assertIn("DUPLICATE_TARGET_WRITE", codes)

    def test_probe_matches_full_compile(self):
        """probe 结论与完整编译一致 (同管线, 不会出现 probe 通过但 compile 拒绝)."""
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "per_row": {"G": "IFERROR(ROUND(A{r}-B{r},2),0)"}}
        r = self._probe(spec)
        self.assertTrue(r["accepted"])
        plan = compile_spec_with(self.wd, spec)
        self.assertEqual(r["operations"], len(plan["operations"]))

    def test_probe_writes_no_artifacts(self):
        """probe 零副作用: 不写 execution_plan.json / mapping.md / run_timing.json."""
        spec = spec_with(self.wd)
        self._probe(spec)
        self.assertFalse((self.tmp / "execution_plan.json").exists())
        self.assertFalse((self.tmp / "mapping.md").exists())
        self.assertFalse((self.tmp / "run_timing.json").exists())

    def test_probe_rejected_writes_no_artifacts(self):
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["formulas"] = {
            "aggregates": [{"col": "G", "rows": "1:9",
                            "formula": "SUM(A{r1}:A{r2})", "style": "anchor"}]}
        r = self._probe(spec)
        self.assertFalse(r["accepted"])
        self.assertFalse((self.tmp / "execution_plan.json").exists())
        self.assertFalse((self.tmp / "mapping.md").exists())
        self.assertFalse((self.tmp / "run_timing.json").exists())

    def test_combination_pattern_renamed_columns_compiles(self):
        """组合模式片段复制即用实证: combination_patterns.yaml
        `per_group_total_explicit_ranges` 片段改列名 (V→G, W→H) 后直接编译
        通过 — 零 probe 起步承诺 (2026-08-13 契约修正)."""
        import yaml
        patterns = yaml.safe_load(
            (SKILL_ROOT / "assets" / "combination_patterns.yaml").read_text(encoding="utf-8"))
        frag = next(p["fragment"] for p in patterns["patterns"]
                    if p["id"] == "per_group_total_explicit_ranges")
        block = yaml.safe_load(frag)["formulas"]["aggregates"]
        rename = {"V": "G", "W": "H"}
        aggregates = [dict(a, col=rename.get(a["col"], a["col"])) for a in block]
        self.wd = make_workdir(self.tmp, n_source_rows=5)  # 片段范围 1:2 / 3:5 需 ≥5 行
        self.wd["workdir"] = self.tmp
        spec = spec_with(self.wd)
        spec["mapping"]["targets"][0]["clone_roles"] = [
            {"role": "spacer"},
            {"role": "title", "template_row": 1, "value": "块标题"},
            {"role": "header", "template_row": 2},
            {"role": "data", "template_row": 3},
        ]
        spec["mapping"]["targets"][0]["formulas"] = {"aggregates": aggregates}
        spec["validation"]["key_outputs"] = ["A8", "G8", "H8", "G10", "H10"]
        r = self._probe(spec)
        self.assertTrue(r["accepted"], f"改列名后应编译通过: {r}")
        self.assertEqual(r["defects"], [])


class CapabilitiesTests(unittest.TestCase):
    """compile_fill.py --capabilities: the contract matrix as the compiler
    itself judges it — same PROBE_CASES list as the contract tests, so the
    runtime report, the tests and FILLSPEC.md can never drift apart."""

    def test_capabilities_matrix_matches_expectations(self):
        from _probe_fixtures import PROBE_CASES
        with tempfile.TemporaryDirectory() as tmp:
            results = compile_fill.run_probe_cases(Path(tmp))
        by_id = {r["id"]: r for r in results}
        self.assertEqual(sorted(by_id), sorted(c["id"] for c in PROBE_CASES))
        for case in PROBE_CASES:
            r = by_id[case["id"]]
            if case["expect"] == "accept":
                self.assertTrue(r["accepted"], f"{case['id']} 应被接受")
                self.assertIsNone(r["code"])
            else:
                self.assertFalse(r["accepted"], f"{case['id']} 应被拒绝")
                self.assertEqual(r["code"], case["expect"],
                                 f"{case['id']} 错误码应为 {case['expect']}")

    def test_capabilities_covers_contract_claims(self):
        """契约关键组合必须在 capabilities 报告中 (防探针集被误删)."""
        with tempfile.TemporaryDirectory() as tmp:
            results = compile_fill.run_probe_cases(Path(tmp))
        by_id = {r["id"]: r for r in results}
        for cid in ("group_merges_aggregate_same_col", "group_merges_aggregate_diff_col",
                    "derived_subtraction_pattern", "mapped_group_column_anchor",
                    "lookup_missing_empty", "precision_keep", "per_group_total_blocks",
                    "per_group_total_hardcoded_ranges", "nulls_aggregate_same_col",
                    "pptx_group_merges", "pptx_inplace", "pptx_group_aggregates",
                    "pptx_per_row_formula", "pptx_aggregates", "pptx_merges",
                    "pptx_nulls", "pptx_remove_rows", "pptx_rows_out_of_bounds",
                    "pptx_basic_fill", "group_aggregates_egypt_3_groups",
                    "group_aggregates_whole_run_gate", "lookup_table_empty",
                    "lookup_column_all_missing", "block_top_aggregates_rejected",
                    "block_top_unknown_key_rejected",
                    "block_formulas_replaces_target_per_row",
                    "block_no_formulas_inherits_target_per_row",
                    "multiproduct_block_append",
                    "header_row_considered_data", "header_row_excluded_by_selector",
                    "single_quotation_block_append"):
            self.assertIn(cid, by_id, f"capabilities 缺契约探针 {cid}")

    def test_capabilities_header_row_guard_warning_surface(self):
        """表头行守卫探针在 capabilities 报告呈现代码: 无 selector → 警告代码
        出现, 排除表头 → 无警告 (矩阵与契约测试同源, 不会漂移)."""
        with tempfile.TemporaryDirectory() as tmp:
            results = compile_fill.run_probe_cases(Path(tmp))
        by_id = {r["id"]: r for r in results}
        self.assertIn("HEADER_ROW_CONSIDERED_DATA",
                      by_id["header_row_considered_data"]["warnings"])
        self.assertNotIn("HEADER_ROW_CONSIDERED_DATA",
                         by_id["header_row_excluded_by_selector"]["warnings"])


class ProbeScaffoldTests(unittest.TestCase):
    """make_probe_spec.py: skeleton spec with fingerprints/inputs auto-filled
    from the manifest — the boilerplate elimination that makes --probe cheap."""

    def _scaffold(self, tmp: Path) -> dict:
        import json as _json
        import subprocess
        sys.path.insert(0, str(SKILL_ROOT / "scripts"))
        out = tmp / "probe_spec.yaml"
        r = subprocess.run(
            [sys.executable, str(SKILL_ROOT / "scripts" / "make_probe_spec.py"),
             "--workdir", str(tmp), "--out", str(out)],
            capture_output=True, text=True, encoding="utf-8")
        self.assertEqual(r.returncode, 0, r.stderr)
        payload = _json.loads(r.stdout)
        self.assertEqual(payload["code"], "PROBE_SCAFFOLD_WRITTEN")
        text = out.read_text(encoding="utf-8")
        return _json.loads(text[text.index("{"):])  # 跳过注释头

    def test_scaffold_fills_fingerprints_and_inputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            spec = self._scaffold(tmp)
            mfp = wd["manifest"]["fingerprints"]
            self.assertEqual(spec["fingerprints"]["source_structure"], mfp["source_structure"])
            self.assertEqual(spec["fingerprints"]["target_structure"], mfp["target_structure"])
            self.assertEqual(spec["inputs"]["target"], "target.xlsx")
            self.assertEqual(spec["inputs"]["target_sheet"], "S")
            self.assertEqual(spec["inputs"]["sources"], ["source_maoli.xlsx"])

    def test_scaffold_probes_without_fingerprint_mismatch(self):
        """骨架生成的 spec 直接 probe 不因指纹报错 (样板是骨架, 结果接受与否
        取决于片段; 至少不该死在指纹样板手上)."""
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            spec = self._scaffold(tmp)
            r = compile_fill.probe_spec(spec, wd["manifest"], tmp)
            self.assertNotEqual(r.get("code"), "FILLSPEC_FINGERPRINT_MISMATCH")
            self.assertIn("accepted", r)

    def test_scaffold_missing_manifest_fails(self):
        import subprocess
        with tempfile.TemporaryDirectory() as tmp:
            r = subprocess.run(
                [sys.executable, str(SKILL_ROOT / "scripts" / "make_probe_spec.py"),
                 "--workdir", tmp, "--out", str(Path(tmp) / "p.yaml")],
                capture_output=True, text=True, encoding="utf-8")
            self.assertEqual(r.returncode, 3)
            self.assertIn("MANIFEST_NOT_FOUND", r.stdout)

    def test_scaffold_refuses_overwrite_without_force(self):
        """默认输出 probe_spec.yaml; 已存在文件须 --force, 防止误覆盖
        真实 fill_spec.yaml (2026-08-12 round1: 骨架曾静默覆盖丢失 spec)."""
        import json as _json
        import subprocess
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            wd = make_workdir(tmp)
            # 默认路径 probe_spec.yaml 首次生成成功
            r = subprocess.run(
                [sys.executable, str(SKILL_ROOT / "scripts" / "make_probe_spec.py"),
                 "--workdir", str(tmp)],
                capture_output=True, text=True, encoding="utf-8")
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertTrue((tmp / "probe_spec.yaml").is_file())
            # 再跑一次 → 拒绝覆盖
            r = subprocess.run(
                [sys.executable, str(SKILL_ROOT / "scripts" / "make_probe_spec.py"),
                 "--workdir", str(tmp)],
                capture_output=True, text=True, encoding="utf-8")
            self.assertEqual(r.returncode, 3)
            self.assertIn("PROBE_SCAFFOLD_EXISTS", r.stdout)
            # --force → 覆盖成功
            r = subprocess.run(
                [sys.executable, str(SKILL_ROOT / "scripts" / "make_probe_spec.py"),
                 "--workdir", str(tmp), "--force"],
                capture_output=True, text=True, encoding="utf-8")
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertIn("PROBE_SCAFFOLD_WRITTEN", r.stdout)


class DocCoverageGuardTests(unittest.TestCase):
    """doc-coverage 守卫: 关键交互词必须留在文档中, 章节被误删时测试变红.

    组合行为契约只写经编译用例背书的声明 (见 FillSpecContractTests /
    CapabilityMappingContractTests); 本类只做存在性断言, 不复制行为声明.
    """

    def _fillspec_section(self, heading: str) -> str:
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(encoding="utf-8")
        m = re.search(rf"^##\s+{re.escape(heading)}", text, re.MULTILINE)
        self.assertIsNotNone(m, f"FILLSPEC.md 缺章节 ## {heading}")
        nxt = re.search(r"^##\s", text[m.end():], re.MULTILINE)
        return text[m.end():m.end() + (nxt.start() if nxt else len(text))]

    def _error_code_table(self) -> str:
        """常见编译错误速查表 (错误码表项守卫只认表内条目)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(encoding="utf-8")
        m = re.search(r"^##\s+常见编译错误速查", text, re.MULTILINE)
        self.assertIsNotNone(m, "FILLSPEC.md 缺章节 ## 常见编译错误速查")
        return text[m.end():]

    def test_fillspec_combination_contract_section(self):
        """「组合行为契约」章节同时解释 group_merges 与 aggregates 的交互
        (Agent 按问题定位, 而不是按特性清单)."""
        section = self._fillspec_section("组合行为契约")
        self.assertIn("group_merges", section)
        self.assertIn("aggregates", section)
        self.assertIn("DUPLICATE_TARGET_WRITE", section)

    def test_fillspec_derived_column_pattern_word(self):
        """减法派生列标准模式词 (FLD-006) 在契约章节中."""
        section = self._fillspec_section("组合行为契约")
        self.assertIn("减法", section)
        self.assertIn("per_row", section)
        self.assertIn("ROUND", section)

    def test_fillspec_q5_zero_policy_duality_words(self):
        """FILLSPEC Q5 含 0-口径二分契约词 (issue 01): 入公式链字段 → 数值 0;
        独立展示、不入公式链字段才可留空 — Case 07 §8 教训 (空串进公式链 →
        #VALUE! → IFERROR 兜底 0)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q5:.*\n", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q5 小节")
        q5 = section[m.end():]
        for word in ("入公式链", "数值 0", "value: \"0\"", "独立展示",
                     "IFERROR", "兜底"):
            self.assertIn(word, q5, f"FILLSPEC Q5 缺 0-口径契约词 {word!r}")

    def test_combination_patterns_zero_policy_duality(self):
        """combination_patterns.yaml zero_policy 同步 0-口径二分 (issue 01):
        入公式链 → value "0"; 独立展示才可留空; 含反例警示 (空串 → IFERROR 兜底)."""
        text = (SKILL_ROOT / "assets" / "combination_patterns.yaml").read_text(
            encoding="utf-8")
        m = re.search(r"- id: zero_policy", text)
        self.assertIsNotNone(m, "缺 zero_policy pattern")
        entry = text[m.start():m.start() + 800]
        for word in ("入公式链", "数值 0", "独立展示", "IFERROR", "兜底"):
            self.assertIn(word, entry, f"zero_policy 缺 0-口径二分词 {word!r}")

    def test_skill_md_formula_zero_policy_duality(self):
        """SKILL「公式约定」0-口径段含二分词 (issue 01): 入公式链 → 数值 0,
        独立展示才可留空."""
        text = self._skill_md_text()
        for word in ("入公式链", "数值 0", "独立展示", "IFERROR", "兜底"):
            self.assertIn(word, text, f"SKILL.md 缺 0-口径二分词 {word!r}")

    def test_known_traps_zero_policy_formula_chain(self):
        """KNOWN_TRAPS 含「费用列空串进公式链 → 净价兜底 0」机械事实
        (issue 01, Case 07 §8): 空串被公式链按非空文本判错 → IFERROR 兜底 0."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(
            encoding="utf-8")
        for word in ("费用列空串进公式链", "兜底", "IFERROR", "0-口径二分"):
            self.assertIn(word, text, f"KNOWN_TRAPS 缺 0-口径机械事实词 {word!r}")

    def test_fillspec_precision_recommendation_order(self):
        """round4 推荐次序在契约章节中: Q7 小节正文内 round4 必须先于 keep
        出现 (推荐序被反转时测试变红)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q7:.*\n", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q7 小节")
        q7 = section[m.end():]
        self.assertIn("round4", q7)
        self.assertIn("keep", q7)
        self.assertLess(q7.index("round4"), q7.index("keep"))

    def test_fillspec_q7_keep_column_width_check(self):
        """Q7 含 keep 列宽实测背书契约 (issue 04): 缺陷码 + 列宽未知豁免
        语义都必须在正文 (防契约条目被误删)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q7:.*\n", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q7 小节")
        q7 = section[m.end():]
        self.assertIn("PRECISION_KEEP_NARROW_COLUMN", q7)
        self.assertIn("PRECISION_KEEP_WIDTH_UNVERIFIED", q7)
        self.assertIn("column_width", q7)
        self.assertIn("PRECISION_KEEP_NARROW_COLUMN", self._error_code_table())

    def test_fillspec_capability_mapping_table(self):
        """能力映射表章节覆盖 MOD 规则类型 × 支持状态."""
        section = self._fillspec_section("能力映射表")
        for word in ("一等", "变通", "暂无"):
            self.assertIn(word, section)

    def test_fillspec_error_code_table_has_duplicate_target_write(self):
        """DUPLICATE_TARGET_WRITE 必须在「常见编译错误速查」表内 (防章节误删)."""
        self.assertIn("DUPLICATE_TARGET_WRITE", self._error_code_table())

    def test_fillspec_error_code_table_has_append_remove_zone(self):
        """REMOVE_TARGETS_APPEND_ZONE 必须在「常见编译错误速查」表内且行内
        首选语义为 append-only (inplace 仅是带样式时的条件选项, 防无条件
        指向 inplace 的误导)."""
        table = self._error_code_table()
        self.assertIn("REMOVE_TARGETS_APPEND_ZONE", table)
        m = re.search(r"^\|\s*REMOVE_TARGETS_APPEND_ZONE.*$", table,
                      re.MULTILINE)
        self.assertIsNotNone(m, "速查表缺 REMOVE_TARGETS_APPEND_ZONE 行")
        row = m.group(0)
        self.assertIn("append-only", row)
        self.assertIn("inplace", row)
        self.assertIn("样式", row)
        self.assertLess(row.index("append-only"), row.index("inplace"),
                        "append-only 是首选, inplace 只能是条件选项")

    def test_fillspec_pptx_capability_boundary(self):
        """「多目标与 PPTX」章节声明 pptx 支持矩阵 + 行边界缺陷码 (issue 06)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        m = re.search(r"^##\s+多目标与 PPTX", text, re.MULTILINE)
        self.assertIsNotNone(m, "FILLSPEC.md 缺章节 ## 多目标与 PPTX")
        section = text[m.end():]
        for word in ("PPTX_CAPABILITY_NOT_ROLLED_OUT",
                     "PPTX_TARGET_ROWS_OUT_OF_BOUNDS", "fail-closed",
                     "不再静默丢弃"):
            self.assertIn(word, section, f"多目标与 PPTX 缺词 {word!r}")
        self.assertIn("PPTX_TARGET_ROWS_OUT_OF_BOUNDS", self._error_code_table())
        self.assertIn("PPTX_CAPABILITY_NOT_ROLLED_OUT", self._error_code_table())

    def test_fillspec_q18_pptx_capability(self):
        """契约章节含 Q18 小节: pptx 支持矩阵与行边界 (issue 06;
        Q17 = issue 05 恰好一次, 撞号重排为 Q18)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q18:.*\n", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q18 小节")
        q18 = section[m.end():]
        self.assertIn("PPTX_CAPABILITY_NOT_ROLLED_OUT", q18)
        self.assertIn("PPTX_TARGET_ROWS_OUT_OF_BOUNDS", q18)
        self.assertIn("test_pptx_e2e.py", q18)

    def test_skill_md_pptx_support_matrix(self):
        """SKILL.md PPTX 小节声明支持矩阵 (issue 06) — frontmatter「任意方向」
        与能力边界不再矛盾 (静默丢弃措辞被 fail-closed 替代)."""
        text = self._skill_md_text()
        self.assertIn("PPTX_CAPABILITY_NOT_ROLLED_OUT", text)
        self.assertIn("不再静默丢弃", text)
        self.assertIn("PPTX_TARGET_ROWS_OUT_OF_BOUNDS", text)
        self.assertIn("支持矩阵 (issue 06 fail-closed)", text)

    def test_known_traps_pptx_silent_drop(self):
        """KNOWN_TRAPS 含「PPTX 未支持声明静默丢失」机械事实 (issue 06)."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(
            encoding="utf-8")
        self.assertIn("PPTX 未支持声明静默丢失", text)
        self.assertIn("PPTX_TARGET_ROWS_OUT_OF_BOUNDS", text)

    def test_skill_md_authoring_procedure_words(self):
        """SKILL.md 撰写规程: 先写后编译 + scratch 纪律词存在."""
        text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("先写后编译", text)
        self.assertIn("scratch", text)

    def test_known_traps_source_reading_discipline(self):
        """KNOWN_TRAPS 含源码阅读/实验纪律条目 (precision: keep 反例)."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        self.assertIn("precision: keep", text)
        self.assertIn("scratch", text)

    def _skill_md_text(self) -> str:
        return (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")

    def _capability_evidence_text(self) -> str:
        p = SKILL_ROOT / "references" / "CAPABILITY_EVIDENCE.md"
        self.assertTrue(p.is_file(), "缺 references/CAPABILITY_EVIDENCE.md")
        return p.read_text(encoding="utf-8")

    def _skill_routing_section(self) -> str:
        """SKILL.md §1.5 Task Shape Check (Routing V2) 区域 (防跨章节误伤)."""
        text = self._skill_md_text()
        m = re.search(r"^### 1\.5 Task Shape Check.*?(?=^### 2\. )",
                      text, re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "SKILL.md 缺 §1.5 Task Shape Check 段")
        return m.group(0)

    def _capability_evidence_section0(self) -> str:
        """CAPABILITY_EVIDENCE.md §0 (能力适用性 × 执行选择) 区域."""
        text = self._capability_evidence_text()
        m = re.search(r"^## 0\..*?(?=^## 1\.)", text,
                      re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "CAPABILITY_EVIDENCE.md 缺 §0 区域")
        return m.group(0)

    def _table_rows_after(self, text: str,
                          header_cells: tuple[str, ...]) -> list[list[str]]:
        """定位以指定表头单元格开头的 markdown 表格, 返回其数据行单元格
        列表 (剥掉反引号与首尾空白, 跳过 --- 分隔行); 表后第一个非表行停止."""
        header_re = r"^\|\s*" + r"\s*\|\s*".join(
            re.escape(h) for h in header_cells) + r"\s*\|"
        m = re.search(header_re, text, re.MULTILINE)
        self.assertIsNotNone(
            m, f"缺表头行: {' | '.join(header_cells)}")
        rows: list[list[str]] = []
        for line in text[m.end():].splitlines():
            line = line.rstrip()
            if not line:
                continue  # 表头后首个换行产生的空串, 跳过
            if not line.startswith("|"):
                break
            if re.fullmatch(
                    r"\|\s*:?-{2,}:?\s*(\|\s*:?-{2,}:?\s*)*\|", line):
                continue  # 分隔行
            cells = [c.strip().strip("`") for c in
                     line.strip().strip("|").split("|")]
            rows.append(cells)
        return rows

    def test_skill_md_capability_resolution_trigger(self):
        """SKILL.md 撰写规程: Capability Question 按需触发指针 + 三态 + 两项
        预算 + TASK MODE 禁区 + Run Verification 不可跳过 (防新停止规则被误删)."""
        text = self._skill_md_text()
        for word in ("Capability Question", "Known Supported", "Known Rejected",
                     "Capability Unknown", "Extra Capability Probe",
                     "Bounded Rescue", "CAPABILITY_EVIDENCE.md"):
            self.assertIn(word, text, f"SKILL.md 缺能力求证词 {word!r}")
        for word in ("不读实现源码", "测试套件", "不修改 Skill", "scratch"):
            self.assertIn(word, text, f"SKILL.md 缺 TASK MODE 禁区词 {word!r}")
        self.assertIn("Run Verification", text)

    def test_skill_md_rescue_formalities_pointer(self):
        """SKILL.md 预算行 (issue 03 rehearsal S6 发现): Rescue 使用前按
        CAPABILITY_EVIDENCE.md §4 合同执行 — 预声明 question/plan/verdict、
        workdir Run-local 记录、Gate 披露一句、无结论时 ASK/STOP 边界 —
        主 Skill 至少给出形式要求指针 (详细合同仍在按需参考, 不复制政策)."""
        text = self._skill_md_text()
        for word in ("预声明", "Run-local 记录", "Gate 披露", "未制度化",
                     "CAPABILITY_EVIDENCE.md", "ASK", "STOP"):
            self.assertIn(word, text, f"SKILL.md 缺 Rescue 形式指针词 {word!r}")

    def test_capability_evidence_reference_owns_policy(self):
        """CAPABILITY_EVIDENCE.md 是唯一详细 policy 源: 三态、Evidence Fit 与
        scope、封闭 Standard Evidence Paths 表、Probe 与 Rescue 合同、ASKS/STOP
        与 Run-local 证据 (防详细政策被拆散或章节被误删)."""
        text = self._capability_evidence_text()
        for word in ("Known Supported", "Known Rejected", "Capability Unknown",
                     "Evidence Fit", "Standard Evidence Path", "scope"):
            self.assertIn(word, text, f"CAPABILITY_EVIDENCE.md 缺词 {word!r}")
        for path_word in ("--capabilities", "formal compile", "officecli help",
                          "KNOWN_TRAPS", "readback", "结构验证", "Render QA"):
            self.assertIn(path_word, text, f"Standard Evidence Paths 缺 {path_word!r}")
        for probe_word in ("ACCEPTED", "REJECTED", "架构分叉"):
            self.assertIn(probe_word, text, f"Probe 合同缺 {probe_word!r}")
        for rescue_word in ("scratch", "Sufficient Evidence", "ASK", "STOP",
                            "Run-Local"):
            self.assertIn(rescue_word, text, f"Rescue 合同缺 {rescue_word!r}")

    def test_capability_evidence_reference_is_on_demand_pointer_target(self):
        """主 Skill 只放短算法与预算, 详细政策唯一住在 CAPABILITY_EVIDENCE.md:
        主 Skill 含按需加载声明 (正常 Run 不预读), 参考文件含五部分骨架."""
        skill = self._skill_md_text()
        self.assertIn("按需", skill)
        self.assertIn("不预读", skill)
        evidence = self._capability_evidence_text()
        self.assertIn("Capability Question", evidence)
        self.assertIn("Standard Evidence Paths", evidence)
        self.assertIn("Extra Capability Probe", evidence)
        self.assertIn("Bounded Rescue", evidence)

    # ── issue 04 (task-shape-routing) + issue 03 (routing-v2):
    #    路由术语一致性 + Routing V2 架构不变量守卫 ──

    def test_skill_md_task_shape_routing_terms(self):
        """SKILL.md 含 Routing V2 分流路由术语 (§1.5, tickets 01/03 落地):
        Task Shape Check / grid_record / form_content / mixed / uncertain /
        officecli_native / combined / obvious_grid / task_shape /
        task_shape.json / NOT_APPLICABLE — 防章节被误删或路由词被误改回 V1."""
        text = self._skill_md_text()
        for word in ("Task Shape Check", "grid_record", "form_content",
                     "mixed", "uncertain", "officecli_native", "combined",
                     "obvious_grid", "task_shape", "task_shape.json",
                     "NOT_APPLICABLE", "officecli native"):
            self.assertIn(word, text, f"SKILL.md 缺任务形态路由词 {word!r}")

    def test_capability_evidence_task_shape_matrix_terms(self):
        """CAPABILITY_EVIDENCE.md §0 重构为两张正交表 (issue 02/03, R2-Q6):
        能力语义表 (0.1) 恰好两行基础 workload, FillSpec Model 列 =
        APPLICABLE / NOT_APPLICABLE (grid 行 APPLICABLE, non-grid 行
        NOT_APPLICABLE; mixed 不占行); 路由决策表 (0.2) 四行且含 combined;
        旧式 1:1 单矩阵不得回归 — §0 无 Executor 列名、能力表行不绑
        fillspec/officecli_native 路由词."""
        section0 = self._capability_evidence_section0()
        # 旧职责适用部分: §0 仍携带两层语义词 (产品层 SUPPORTED / 引擎层
        # NOT_APPLICABLE / 路由 fillspec·officecli_native / Task Shape 联动)
        for word in ("Task Shape", "SUPPORTED", "NOT_APPLICABLE", "fillspec",
                     "officecli_native"):
            self.assertIn(word, section0,
                          f"CAPABILITY_EVIDENCE.md §0 缺两表语义词 {word!r}")
        # 旧式 1:1 矩阵回归禁止: 无 Executor 列名 (列曾把 shape 绑死执行器)
        self.assertNotIn("Executor", section0,
                         "§0 不得回归旧单矩阵的 Executor 列 (1:1 绑定)")
        # 0.1 能力语义表 code block = 表头 + 恰好两行
        m = re.search(r"### 0\.1.*?```text\n(.*?)```", section0, re.DOTALL)
        self.assertIsNotNone(m, "§0 缺 0.1 能力语义表 code block")
        block = m.group(1).strip().splitlines()
        self.assertEqual(len(block), 3,
                         "能力语义表必须为表头 + 恰好两行基础 workload")
        self.assertIn("FillSpec Model", block[0], "能力表缺 FillSpec Model 列")
        self.assertIn("table-fill Product", block[0])
        grid_row = block[1]
        self.assertIn("Grid / record transformation", grid_row)
        self.assertIn("APPLICABLE", grid_row,
                      "grid 行 FillSpec Model 列必须是 APPLICABLE (不是 SUPPORTED)")
        self.assertIn("SUPPORTED", grid_row, "grid 行产品层仍 SUPPORTED")
        self.assertNotIn("fillspec", grid_row,
                         "能力表行不得绑执行路由 (旧 1:1 矩阵写法)")
        non_grid_row = block[2]
        self.assertIn("Non-grid Office operation", non_grid_row)
        self.assertIn("NOT_APPLICABLE", non_grid_row,
                      "non-grid 行 FillSpec Model 列必须是 NOT_APPLICABLE")
        self.assertIn("SUPPORTED", non_grid_row)
        self.assertNotIn("officecli_native", non_grid_row,
                         "能力表行不得绑执行路由 (旧 1:1 矩阵写法)")
        self.assertIn("不占行", section0,
                      "§0 必须声明 mixed 不占能力表行")
        # 0.2 路由决策表: 四行 (Fast Path / Direct / Non-Grid / Combined)
        rows = self._table_rows_after(
            section0, ("Workload situation", "route", "典型 evidence"))
        self.assertEqual(len(rows), 4, "路由决策表必须恰好四行")
        routes = [r[1] for r in rows]
        self.assertIn("combined", routes,
                      "路由决策表必须含 combined (第 4 行 Combined 分支)")
        self.assertTrue(all(r in ("fillspec", "officecli_native", "combined")
                            for r in routes),
                        f"路由决策表 route 值越界: {routes!r}")

    def test_capability_evidence_form_content_not_unsupported_drift_guard(self):
        """防措辞回退 (issue 02/03, R2-Q6): form_content 不再占能力表行,
        但任何把 form_content 标成 UNSUPPORTED / Known Rejected 的语境仍是
        红线 — 对 §0 中含 form_content 的每一行做行作用域负断言 (0.6 纪律句
        本身含反例词但不同现一行, 防自伤); 且 0.6 措辞纪律声明必须存在:
        NOT_APPLICABLE ≠ UNSUPPORTED / ≠ Known Rejected."""
        section0 = self._capability_evidence_section0()
        for line in section0.splitlines():
            if "form_content" not in line:
                continue
            self.assertNotIn("UNSUPPORTED", line,
                             f"含 form_content 的行不得标 UNSUPPORTED: {line!r}")
            self.assertNotIn("Known Rejected", line,
                             f"含 form_content 的行不得标 Known Rejected: {line!r}")
        m = re.search(r"### 0\.6 措辞纪律", section0)
        self.assertIsNotNone(m, "§0 缺 0.6 措辞纪律段")
        stripped = re.sub(r"[\s`]", "", section0[m.end():])
        self.assertIn("NOT_APPLICABLE≠UNSUPPORTED", stripped,
                      "0.6 缺 NOT_APPLICABLE ≠ UNSUPPORTED 声明")
        self.assertIn("NOT_APPLICABLE≠KnownRejected", stripped,
                      "0.6 缺 NOT_APPLICABLE ≠ Known Rejected 声明")

    def test_skill_md_routing_shape_domain_four_values(self):
        """§1.5 值域表 shape 维度恰好四值: grid_record / form_content /
        mixed / uncertain (R2-Q1: shape 新增且仅新增 mixed, V1 三态 → V2
        四态; direct 永不作为 shape)."""
        section = self._skill_routing_section()
        rows = self._table_rows_after(
            section, ("task_shape", "含义", "合法 route", "典型 evidence"))
        self.assertEqual(len(rows), 4, "task_shape 值域表必须恰好四行")
        shapes = {r[0] for r in rows}
        self.assertEqual(
            shapes, {"grid_record", "form_content", "mixed", "uncertain"},
            "task_shape 值域表缺值或混入非法 shape 值")
        self.assertIn("永不作为 shape", section,
                      "§1.5 缺 direct 永不作为 shape 锚点句")

    def test_skill_md_routing_route_domain_three_values_no_hybrid(self):
        """§1.5 route 值域恰好三值: fillspec / officecli_native / combined
        (R2-Q1: 用 combined 而非 hybrid); 值域表 '合法 route' 列各分支绑定
        正确 (grid_record → fillspec/officecli_native, form_content →
        officecli_native, mixed → combined, uncertain 不落 route);
        hybrid 不得作为 route 值出现在 §1.5 路由语境."""
        section = self._skill_routing_section()
        rows = self._table_rows_after(
            section, ("task_shape", "含义", "合法 route", "典型 evidence"))
        by_shape = {r[0]: r[2] for r in rows}
        self.assertIn("fillspec", by_shape["grid_record"],
                      "grid_record 合法 route 缺 fillspec (Fast Path)")
        self.assertIn("officecli_native", by_shape["grid_record"],
                      "grid_record 合法 route 缺 officecli_native (Direct)")
        self.assertEqual(by_shape["form_content"], "officecli_native",
                         "form_content 合法 route 必须是 officecli_native")
        self.assertEqual(by_shape["mixed"], "combined",
                         "mixed 合法 route 必须是 combined (改回 hybrid 变红)")
        self.assertIn("不落执行", by_shape["uncertain"],
                      "uncertain 是临时判定态, 不落执行 route")
        stripped = re.sub(r"[\s`]", "", section)
        self.assertIn("route值域仅fillspec/officecli_native/combined",
                      stripped, "§1.5 缺 route 值域仅三值定义句")
        self.assertNotIn("hybrid", stripped,
                         "hybrid 不得作为 route 值出现在 §1.5 路由语境")

    def test_capability_evidence_route_domain_three_values_no_hybrid(self):
        """§0.2 路由决策表 route 值域恰好三值: fillspec / officecli_native /
        combined (R2-Q1); route 定义句声明仅三值; 0.6 消歧陈述 (combined ≠
        hybrid, 点名 FILLSPEC "hybrid overflow") 是 hybrid 唯一合法语境 —
        路由决策表行与定义句不得含 hybrid (最小作用域, 防误伤消歧句)."""
        section0 = self._capability_evidence_section0()
        rows = self._table_rows_after(
            section0, ("Workload situation", "route", "典型 evidence"))
        routes = {r[1] for r in rows}
        self.assertEqual(
            routes, {"fillspec", "officecli_native", "combined"},
            "路由决策表 route 值域必须恰为 fillspec / officecli_native / combined")
        rows_text = "\n".join("|".join(r) for r in rows)
        self.assertNotIn("hybrid", rows_text,
                         "路由决策表行不得出现 hybrid (消歧句在 0.6, 不在此域)")
        stripped = re.sub(r"[\s`]", "", section0)
        self.assertIn("route值域仅fillspec/officecli_native/combined",
                      stripped, "§0 缺 route 值域仅三值定义句")
        # 0.6 消歧陈述存在 — hybrid 在路由文档的唯一合法语境被保留
        m = re.search(r"### 0\.6 措辞纪律", section0)
        self.assertIsNotNone(m, "§0 缺 0.6 措辞纪律段")
        disc = re.sub(r"[\s`]", "", section0[m.end():])
        self.assertIn("combined≠hybrid", disc,
                      "0.6 缺 combined ≠ hybrid 消歧陈述")
        self.assertIn("hybridoverflow", disc,
                      "0.6 消歧陈述须点名 FILLSPEC hybrid overflow")

    def test_skill_md_routing_legal_shape_route_combinations(self):
        """防 shape×route 1:1 绑定回退 (R2-Q1/Q3/Q4): 路由分流 ASCII 块同时
        含 Direct 组合 grid_record + officecli_native 与 Combined 组合
        mixed + combined; §1.5 声明两维度不再 1:1 绑定 (删任一合法组合分支
        或改回 1:1 矩阵时变红)."""
        section = self._skill_routing_section()
        ascii_m = re.search(r"```text\n(.*?)```", section, re.DOTALL)
        self.assertIsNotNone(ascii_m, "§1.5 缺路由分流 ASCII 块")
        stripped = re.sub(r"[\s`]", "", ascii_m.group(1))
        self.assertIn("grid_record+officecli_native", stripped,
                      "路由 ASCII 缺 Direct 组合 (grid_record + officecli_native)")
        self.assertIn("mixed+combined", stripped,
                      "路由 ASCII 缺 Combined 组合 (mixed + combined)")
        self.assertIn("form_content+officecli_native", stripped,
                      "路由 ASCII 缺 Non-Grid 组合 (form_content + officecli_native)")
        whole = re.sub(r"[\s`]", "", section)
        self.assertIn("不再1:1绑定", whole,
                      "§1.5 缺 shape×route 正交声明 (不再 1:1 绑定)")

    def test_capability_evidence_applicability_not_justification(self):
        """§0 声明 Applicability ≠ Justification (issue 02/03, R2-Q6 入文):
        0.3 正式定义段同时给出 Applicability (适用性) 与 Justification
        (启动理由), 并含 '不相等' 反例示例 (3~5 固定 cell: APPLICABLE 但
        NOT JUSTIFIED → Direct) — 防适用性/正当性被揉回单一矩阵."""
        section0 = self._capability_evidence_section0()
        m = re.search(
            r"### 0\.3 Applicability ≠ Justification.*?(?=^### 0\.4)",
            section0, re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(
            m, "§0 缺 0.3 Applicability ≠ Justification 正式定义段")
        sec = m.group(0)
        for word in ("Applicability", "Justification", "适用性", "启动理由"):
            self.assertIn(word, sec, f"0.3 定义段缺 {word!r}")
        stripped = re.sub(r"[\s`]", "", section0)
        self.assertIn("Applicability≠Justification", stripped,
                      "§0 缺 Applicability ≠ Justification 声明")
        self.assertIn("NOTJUSTIFIED", stripped,
                      "0.3 缺 NOT JUSTIFIED 反例示例")

    def test_skill_md_obvious_grid_fast_path(self):
        """Level 0 Obvious Grid Fast Path 锚点 (issue 01/03, R2-Q2):
        §1.5 含 evidence code obvious_grid、Obvious Grid 分支与 '立即进 MOD'
        stop-rule 词 (删 Fast Path evidence 或整段时变红)."""
        section = self._skill_routing_section()
        stripped = re.sub(r"[\s`]", "", section)
        self.assertIn("obvious_grid", stripped,
                      "§1.5 缺 obvious_grid evidence code (Fast Path)")
        self.assertIn("ObviousGrid", stripped, "§1.5 缺 Obvious Grid 分支")
        self.assertIn("FastPath", stripped, "§1.5 缺 Fast Path 表述")
        self.assertIn("立即进MOD", stripped,
                      "§1.5 缺立即进 MOD stop-rule 词")

    def test_skill_md_combined_final_gate_order(self):
        """Combined 最小契约 (issue 01/03, R2-Q4): 单一 Final Gate 延后至
        全部写操作完成 — 契约块含 单一 Final Gate, OfficeCLI finishing 步骤
        在 Gate 之前; 正文声明 'officecli finishing 在 Gate 之前执行'.
        (finishing 与 Gate 的相对顺序被颠倒时变红.)"""
        section = self._skill_routing_section()
        m = re.search(r"#### Combined 最小契约.*?```text\n(.*?)```",
                      section, re.DOTALL)
        self.assertIsNotNone(m, "§1.5 缺 Combined 最小契约块")
        block = re.sub(r"[\s`]", "", m.group(1))
        self.assertIn("单一FinalGate", block,
                      "Combined 契约缺 单一 Final Gate")
        self.assertIn("OfficeCLIfinishing", block,
                      "Combined 契约缺 OfficeCLI finishing 步骤")
        self.assertLess(block.index("OfficeCLIfinishing"),
                        block.index("单一FinalGate"),
                        "finishing 必须先于 Final Gate (单一 Final Gate 延后至全部写操作完成)")
        stripped = re.sub(r"[\s`]", "", section)
        self.assertIn("finishing在Gate之前执行", stripped,
                      "§1.5 缺 finishing 在 Gate 之前执行的顺序关键句")

    def test_old_universal_probe_and_escape_hatch_removed(self):
        """旧流程被原子替换 (migrate facts, replace process): 普遍 probe、
        TASK MODE 源码/spike escape hatch 与任务内强制制度化措辞不再存在,
        不以优先级说明保留并行旧流程. 三件套完整性标准 (缺一视为未完成)
        属于 Skill Development 制度化标准, 不在此禁止之列."""
        skill = self._skill_md_text()
        for phrase in ("probe = 唯一允许的确认手段", "源码/spike 边界",
                       "才读源码或做", "必须强制转换"):
            self.assertNotIn(phrase, skill, f"SKILL.md 仍含旧流程措辞 {phrase!r}")
        fillspec = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        for phrase in ("不确定 → `compile_fill.py --probe`",
                       "probe/编译报错无法解释"):
            self.assertNotIn(phrase, fillspec, f"FILLSPEC.md 仍含旧流程措辞 {phrase!r}")
        traps = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(
            encoding="utf-8")
        for phrase in ("仅「文档未覆盖 且 报错无法解释」才读源码",):
            self.assertNotIn(phrase, traps, f"KNOWN_TRAPS.md 仍含旧流程措辞 {phrase!r}")
        catalog = (SKILL_ROOT / "assets" / "combination_patterns.yaml").read_text(
            encoding="utf-8")
        self.assertNotIn("仍不确定 → compile_fill.py --probe", catalog,
                         "combination_patterns.yaml 仍含普遍 probe 指令")
        self.assertIn("CAPABILITY_EVIDENCE.md", catalog,
                      "combination_patterns.yaml 缺能力政策按需指针")

    def test_skill_md_distrust_recorded_not_institutionalized(self):
        """不信任事件纪律: TASK MODE 只记录 (Capability Gap Discovery /
        Contract Drift), 三件套制度化只属于用户发起的 Skill Development
        (旧'当前任务内必须强制转换'被替换, 触发条件与契约漂移保留)."""
        text = self._skill_md_text()
        for word in ("不信任事件", "契约漂移", "触发条件", "三件套", "Skill Development"):
            self.assertIn(word, text, f"SKILL.md 缺不信任事件纪律词 {word!r}")
        m = re.search(r"最高优先[^\n]*", text)
        self.assertIsNotNone(m, "缺最高优先触发条件声明")
        self.assertIn("契约漂移", m.group(0))
        self.assertIn("needs-triage", text)
        self.assertNotIn("必须强制转换", text)

    def test_skill_md_probe_and_capabilities_tools_mentioned(self):
        """SKILL.md 能力求证段保留工具名 (事实): --probe / --capabilities /
        make_probe_spec.py 仍在, 但角色是预算内的架构分叉工具, 不是普遍
        确认手段."""
        text = self._skill_md_text()
        for word in ("--probe", "--capabilities", "make_probe_spec.py"):
            self.assertIn(word, text, f"SKILL.md 缺工具名 {word!r}")

    def test_combination_patterns_exist(self):
        """assets/combination_patterns.yaml 存在且含关键模式 (防模板被误删)."""
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        self.assertTrue(p.is_file(), "缺 assets/combination_patterns.yaml")
        text = p.read_text(encoding="utf-8")
        self.assertIn("group_merges", text)
        self.assertIn("round4", text)
        self.assertIn("per_group_total_explicit_ranges", text)

    def test_fillspec_q8_anchor_scope_words(self):
        """契约章节含 Q8 小节: data role 作用域 + title/header 无检查边界."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q8:.*\n", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q8 小节")
        q8 = section[m.end():]
        self.assertIn("data", q8)
        self.assertIn("title/header", q8)

    def test_fillspec_q9_q10_sections(self):
        """契约章节含 Q9 (value 延迟写入) 与 Q10 (readback 种类) 小节."""
        section = self._fillspec_section("组合行为契约")
        self.assertIsNotNone(re.search(r"^### Q9:", section, re.MULTILINE), "缺 Q9")
        self.assertIsNotNone(re.search(r"^### Q10:", section, re.MULTILINE), "缺 Q10")
        self.assertIn("deferred_values", section)
        self.assertIn("nonempty", section)

    def test_skill_md_failure_cost_quantified(self):
        """SKILL.md 撰写规程量化失败成本: 第 1 轮失败是预期路径, 修复 <2 分钟,
        预算约束连续失败而非单次失败 (消除'怕失败读源码'的动机)."""
        text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("预期路径", text)
        self.assertIn("2 分钟", text)
        self.assertIn("连续失败", text)

    def test_fillspec_q11_q12_sections(self):
        """契约章节含 Q11 (克隆携带合并) 与 Q12 (merges×aggregates/多组聚合)."""
        section = self._fillspec_section("组合行为契约")
        self.assertIsNotNone(re.search(r"^### Q11:", section, re.MULTILINE), "缺 Q11")
        self.assertIsNotNone(re.search(r"^### Q12:", section, re.MULTILINE), "缺 Q12")
        self.assertIn("mergeCell", section)
        self.assertIn("显式范围", section)

    def test_fillspec_q13_per_group_total_boundary(self):
        """契约章节含 Q13: 每组合计接受边界 — 聚合列不进 nulls 是触发因素
        (负面表达与通过形态同源, fixture 漂移以 Q&A 文字锁定)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q13:", section, re.MULTILINE)
        self.assertIsNotNone(m, "缺 Q13")
        q13 = section[m.end():]
        self.assertIn("nulls", q13)
        self.assertIn("DUPLICATE_TARGET_WRITE", q13)
        self.assertIn("first as empty", q13)
        self.assertIn("per_group_total_explicit_ranges", q13)

    def test_fillspec_q14_group_aggregates_section(self):
        """契约章节含 Q14 小节 (group_aggregates 一等能力 + whole_run 门)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q14:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q14 小节")
        q14 = section[m.end():]
        self.assertIn("group_aggregates", q14)
        self.assertIn("group_by", q14)
        self.assertIn("CAPABILITY_NOT_ROLLED_OUT", q14)
        self.assertIn("DUPLICATE_TARGET_WRITE", q14)

    def test_fillspec_capability_table_group_aggregates_first_class(self):
        """能力映射表「每组合计」→ 一等 (group_aggregates 表达)."""
        section = self._fillspec_section("能力映射表")
        self.assertIn("group_aggregates", section)
        self.assertIn("一等", section)
        self.assertIn("CAPABILITY_NOT_ROLLED_OUT", section)

    def test_fillspec_error_code_table_has_capability_gate(self):
        """CAPABILITY_NOT_ROLLED_OUT 在「常见编译错误速查」表内 (whole_run 门)."""
        self.assertIn("CAPABILITY_NOT_ROLLED_OUT", self._error_code_table())

    def test_combination_patterns_group_aggregates_pattern(self):
        """combination_patterns.yaml 含 group_aggregates 一等模式 (改列名即可)."""
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        text = p.read_text(encoding="utf-8")
        self.assertIn("group_aggregates", text)
        self.assertIn("group_by", text)

    # ── issue 02: 完整 Canonical Pattern (preformatted_quotation_inplace) ──

    def test_combination_patterns_preformatted_quotation_inplace_entry(self):
        """Catalog 含唯一 preformatted_quotation_inplace entry, 命名与问题描述
        让 Agent 在预格式报价模板场景可定位; 同一条目 (fragment) 同时携带全部
        结构职责词 — 不退化为独立 feature snippets 的链接列表 (只守词, 不复制
        整份 skeleton)."""
        import yaml
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        entries = yaml.safe_load(p.read_text(encoding="utf-8"))["patterns"]
        hits = [e for e in entries if e["id"] == "preformatted_quotation_inplace"]
        self.assertEqual(len(hits), 1, "pattern id 必须唯一")
        entry = hits[0]
        for key in ("id", "question", "answer", "fragment", "note"):
            self.assertIn(key, entry, f"entry 缺字段 {key!r}")
        self.assertIn("报价", entry["question"], "问题描述应可定位预格式报价场景")
        for word in ("mode: inplace", "start_row", "capacity", "template_row",
                     "group_merges", "label", "sets", "numberformat",
                     "key_outputs"):
            self.assertIn(word, entry["fragment"],
                          f"fragment 缺结构职责词 {word!r}")
        # Trim/overflow 是编译器从 inplace 几何推导的后果, 由 answer/note 表达
        # (fragment 只声明几何); 文档守卫同时守住这段职责说明
        for word in ("Trim", "overflow"):
            self.assertIn(word, entry["answer"] + entry["note"],
                          f"entry 缺 {word} 职责说明")
        # 结构决策完整保留在同一 fragment (含几何与两类 merge 与绝对写),
        # 而不是只给链接列表
        self.assertIn("base_last_row", entry["fragment"])
        self.assertIn("group_by", entry["fragment"])
        self.assertIn("path:", entry["fragment"])

    def test_combination_patterns_preformatted_quotation_data_neutral(self):
        """MXP 单次业务事实不泄漏进 fragment: 客户/sheet/行号/价格/lookup/文案
        一律参数化为占位标记 (含行号: region/trim 数字全部在替换表里)."""
        import yaml
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        entry = next(e for e in yaml.safe_load(p.read_text(encoding="utf-8"))["patterns"]
                     if e["id"] == "preformatted_quotation_inplace")
        frag = entry["fragment"]
        for leaked in ("MXP", "ATLAS", "Algeria", "To Messrs", "2026",
                       "source_17", "158", "5035"):
            self.assertNotIn(leaked, frag, f"fragment 泄漏 MXP 单次事实 {leaked!r}")
        for token in ("<TARGET_SHEET>", "<BASE_LAST_ROW>", "<REGION_START>",
                      "<REGION_CAPACITY>", "<TEMPLATE_ROW>", "<SOURCE_NAME>",
                      "<NUMFMT>", "<HDR_ADDR>", "<FOOT_ADDR>", "<KO1>"):
            self.assertIn(token, frag, f"fragment 缺参数占位符 {token}")
        # 明确是推荐构造路径, 不是 whitelist / cross-Run support claim
        text = p.read_text(encoding="utf-8")
        self.assertIn("构造路径", text)
        self.assertIn("白名单", text)

    def test_combination_patterns_canonical_admission_boundary(self):
        """完整 Canonical Pattern 准入边界: 真实任务 + 通过的 Validated Draft +
        显著减少决策; 局部 group_aggregates Pattern 保留; 不拼接未验证的
        append+group_aggregates 完整骨架; 无 Candidate/Provisional 层级."""
        import yaml
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        text = p.read_text(encoding="utf-8")
        for word in ("真实任务", "Validated Draft", "显著减少"):
            self.assertIn(word, text, f"catalog 头部缺准入词 {word!r}")
        ids = [e["id"] for e in yaml.safe_load(text)["patterns"]]
        self.assertIn("per_group_total_group_aggregates", ids,
                      "局部 group_aggregates Pattern 必须保留")
        self.assertIn("preformatted_quotation_inplace", ids)
        self.assertNotIn("append_group_aggregates", ids,
                         "未验证的 append+group_aggregates 完整骨架不得入库")
        self.assertFalse(any("candidate" in i.lower() or "provisional" in i.lower()
                             for i in ids),
                         "不得新增 Candidate/Provisional Pattern 层级")

    # ── issue 03: 完整 Canonical Pattern (multiproduct_block_append) ──

    def _multiproduct_entry(self) -> dict:
        import yaml
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        entries = yaml.safe_load(p.read_text(encoding="utf-8"))["patterns"]
        hits = [e for e in entries if e["id"] == "multiproduct_block_append"]
        self.assertEqual(len(hits), 1, "pattern id 必须唯一")
        return hits[0]

    def test_combination_patterns_multiproduct_block_entry(self):
        """Catalog 含唯一 multiproduct_block_append entry, 问题描述可定位家用/商用
        双数据块追加场景; 同一 fragment 携带全部结构职责词 (clone_roles /
        group_merges / nulls / merges / aggregates / key_outputs), note 明确
        U4 机械事实 (不自动建合并区 / 显式 merges 覆盖非锚点残留)."""
        entry = self._multiproduct_entry()
        for key in ("id", "question", "answer", "fragment", "note"):
            self.assertIn(key, entry, f"entry 缺字段 {key!r}")
        for word in ("家用", "商用", "多产品"):
            self.assertIn(word, entry["question"] + entry["answer"],
                          f"问题描述缺定位词 {word!r}")
        for word in ("clone_roles", "group_merges", "nulls", "merges",
                     "aggregates", "key_outputs", "1:{n}"):
            self.assertIn(word, entry["fragment"],
                          f"fragment 缺结构职责词 {word!r}")
        self.assertIn("label", entry["fragment"])
        # U4 机械事实必须同时落在 answer 与 note (供 Agent 定位与契约固化)
        for sect in (entry["answer"], entry["note"]):
            self.assertIn("不自动创建合并区", sect,
                          f"{sect} 缺「不自动创建合并区」词")
            self.assertIn("显式 merges", sect,
                          f"{sect} 缺「显式 merges」词")
        # note 指明 deviating 变体指向既有局部片段 (块内多组显式范围)
        self.assertIn("per_group_total_explicit_ranges", entry["note"])

    def test_combination_patterns_multiproduct_block_data_neutral(self):
        """真实任务业务事实 (客户/sheet/型号/价格/行号) 不泄漏进 fragment:
        全部参数化为占位标记; 组范围/总盈亏范围用占位符 + 1:{n}, 不写死行数."""
        entry = self._multiproduct_entry()
        frag = entry["fragment"]
        for leaked in ("家用悦风", "清爽星", "Z001", "11_FRESH", "毛利表",
                       "105000", "ATLAS", "埃及"):
            self.assertNotIn(leaked, frag, f"fragment 泄漏业务事实 {leaked!r}")
        for token in ("<TARGET_SHEET>", "<BASE_LAST_ROW>", "<SOURCE_HOUSE>",
                      "<SOURCE_COMMERCIAL>", "<TITLE_HOUSE>", "<GROUP_COL>",
                      "<LABEL_COL>", "<NULL_COL_1>", "<AGG_COL_V>",
                      "<TOTAL_COL_W>", "<V_ROWS_H1>", "<V_ROWS_C1>", "<KO1>"):
            self.assertIn(token, frag, f"fragment 缺参数占位符 {token}")
        # 每源分组 V 显式范围 + 总盈亏 W 1:{n} 的可复制声明在 answer/note
        for phrase in ("每源分组一条 V", "总盈亏 W 一条 1:{n}"):
            self.assertIn(phrase, entry["answer"] + entry["note"],
                          f"缺 {phrase}")
        # 明确是推荐构造路径, 不是 whitelist / cross-Run support claim
        self.assertIn("构造路径", entry["note"])

    # ── issue 02: 完整 Canonical Pattern (single_quotation_block_append) ──

    def _single_block_entry(self) -> dict:
        import yaml
        p = SKILL_ROOT / "assets" / "combination_patterns.yaml"
        entries = yaml.safe_load(p.read_text(encoding="utf-8"))["patterns"]
        hits = [e for e in entries if e["id"] == "single_quotation_block_append"]
        self.assertEqual(len(hits), 1, "pattern id 必须唯一")
        return hits[0]

    def test_combination_patterns_single_quotation_block_entry(self):
        """Catalog 含唯一 single_quotation_block_append entry, 问题描述可定位
        核价/报价汇总单块 append 场景; 同一 fragment 携带全部结构职责词
        (clone_roles / selectors / 0-口径 / nulls / merges / aggregates /
        key_outputs), note 锁定 U1 表头行契约与单块同构、聚合列不进
        nulls/group_merges、[1] 外部引用列直接 null."""
        entry = self._single_block_entry()
        for key in ("id", "question", "answer", "fragment", "note"):
            self.assertIn(key, entry, f"entry 缺字段 {key!r}")
        for word in ("核价", "报价汇总", "单块"):
            self.assertIn(word, entry["question"] + entry["answer"],
                          f"问题描述缺定位词 {word!r}")
        for word in ("clone_roles", "selectors", "columns", "formulas",
                     "per_row", "aggregates", "merges", "nulls", "key_outputs"):
            self.assertIn(word, entry["fragment"],
                          f"fragment 缺结构职责词 {word!r}")
        # 完整骨架必须显式 selectors 排除表头 (U1) — 词落在 fragment 与 note
        for sect in (entry["answer"], entry["note"]):
            self.assertIn("排除表头行", sect, f"{sect} 缺「排除表头行」词")
        self.assertIn("HEADER_ROW_CONSIDERED_DATA", entry["note"],
                      "note 缺表头行守卫缺陷码词")
        # 单块同构 + 聚合列不变量 + [1] 外部引用机械事实
        for word in ("multiproduct_block_append", "单块同构"):
            self.assertIn(word, entry["note"], f"note 缺 {word!r} 词")
        for word in ("不进 nulls", "不进 group_merges"):
            self.assertIn(word, entry["note"], f"note 缺「{word}」词")
        for word in ("[1]", "外部工作簿引用", "formula_not_evaluated"):
            self.assertIn(word, entry["note"] + entry["answer"],
                          f"note/answer 缺 [1] 外部引用词 {word!r}")
        self.assertIn("0-口径", entry["answer"] + entry["note"])
        self.assertIn("ROUND", entry["answer"] + entry["note"])

    def test_combination_patterns_single_quotation_block_data_neutral(self):
        """真实任务业务事实 (客户/国家/型号/价格/行号) 不泄漏进 fragment:
        全部参数化为占位标记; 只写一块不写死行数占位 (范围用 1:{n} / 占位符)."""
        entry = self._single_block_entry()
        frag = entry["fragment"]
        for leaked in ("MXP", "ATLAS", "阿尔及利亚", "一拖多外机", "Z码",
                       "105000", "8/", "2026-08-18"):
            self.assertNotIn(leaked, frag, f"fragment 泄漏业务事实 {leaked!r}")
        for token in ("<TARGET_SHEET>", "<BASE_LAST_ROW>", "<TITLE>",
                      "<SOURCE_NAME>", "<GROUP_COL>", "<MODEL_COL>",
                      "<COST_COL>", "<CMD_COL>", "<SELECTOR_COL>",
                      "<SELECTOR_PATTERN>", "<TOTAL_COL>",
                      "<NULL_COL_1>", "<NULL_COL_2>", "<NULL_COL_3>", "<KO1>"):
            self.assertIn(token, frag, f"fragment 缺参数占位符 {token}")
        # 明确是推荐构造路径, 不是 whitelist / cross-Run support claim
        self.assertIn("构造路径", entry["note"])
        self.assertIn("白名单", entry["note"])

    def test_fillspec_selectors_header_row_contract_word(self):
        """FILLSPEC selectors 段含表头行契约词 (issue 02 / Case 08 U1):
        「展平 CSV 首行（表头）是候选数据行」+ 「无 selector 会把表头映射进数据区」+
        pattern/not_pattern 排除示例 + HEADER_ROW_CONSIDERED_DATA 警告."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(encoding="utf-8")
        m = re.search(r"^### selectors\n", text, re.MULTILINE)
        self.assertIsNotNone(m, "FILLSPEC.md 缺 ### selectors 小节")
        nxt = re.search(r"^### ", text[m.end():], re.MULTILINE)
        section = text[m.end():m.end() + (nxt.start() if nxt else len(text))]
        for word in ("展平 CSV 首行", "候选数据行", "表头", "pattern", "not_pattern",
                     "HEADER_ROW_CONSIDERED_DATA"):
            self.assertIn(word, section, f"selectors 段缺契约词 {word!r}")

    def test_error_code_table_has_header_row_considered_data(self):
        """HEADER_ROW_CONSIDERED_DATA 必须在「常见编译错误速查」表内, corrective
        action 指向 pattern/not_pattern 排除表头行 (防章节误删/指引回退)."""
        table = self._error_code_table()
        self.assertIn("HEADER_ROW_CONSIDERED_DATA", table)
        m = re.search(r"^\|\s*HEADER_ROW_CONSIDERED_DATA.*$", table, re.MULTILINE)
        self.assertIsNotNone(m, "速查表缺 HEADER_ROW_CONSIDERED_DATA 行")
        row = m.group(0)
        self.assertIn("排除表头行", row)
        self.assertIn("pattern", row)
        self.assertIn("not_pattern", row)

    def test_skill_md_single_block_minimal_literature(self):
        """SKILL 含本场景最小文献面清单 (issue 02 / Case 08 R3 收敛): 单块核价/
        报价块 append 只读 决策树 + Q5/Q8/Q12/Q19 + ROUND 精准 + 单块骨架,
        不全文通读 FILLSPEC."""
        text = self._skill_md_text()
        for word in ("最小文献面", "单块", "布局决策树", "Q5", "Q8", "Q12", "Q19",
                     "ROUND", "single_quotation_block_append", "全文通读"):
            self.assertIn(word, text, f"SKILL.md 缺最小文献面词 {word!r}")

    def test_fillspec_q19_aggregate_merge_region_word(self):
        """契约章节含 Q19: aggregates/group_aggregates 不自动创建合并区,
        聚合列非锚点残留需显式 merges 覆盖 (U4 组合空缺, issue 03)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q19:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q19 小节")
        q19 = section[m.end():]
        for word in ("不自动创建合并区", "显式", "merges", "非锚点残留",
                     "multiproduct_block_append", "DUPLICATE_TARGET_WRITE",
                     "1:{n}"):
            self.assertIn(word, q19, f"Q19 缺词 {word!r}")

    def test_known_traps_aggregate_merge_region_fact(self):
        """KNOWN_TRAPS 沉淀 U4 机械事实双件: ① block 顶层聚合键静默丢弃 → 编译期
        拒绝 (issue 01); ② group_aggregates/aggregates 不自动建合并区, 聚合列
        非锚点残留需显式 merges 覆盖 + 指向 multiproduct_block_append 骨架."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(
            encoding="utf-8")
        m = re.search(r"^\|\s*\*\*block 顶层聚合键静默丢弃.*$", text, re.MULTILINE)
        self.assertIsNotNone(m, "KNOWN_TRAPS 缺 block 顶层聚合键静默丢弃 行")
        self.assertIn("BLOCK_KEY_STRUCTURE_INVALID", m.group(0))
        m2 = re.search(r"^\|\s*\*\*aggregates/group_aggregates 不自动建合并区.*$",
                       text, re.MULTILINE)
        self.assertIsNotNone(m2, "KNOWN_TRAPS 缺 aggregates 不自动建合并区 行")
        row = m2.group(0)
        for word in ("非锚点残留", "显式", "merges", "multiproduct_block_append",
                     "DUPLICATE_TARGET_WRITE"):
            self.assertIn(word, row, f"U4 机械事实行缺 {word!r}")

    def test_known_traps_spike_facts(self):
        """KNOWN_TRAPS 沉淀已 spike 机械事实 (克隆携带合并 / merges×aggregates)."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        self.assertIn("克隆携带合并", text)
        self.assertIn("mergeCell", text)
        self.assertIn("A41:F41", text)

    # ── issue 03: MERGE_MODE_CONFLICT 指引 + [1] 外部引用 + 探索终止硬化 ──

    def test_fillspec_q19_aggregate_column_not_in_group_merges(self):
        """Q19 补「聚合列不进 group_merges」单一 owner 不变量 + MERGE_MODE_CONFLICT
        正确修复 (删 group_merges、用同范围 merges+aggregates 对, 聚合锚点=合并锚点)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q19:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q19 小节")
        q19 = section[m.end():]
        for word in ("不进 group_merges", "MERGE_MODE_CONFLICT", "merges + aggregates",
                     "聚合锚点=合并锚点", "不绑定 pattern 名"):
            self.assertIn(word, q19, f"Q19 缺聚合列×group_merges 指引词 {word!r}")

    def test_fillspec_q1_aggregate_single_owner_extension(self):
        """Q1 规则补「聚合列不进 group_merges」— 与「聚合列不进 nulls」并列的单一
        owner 不变量 (MERGE_MODE_CONFLICT 属同列双写)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q1:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q1 小节")
        q1 = section[m.end():]
        for word in ("聚合列不进", "group_merges", "MERGE_MODE_CONFLICT", "不进 nulls"):
            self.assertIn(word, q1, f"Q1 规则缺聚合列不进 group_merges 词 {word!r}")

    def test_error_code_table_merge_mode_conflict_wording(self):
        """MERGE_MODE_CONFLICT 速查表行 corrective_action 指向正确组合: 聚合列 →
        删 group_merges + 同范围 merges+aggregates 对; 普通标签列 → 保留一种模式
        (防指引回退到「每列只用一种合并模式」)."""
        table = self._error_code_table()
        m = re.search(r"^\|\s*MERGE_MODE_CONFLICT.*$", table, re.MULTILINE)
        self.assertIsNotNone(m, "速查表缺 MERGE_MODE_CONFLICT 行")
        row = m.group(0)
        for word in ("group_merges", "merges + aggregates", "聚合锚点=合并锚点",
                     "普通标签列", "保留一种", "聚合"):
            self.assertIn(word, row, f"MERGE_MODE_CONFLICT 行缺 {word!r}")

    def test_known_traps_external_workbook_reference(self):
        """KNOWN_TRAPS 沉淀 `[1]` 前缀 = 外部工作簿引用机械事实: 新写公式未重建外部
        链接不可求值 (formula_not_evaluated), 该列直接 null 而非常规复制公式."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(
            encoding="utf-8")
        m = re.search(r"^\|\s*\*\*.*\[\d+\].*前缀 = 外部工作簿引用.*$", text, re.MULTILINE)
        self.assertIsNotNone(m, "KNOWN_TRAPS 缺 [1] 外部工作簿引用 行")
        row = m.group(0)
        for word in ("[1]", "外部工作簿引用", "formula_not_evaluated", "直接",
                     "null", "外部工作簿", "XLOOKUP"):
            self.assertIn(word, row, f"[1] 外部引用行缺 {word!r}")

    def test_layer4_failure_classes_render_default_html(self):
        """LAYER4 与 FAILURE_CLASSES 同步 render 默认值: 省略 --render 时按 html
        执行 (进出接口契约一致)."""
        layer4 = (SKILL_ROOT / "references" / "LAYER4_EXECUTE_LOOP.md").read_text(
            encoding="utf-8")
        self.assertIn("默认 `html`", layer4,
                      "LAYER4 缺 render 默认 html 措辞")
        classes = (SKILL_ROOT / "references" / "FAILURE_CLASSES.md").read_text(
            encoding="utf-8")
        self.assertIn("默认 `--render html`", classes,
                      "FAILURE_CLASSES 缺 render 默认 html 措辞")

    def test_skill_md_machine_evidence_termination_extended(self):
        """机器证据终止条件扩展: execute 已返回机器证据后禁止 officecli get 逐格
        复核 + 禁止读 case 复盘/测试病历作证据; 唯一例外 = 异常驱动定向 get ≤2."""
        text = self._skill_md_text()
        for word in ("机器证据终止条件", "逐格复核", "officecli get", "case 复盘",
                     "测试病历", "异常驱动的定向检查", "禁止"):
            self.assertIn(word, text, f"SKILL.md 机器证据终止条件缺词 {word!r}")

    def test_skill_md_canonical_pattern_instantiate_stop(self):
        """结构/层级缺陷预算补 canonical→STOP: 命中 combination_patterns 的
        canonical pattern → 直接实例化, 不再读 case 复盘重推组合."""
        text = self._skill_md_text()
        for word in ("结构/层级缺陷预算", "直接实例化", "不再读 case 复盘",
                     "canonical pattern", "重推组合"):
            self.assertIn(word, text, f"SKILL.md canonical→STOP 缺词 {word!r}")

    def test_skill_md_mod_conflict_no_full_content_recheck(self):
        """MOD Resolution 补: MOD conflict 且排除信号命中 → 不再读 MOD 全文核对
        排除信号是否误报, 直接 fail-closed ASK (领域判断不改变裁决机制)."""
        text = self._skill_md_text()
        m = re.search(r"^### 2\. MOD Resolution.*?(?=^### 3\.)",
                      text, re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "SKILL.md 缺 MOD Resolution 段")
        section = m.group(0)
        for word in ("不再读 MOD 全文", "排除信号", "fail-closed ASK", "误报"):
            self.assertIn(word, section,
                          f"MOD Resolution 缺不读全文词 {word!r}")

    def test_skill_md_mod_ask_checklist_templated(self):
        """MOD ASK 必问清单模板化: 成本口径 / 面价 vs 散件 / 缺失稳定属性 / 费用
        组成 / 含管口径 / 输出文件形态 一次性枚举 (防第二轮补问)."""
        text = self._skill_md_text()
        for word in ("必问清单", "成本口径", "面价 vs 散件", "缺失稳定属性",
                     "费用组成", "含管口径", "输出文件形态", "一轮问全"):
            self.assertIn(word, text, f"SKILL.md MOD ASK 清单缺词 {word!r}")

    def test_fillspec_execution_order_section(self):
        """「执行顺序保证」章节存在且覆盖全部四条锁定声明: op 顺序不变量 /
        remove 目标身份 / 自底向上 / 坐标翻译边界 (防章节误删或声明回退)."""
        section = self._fillspec_section("执行顺序保证")
        for word in ("clear", "add", "remove", "merge", "fill",
                     "duplicate_row", "模板坐标", "自底向上", "readback",
                     "REMOVE_TARGETS_APPEND_ZONE", "deferred_values",
                     "final_row"):
            self.assertIn(word, section, f"执行顺序保证章节缺 {word!r}")
        self.assertIsNotNone(re.search(r"^### E1:", section, re.MULTILINE), "缺 E1")
        self.assertIsNotNone(re.search(r"^### E2:", section, re.MULTILINE), "缺 E2")
        self.assertIsNotNone(re.search(r"^### E3:", section, re.MULTILINE), "缺 E3")
        self.assertIsNotNone(re.search(r"^### E4:", section, re.MULTILINE), "缺 E4")

    def test_known_traps_remove_add_interaction(self):
        """KNOWN_TRAPS 沉淀 remove/add 交互权威事实 (埃及案例重放 oracle:
        Agent 无需读源码即可回答 add 后 remove 目标是谁 — 指向 FILLSPEC 章节)."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        self.assertIn("不被 add 推移", text)
        self.assertIn("执行顺序保证", text)
        self.assertIn("mechanical_facts", text)

    def test_mod_index_execution_boundary(self):
        """MOD_INDEX 标注执行 vs 治理文档边界: 执行任务不读 MOD_TEMPLATE."""
        text = (SKILL_ROOT / "references" / "MOD_INDEX.md").read_text(encoding="utf-8")
        self.assertIn("不读", text)
        self.assertIn("MOD_TEMPLATE", text)

    def test_skill_md_prepare_sheets_all_at_once(self):
        """SKILL.md prepare 阶段 B: 一次列出全部源 sheet, 增量展平是兜底."""
        text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("一次列出全部源 sheet", text)
        self.assertIn("兜底", text)

    def test_fillspec_q15_lookup_integrity_section(self):
        """契约章节含 Q15: 空索引 → LOOKUP_TABLE_EMPTY (exit 3); 整列未命中 →
        LOOKUP_COLUMN_ALL_MISSING 警告 (防章节误删 / 静默全空回退)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q15:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q15 小节")
        q15 = section[m.end():]
        self.assertIn("LOOKUP_TABLE_EMPTY", q15)
        self.assertIn("LOOKUP_COLUMN_ALL_MISSING", q15)
        self.assertIn("field_consensus", q15)

    def test_fillspec_error_code_table_has_lookup_integrity(self):
        """LOOKUP_TABLE_EMPTY / LOOKUP_COLUMN_ALL_MISSING 在「常见编译错误速查」
        表内 (防章节误删)."""
        table = self._error_code_table()
        self.assertIn("LOOKUP_TABLE_EMPTY", table)
        self.assertIn("LOOKUP_COLUMN_ALL_MISSING", table)

    def test_known_traps_lookup_index_rebuild(self):
        """KNOWN_TRAPS 沉淀索引清洗机械事实: LOOKUP_TABLE_EMPTY 拦截 + 禁止手改
        JSON, 用 build_inheritance_index.py 重建."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        self.assertIn("LOOKUP_TABLE_EMPTY", text)
        self.assertIn("手改 JSON", text)
        self.assertIn("build_inheritance_index.py", text)

    def test_fillspec_lookups_source_excludes_target(self):
        """「lookups」契约条目: 索引输入 sheet 排除目标 sheet — 自引用 → 共识
        conflict → 按缺失处理 (静默), 排查线索见 KNOWN_TRAPS (埃及 FRESH 坑 2)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(encoding="utf-8")
        m = re.search(r"^### lookups\s*$", text, re.MULTILINE)
        self.assertIsNotNone(m, "FILLSPEC.md 缺 ### lookups 小节")
        nxt = re.search(r"^### ", text[m.end():], re.MULTILINE)
        section = text[m.end():m.end() + (nxt.start() if nxt else len(text))]
        self.assertIn("目标 sheet", section)
        self.assertIn("自引用", section)
        self.assertIn("KNOWN_TRAPS", section)

    def test_fillspec_q15_self_reference_hint(self):
        """Q15 整列未命中排查含自引用线索 (索引输入误含目标 sheet)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q15:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q15 小节")
        q15 = section[m.end():]
        self.assertIn("目标 sheet", q15)
        self.assertIn("自引用", q15)

    def test_known_traps_lookup_source_self_reference(self):
        """KNOWN_TRAPS 沉淀索引自引用机械事实: 输入含目标 sheet → 同 SKU 多值
        → 共识 conflict → 按缺失处理; 构建索引只喂独立数据 sheet."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        self.assertIn("目标 sheet", text)
        self.assertIn("共识 conflict", text)

    def test_fillspec_layout_decision_tree_style_first(self):
        """FILLSPEC 布局决策树: 以样式为第一判定条件 (带样式分支先于裸行分支),
        三分支齐全, 各分支与缺陷码对应 (文档与编译器裁决同源)."""
        section = self._fillspec_section("布局决策树")
        self.assertIn("以样式为第一判定条件", section)
        for word in ("带样式", "裸行", "inplace", "clone-append",
                     "append-only", "remove_rows", "自然下沉", "收缩"):
            self.assertIn(word, section)
        self.assertLess(section.index("带样式"), section.index("裸行"),
                        "样式条件必须作为第一判定条件: 带样式分支先于裸行分支")
        self.assertLess(section.index("①"), section.index("②"),
                        "分支 ① (带样式→inplace) 必须先于分支 ② (裸行→clone-append)")
        self.assertLess(section.index("②"), section.index("③"),
                        "分支 ② 必须先于分支 ③ (既有块收缩)")
        for code in ("REMOVE_TARGETS_APPEND_ZONE", "TEMPLATE_ROW_GAP",
                     "STRUCTURAL_OP_OUT_OF_ZONE"):
            self.assertIn(code, section)
        self.assertLess(section.index("≤ base_last_row"),
                        section.index("TEMPLATE_ROW_GAP"),
                        "分支 ③ 必须声明 remove_rows ≤ base_last_row 边界")

    def test_skill_md_mod_loading_output_form(self):
        """SKILL.md MOD 段输出形态优化: 提名阶段只给摘要不含完整规则集,
        裁决后才加载选中 MOD 完整规则; 「候选规则必须加载后才可写 spec」
        的硬性要求保留 (改变加载时机与粒度, 不是是否加载)."""
        text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        m = re.search(r"^### 2\. MOD Resolution.*?(?=^### 3\.)",
                      text, re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "SKILL.md 缺 MOD Resolution 段")
        section = m.group(0)
        for word in ("摘要", "不含完整规则集", "裁决后", "选中"):
            self.assertIn(word, section)
        self.assertIn("必须加载后才可写 spec", section)
        self.assertIn("不因输出形态优化放宽", section)

    def test_known_traps_three_workflows_oracles(self):
        """KNOWN_TRAPS 与三个工作流产物对应 (07 验收 #2): remove/add 交互
        (01/02), 裸行占位 → append-only 终态 (03/04), 显式范围 nulls 触发
        条件 (05), 组聚合落点 (06) 全部有重放 oracle."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        for word in ("不被 add 推移", "裸行占位", "占位行样式", "first as empty",
                     "group_aggregates", "append-only"):
            self.assertIn(word, text, f"KNOWN_TRAPS 缺机械事实词 {word!r}")

    def test_fillspec_q16_row_gap_fingerprint_sync(self):
        """契约章节 Q16: repair_row_gaps 自动重跑 flatten 同步 manifest 指纹;
        唯一手工动作 = 更新 spec 指纹 (或 --patch-spec) + 重编译 (防流程回退
        到手工三步同步)."""
        section = self._fillspec_section("组合行为契约")
        m = re.search(r"^### Q16:", section, re.MULTILINE)
        self.assertIsNotNone(m, "契约章节缺 Q16 小节")
        q16 = section[m.end():]
        for word in ("repair_row_gaps.py", "自动", "指纹", "patch-spec",
                     "必然变化"):
            self.assertIn(word, q16)

    def test_error_code_table_has_template_row_gap(self):
        """TEMPLATE_ROW_GAP 在「常见编译错误速查」表内 (防章节误删)."""
        table = self._error_code_table()
        self.assertIn("TEMPLATE_ROW_GAP", table)
        self.assertIn("repair_row_gaps.py", table)

    def test_error_code_table_has_block_key_structure_invalid(self):
        """BLOCK_KEY_STRUCTURE_INVALID 在「常见编译错误速查」表内, 行内修复指引
        点名"块级错位键"与"应写 formulas 嵌套" (ID-1: 不再静默忽略)."""
        table = self._error_code_table()
        self.assertIn("BLOCK_KEY_STRUCTURE_INVALID", table)
        m = re.search(r"^\|\s*BLOCK_KEY_STRUCTURE_INVALID.*$", table,
                      re.MULTILINE)
        self.assertIsNotNone(m, "速查表缺 BLOCK_KEY_STRUCTURE_INVALID 行")
        row = m.group(0)
        self.assertIn("块级", row)
        self.assertIn("formulas", row)
        self.assertIn("不再静默", row)

    # ── issue 02: 文档契约硬化 (blocks 继承/取代 + transforms + key_outputs ──
    #    + 机器证据终止 + ROUND 优先序 + 结构缺陷预算) ──

    def test_fillspec_blocks_formulas_replace_contract_word(self):
        """FILLSPEC blocks 段: 块级声明 `formulas` 即整体取代而非合并 target 级
        per_row — "取代" 措辞 + "必须整段携带" 继承契约必须在 (ID-4)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        m = re.search(r"^### blocks:.*?(?=^### )", text,
                      re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "FILLSPEC 缺 blocks 小节")
        section = re.sub(r"\s+", "", m.group(0))  # 跨行折行不破坏断言
        for word in ("取代", "整段携带", "每行物化"):
            self.assertIn(word, section, f"blocks 段缺 {word!r}")

    def test_fillspec_blocks_counterexample_warnings(self):
        """FILLSPEC blocks 段反例警示: 反例警示标题 + 静默丢弃回退措辞 +
        BLOCK_KEY_STRUCTURE_INVALID / group_by 稀疏源列不建组 (ID-4 反例)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        m = re.search(r"^### blocks:.*?(?=^### )", text,
                      re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "FILLSPEC 缺 blocks 小节")
        section = m.group(0)
        self.assertIn("反例警示", section)
        self.assertIn("静默丢弃", section)
        self.assertIn("BLOCK_KEY_STRUCTURE_INVALID", section)
        self.assertIn("稀疏源列", section)

    def test_fillspec_transforms_builtin_only_round(self):
        """FILLSPEC transforms 修正: 内建数值变换仅 round2/round4, strip/regex
        需在 mapping.transforms 定义 (ID-9 — 防"transforms 支持 strip"诱导向
        回退)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        m = re.search(r"^transforms 说明:.*?(?=^### )", text,
                      re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "找不到 transforms 说明段")
        section = re.sub(r"\s+", "", m.group(0))  # 内嵌代码反引号/折行无关
        for word in ("内建数值变换仅", "round2", "round4", "不是内建",
                     "mapping.transforms", "regex_replace", "strip_sku"):
            self.assertIn(word, section, f"transforms 说明缺 {word!r}")

    def test_fillspec_key_outputs_data_start_word(self):
        """FILLSPEC key_outputs 说明: 行号可直接取 plan blocks[].data_start,
        不手工重推 (ID-8 — 防回归到手工行号推导)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        self.assertIn("blocks[].data_start", text)
        self.assertIn("不手工重推", text)
        table = self._error_code_table()
        m = re.search(r"^\|\s*KEY_OUTPUT_UNWRITTEN.*$", table, re.MULTILINE)
        self.assertIsNotNone(m, "速查表缺 KEY_OUTPUT_UNWRITTEN 行")
        self.assertIn("data_start", m.group(0))

    def test_skill_md_machine_evidence_termination(self):
        """SKILL.md 机器证据终止条件: execute 已返回机器证据后禁止 officecli
        issues / 读 execution_plan.json 人工复核; 唯一例外 = 异常驱动的定向
        officecli get ≤2 (ID-5, 防 E5/E6 冗余复核回退)."""
        text = self._skill_md_text()
        self.assertIn("机器证据终止条件", text)
        self.assertIn("officecli", text)
        self.assertIn("issues", text)
        self.assertIn("execution_plan.json", text)
        self.assertIn("异常驱动的定向检查", text)
        self.assertIn("禁止", text)

    def test_skill_md_structure_defect_budget(self):
        """SKILL.md 结构/层级缺陷预算: Compile 结构类缺陷 (KEY_OUTPUT_UNWRITTEN /
        CLONE_RESIDUE_* / block 结构) 默认第一轮先查 combination_patterns 找
        模式, 而非自由改 spec 重编译 (ID-6, 压缩 E4 3次→1次)."""
        text = self._skill_md_text()
        self.assertIn("结构/层级缺陷预算", text)
        self.assertIn("KEY_OUTPUT_UNWRITTEN", text)
        self.assertIn("CLONE_RESIDUE", text)
        self.assertIn("combination_patterns.yaml", text)

    def test_skill_and_fillspec_round_precedence(self):
        """SKILL.md 与 FILLSPEC 公式约定: ROUND 优先序 — 即使模板既有公式无
        ROUND、即使 officecli-xlsx preserve 建议复刻, table-fill ROUND 精准原则
        优先 (ID-7); text_overflow 属 REPAIR 预期路径."""
        skill = self._skill_md_text()
        self.assertIn("ROUND 优先序", skill)
        self.assertIn("preserve existing", skill)
        self.assertIn("REPAIR", skill)
        fillspec = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(
            encoding="utf-8")
        self.assertIn("ROUND 优先序", fillspec)
        self.assertIn("preserve existing", fillspec)

    def test_known_traps_row_gap_auto_resync(self):
        """KNOWN_TRAPS 沉淀行洞修复机械事实: 行洞修复 = staged 文件修改 =
        指纹必然变化; repair 脚本自动重算, Agent 不再手工同步."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        for word in ("指纹必然变化", "自动", "patch-spec", "唯一动作"):
            self.assertIn(word, text)

    def test_skill_md_repair_auto_flatten(self):
        """SKILL.md prepare 段: repair 后 flatten 已自动, 唯一动作 = 更新
        spec 指纹 + 重编译."""
        text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        m = re.search(r"^### 1\. Prepare.*?(?=^### 2\.)",
                      text, re.MULTILINE | re.DOTALL)
        self.assertIsNotNone(m, "SKILL.md 缺 Prepare 段")
        section = m.group(0)
        for word in ("repair_row_gaps.py", "自动重跑 flatten", "patch-spec",
                     "唯一动作"):
            self.assertIn(word, section)

    def test_fillspec_yaml_discipline_whole_line_quote(self):
        """FILLSPEC「YAML 纪律」契约条目: decisions/gaps 条目含 ': ' →
        整行双引号包裹 (含冒号), 漏写 → SPEC_NON_STRING_ITEM exit 3,
        corrective_action 给正确写法 (防契约条目被误删或回退到 dict repr)."""
        text = (SKILL_ROOT / "references" / "FILLSPEC.md").read_text(encoding="utf-8")
        m = re.search(r"^### YAML 纪律", text, re.MULTILINE)
        self.assertIsNotNone(m, "FILLSPEC.md 缺 ### YAML 纪律 小节")
        nxt = re.search(r"^### ", text[m.end():], re.MULTILINE)
        section = text[m.end():m.end() + (nxt.start() if nxt else len(text))]
        for word in ("整行双引号包裹", "SPEC_NON_STRING_ITEM", "exit 3",
                     "corrective_action", "冒号"):
            self.assertIn(word, section)
        self.assertIn('"追加新历史块', section)

    def test_error_code_table_has_spec_non_string_item(self):
        """SPEC_NON_STRING_ITEM 在「常见编译错误速查」表内且修复指引为
        整行双引号包裹 (防误导性修复建议回退)."""
        table = self._error_code_table()
        m = re.search(r"^\|\s*SPEC_NON_STRING_ITEM.*$", table, re.MULTILINE)
        self.assertIsNotNone(m, "速查表缺 SPEC_NON_STRING_ITEM 行")
        self.assertIn("整行双引号包裹", m.group(0))

    def test_skill_md_yaml_discipline_whole_line_quote(self):
        """SKILL.md「YAML 纪律」: decisions/gaps 含 ': ' → 整行 (含冒号)
        加双引号, 漏写 → SPEC_NON_STRING_ITEM exit 3 (防纪律词被降级为
        '统一加引号' 的模糊表达)."""
        text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        m = re.search(r"^-\s+\*\*YAML 纪律\*\*", text, re.MULTILINE)
        self.assertIsNotNone(m, "SKILL.md 缺 YAML 纪律 纪律条")
        nxt = re.search(r"^-\s+\*\*", text[m.end():], re.MULTILINE)
        bullet = text[m.start():m.start() + (nxt.start() if nxt else 400)]
        self.assertIn("整行", bullet)
        self.assertIn("SPEC_NON_STRING_ITEM", bullet)

    def test_known_traps_yaml_quote_mechanical_fact(self):
        """KNOWN_TRAPS 沉淀 YAML 引号机械事实: 整行双引号包裹 + 漏写 →
        SPEC_NON_STRING_ITEM exit 3 + corrective_action 给正确写法."""
        text = (SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(encoding="utf-8")
        m = re.search(r"^\|\s*\*\*YAML 引号漏写\*\*.*$", text, re.MULTILINE)
        self.assertIsNotNone(m, "KNOWN_TRAPS 缺 YAML 引号漏写 行")
        row = m.group(0)
        for word in ("整行双引号包裹", "SPEC_NON_STRING_ITEM", "corrective_action"):
            self.assertIn(word, row)

    # ── issue 07: OFFICECLI_REFERENCE 对齐当前执行契约 + 关键词守卫 ──
    # (只守关键词, 不复制整段文档; 权威契约在 LAYER4 / FILLSPEC / KNOWN_TRAPS)

    def _officecli_reference_text(self) -> str:
        return (SKILL_ROOT / "references" / "OFFICECLI_REFERENCE.md").read_text(
            encoding="utf-8")

    def test_officecli_reference_flush_contract(self):
        """写持久化契约: resident 延迟写 + 显式 close 刷盘 (与 LAYER4 /
        KNOWN_TRAPS 同源); 禁止回退到「立即生效/直接修改文件」表述."""
        text = self._officecli_reference_text()
        for word in ("resident", "close", "刷盘", "延迟"):
            self.assertIn(word, text, f"OFFICECLI_REFERENCE 缺写持久化词 {word!r}")
        for stale in ("立即生效", "直接修改文件", "不需要额外保存"):
            self.assertNotIn(stale, text,
                             f"OFFICECLI_REFERENCE 仍含陈旧表述 {stale!r}")

    def test_officecli_reference_op_order_invariant(self):
        """op 恒序 clear→add→remove→merge→fill (与 FILLSPEC E1 / KNOWN_TRAPS
        同源): OFFICECLI_REFERENCE 与 KNOWN_TRAPS 都按此序含全部五个阶段,
        remove 不得排在 add 之前."""
        norm = lambda s: re.sub(r"\s+", "", s)  # noqa: E731
        invariant = "clear→add→remove→merge→fill"
        ref = norm(self._officecli_reference_text())
        self.assertIn(invariant, ref,
                      "OFFICECLI_REFERENCE 缺 op 恒序 clear→add→remove→merge→fill")
        traps = norm((SKILL_ROOT / "references" / "KNOWN_TRAPS.md").read_text(
            encoding="utf-8"))
        self.assertIn(invariant, traps,
                      "KNOWN_TRAPS 缺 op 恒序 (remove/add 交互行)")

    def test_officecli_reference_xlsx_props(self):
        """属性名: xlsx 用 value/numberformat (PROPS_WHITELIST 同源), pptx 用
        text; 禁止回退到把 xlsx text/numFmt 写成 JSON props."""
        text = self._officecli_reference_text()
        for word in ("value", "numberformat"):
            self.assertIn(word, text, f"OFFICECLI_REFERENCE 缺 xlsx 属性词 {word!r}")
        # 陈旧属性名的"处方形" (JSON prop 写法) 禁止; 文档提及其为陈旧别名
        # (警示) 是允许的 — 同 KNOWN_TRAPS「text 属性漂移」「numFmt 大小写歧义」
        self.assertNotIn('"numFmt"', text,
                         "OFFICECLI_REFERENCE 仍把 numFmt 写成 JSON prop")
        m = re.search(r"```json\n(.*?)```", text, re.DOTALL)
        self.assertIsNotNone(m, "OFFICECLI_REFERENCE 缺 batch JSON 示例")
        batch_example = m.group(1)
        self.assertNotIn('"text"', batch_example,
                         "batch JSON 示例 (xlsx) 仍用 text 属性")
        self.assertIn('"value"', batch_example,
                      "batch JSON 示例 (xlsx) 缺 value 属性")
        self.assertIn('"numberformat"', batch_example,
                      "batch JSON 示例 (xlsx) 缺 numberformat 属性")
        m = re.search(r"### PPTX 特殊要求", text)
        self.assertIsNotNone(m, "OFFICECLI_REFERENCE 缺 PPTX 特殊要求 小节")
        self.assertIn("text", text[m.end():],
                      "OFFICECLI_REFERENCE PPTX 段缺 text 属性")

    # ── issue 09: Office 访问与 UTF-8 适配器规则 (结构解析 vs officecli 适配器) ──

    def _tool_traps_text(self) -> str:
        return (SKILL_ROOT / "references" / "TOOL_TRAPS.md").read_text(
            encoding="utf-8")

    def test_tool_traps_adapter_rule_wording(self):
        """TOOL_TRAPS 编码统一行声明适配器硬性规则 (issue 09): officecli 子进程
        调用一律走 `_officecli.officecli()` (UTF-8 subprocess)."""
        text = self._tool_traps_text()
        self.assertIn("一律走", text)
        self.assertIn("_officecli.officecli()", text)

    def test_tool_traps_structure_parsing_rule_wording(self):
        """TOOL_TRAPS 明确两条规则分开 (issue 09): ZIP/XML 结构解析允许直读,
        与 officecli 适配器规则并存 — 结构解析规则不得被"全走 officecli"吞掉."""
        text = self._tool_traps_text()
        self.assertIn("ZIP/XML", text)
        self.assertIn("结构解析", text)
        self.assertIn("_officecli.officecli()", text)

    def test_skill_md_adapter_rule_wording(self):
        """SKILL.md frontmatter: 「All read/write goes through officecli」已
        精确化 — 两条规则分开 (issue 09): 结构解析允许直读 ZIP/XML, 任何
        officecli 子进程调用必须经 `_officecli.officecli()`."""
        text = self._skill_md_text()
        self.assertIn("ZIP/XML", text)
        self.assertIn("_officecli.officecli()", text)
        self.assertIn("结构解析", text)

    def test_adapter_rule_consistent_across_docs(self):
        """SKILL.md 与 TOOL_TRAPS.md 的适配器措辞一致 (issue 09): 两边都声明
        officecli 调用走 `_officecli.officecli()` 且结构解析 (ZIP/XML) 允许直读."""
        skill = self._skill_md_text()
        traps = self._tool_traps_text()
        for text, name in ((skill, "SKILL.md"), (traps, "TOOL_TRAPS.md")):
            self.assertIn("_officecli.officecli()", text,
                          f"{name} 缺适配器规则词 _officecli.officecli()")
            self.assertIn("ZIP/XML", text,
                          f"{name} 缺结构解析规则词 ZIP/XML")


class OfficecliAdapterAuditTests(unittest.TestCase):
    """issue 09 静态审计: scripts/ 下禁止直接 subprocess.run(["officecli", ...])
    绕过 `_officecli.officecli()` 共享适配器 (Windows 中文输出 / 错误处理 /
    resident 清理行为一致). 豁免需在 EXEMPTIONS 登记理由 — 当前无豁免."""

    # 已登记豁免: {路径: 理由}. 默认收敛, 新增绕过必须先记录豁免理由.
    EXEMPTIONS: dict[str, str] = {}

    # _officecli.py 是适配器本体, 是唯一允许的 officecli subprocess 宿主.
    ADAPTER_FILE = "_officecli.py"

    DIRECT_CALL_RE = re.compile(
        r'subprocess\.run\(\s*\[\s*["\']officecli["\']', re.DOTALL)

    def _scripts(self):
        return sorted((SKILL_ROOT / "scripts").glob("*.py"))

    def test_no_direct_officecli_subprocess_bypass(self):
        """scripts/ 下除适配器本体 (与已登记豁免) 外, 无直接
        subprocess.run(["officecli", ...]) 绕过共享适配器 (issue 09)."""
        offenders = []
        for py in self._scripts():
            if py.name == self.ADAPTER_FILE or py.name in self.EXEMPTIONS:
                continue
            text = py.read_text(encoding="utf-8")
            for m in self.DIRECT_CALL_RE.finditer(text):
                line = text.count("\n", 0, m.start()) + 1
                offenders.append(f"{py.name}:{line}")
        self.assertEqual(
            offenders, [],
            "scripts/ 存在绕过 _officecli.officecli() 的直接 officecli "
            f"subprocess 调用: {', '.join(offenders)}")

    def test_scripts_import_adapter_or_exempt(self):
        """任何含 officecli 关键字调用的脚本必须经 `from _officecli import
        officecli` (豁免除外) — 防止换一种裸调用写法绕过审计 (issue 09)."""
        import re as _re
        call_re = _re.compile(r'\b(officecli|officecli_validate)\(')
        for py in self._scripts():
            if py.name == self.ADAPTER_FILE or py.name in self.EXEMPTIONS:
                continue
            text = py.read_text(encoding="utf-8")
            if not call_re.search(text):
                continue  # 不调用 officecli, 无关
            self.assertIn(
                "from _officecli import", text,
                f"{py.name} 调用 officecli 但未导入共享适配器 (issue 09)")

    def test_officecli_adapter_module_documented_contract(self):
        """适配器本体仍声明为唯一 subprocess 宿主 (issue 09): _officecli.py
        的模块 docstring 描述 UTF-8 子进程包装职责."""
        text = (SKILL_ROOT / "scripts" / self.ADAPTER_FILE).read_text(
            encoding="utf-8")
        self.assertIn("UTF-8 subprocess", text)
        self.assertIn("never raw PowerShell", text)


class ReadbackNormalizationTests(unittest.TestCase):
    """Q10 readback value 断言: 数值归一化只适用于真数值形态; 字母数字标识
    (SKU/型号/Z 码) 一律按文本精确比较 (issue 04 — 阻止 Z001 vs X001 误判相等)."""

    def _check(self, expected: str, actual: str, kind: str = "value") -> dict | None:
        """readback 断言最小面: read_map 命中即不触碰文件系统 (book 不打开)."""
        return execute_batch.check_expectation(
            Path("unused.xlsx"), "/S/A1", kind, expected,
            {"/S/A1": actual})

    def test_number_normalized_rejects_non_numeric_literals(self):
        """Z001 / SN-001 / ABC123 等字母数字标识不是数值字面量 → None
        (禁止归一化把 SN-001 变成 -1.0、Z001/X001 都变成 1.0); 指数/NaN
        形态同样不归一化 (保守方向: 拿不准就文本比较)."""
        for s in ("Z001", "X001", "SN-001", "ABC123", "1e5", "NaN"):
            self.assertIsNone(execute_batch.number_normalized(s),
                              f"{s!r} 不应被数值归一化")

    def test_number_normalized_accepts_true_numeric_forms(self):
        """真数值形态 (含货币/千分位/百分比装饰与前导符号) 仍归一化."""
        cases = {"138.00": 138.0, "138": 138.0, "$1,234.5": 1234.5,
                 "1,234.50": 1234.5, "-45.2%": -45.2, "12.5%": 12.5,
                 " 12 ": 12.0, "+12": 12.0, ".5": 0.5}
        for s, want in cases.items():
            self.assertEqual(execute_batch.number_normalized(s), want, s)

    def test_alphanumeric_id_mismatch_readback_fails(self):
        """Z001 vs X001 → 文本精确比较 → 失败 (不再归一化为 1.0 误判相等)."""
        failure = self._check("Z001", "X001")
        self.assertIsNotNone(failure)
        self.assertEqual(failure["expected"], "Z001")
        self.assertEqual(failure["actual"], "X001")

    def test_sn_style_id_exact_text_compare(self):
        """SN-001 类标识按文本精确比较: 相同通过, 不同失败."""
        self.assertIsNone(self._check("SN-001", "SN-001"))
        failure = self._check("SN-001", "SN-002")
        self.assertIsNotNone(failure)

    def test_numeric_forms_still_normalize_pass(self):
        """真数值形态差异仍被容忍: 138.00 vs 138、$1,234.5 vs 1234.5、百分比."""
        self.assertIsNone(self._check("138.00", "138"))
        self.assertIsNone(self._check("$1,234.5", "1234.5"))
        self.assertIsNone(self._check("12.5%", "12.5"))

    def test_numeric_mismatch_still_fails(self):
        """真数值形态的真正差异仍失败 (容差不掩盖差异)."""
        self.assertIsNotNone(self._check("138.00", "139"))
        self.assertIsNotNone(self._check("$1,234.5", "1,235"))

    def test_empty_nonempty_kinds_unaffected(self):
        """empty/nonempty 断言不经过数值归一化, 行为不变."""
        self.assertIsNone(self._check("", "", kind="empty"))
        self.assertIsNotNone(self._check("", "X", kind="empty"))
        self.assertIsNone(self._check("x", "SN-001", kind="nonempty"))
        self.assertIsNotNone(self._check("x", "", kind="nonempty"))


class ModCatalogIndexTests(unittest.TestCase):
    """MOD_INDEX.md 目录解析: 转义管道 + 修订号漂移守护."""

    def test_escaped_pipe_keeps_revision_column(self):
        """信号格含 \\| 转义时, 列对齐不被破坏, revision 解析正确."""
        text = (
            "## Registered MODs\n\n"
            "| MOD Name | Aliases | Scope Signals (+) | Exclusion Signals (-) | Path | Revision | Visibility |\n"
            "|---|---|---|---|---|---|---|\n"
            "| m1 | alias | semantic_type::quotation,sheet_marker::三三三\\|333 |  | MOD_m1.md | 5 | private |\n"
        )
        entries = parse_mod_index(text)
        self.assertEqual(len(entries), 1)
        e = entries[0]
        self.assertEqual(e.mod_name, "m1")
        self.assertEqual(e.scope_signals, "semantic_type::quotation,sheet_marker::三三三|333")
        self.assertEqual(e.revision, 5)
        self.assertEqual(e.path, "MOD_m1.md")
        self.assertEqual(e.visibility, "private")

    def test_live_index_revisions_synced_with_mod_files(self):
        """真实 MOD_INDEX 与 MOD 文件头修订号一致 (漂移守护).

        发布仓库不携带私有客户 MOD — Registered MODs 为空时本守卫跳过
        (私有 MOD 随捕获流程存在于本地, 不随 office-skills 推送)."""
        refs = SKILL_ROOT / "references"
        index_text = (refs / "MOD_INDEX.md").read_text(encoding="utf-8")
        entries = parse_mod_index(index_text)
        if not entries:
            return
        for entry in entries:
            mod_file = refs / entry.path
            self.assertTrue(mod_file.is_file(), f"索引指向缺失文件: {entry.path}")
            mod_text = mod_file.read_text(encoding="utf-8")
            m = re.search(r"^Revision:\s*(\d+)", mod_text, re.MULTILINE)
            self.assertIsNotNone(m, f"{entry.path} 缺 Revision 头")
            self.assertEqual(
                entry.revision, int(m.group(1)),
                f"{entry.mod_name} 修订号漂移: 索引 {entry.revision} vs 文件 {m.group(1)}")


if __name__ == "__main__":
    unittest.main()
