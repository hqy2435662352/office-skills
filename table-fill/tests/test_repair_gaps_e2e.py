"""Issue 05 — repair_row_gaps.py 自动重算指纹 (消除手工同步).

2026-08-13 埃及 FRESH 运行复盘: 目标 sheet 行号空洞 (row 22 缺失) →
TEMPLATE_ROW_GAP 编译拒绝 → repair_row_gaps.py 物化行 → staged 文件变了 →
指纹变 → 必须手工重跑 prepare_run.py --flatten + 把新 target_structure 抄进
fill_spec.yaml → 重编译。三步手工环节, 每步都可能抄错/漏跑。

本契约把「修复后自动同步」变成脚本行为:

- repair 成功后自动重跑 flatten (仅目标 sheet), 同步 prepare_manifest.json
  指纹 (manifest 指纹 == repair 输出指纹 == 手工重 flatten 结果);
- 行洞修复 = staged 文件修改 = 指纹必然变化 (机械事实, 本文件断言);
- --patch-spec 一步把新 target_structure 写进 fill_spec.yaml;
- 无空洞时是 no-op (NO_ROW_GAPS, 幂等)。

Smoke (脚本集成, 无单元测试 — 见 issue 05): 构造带行洞 fixture →
repair → manifest 指纹 == 重 flatten 结果。officecli 不可用时跳过
(unit-only 环境)。

Run with:
  python -m unittest tests.test_repair_gaps_e2e -v
"""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

import yaml

SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = SKILL_ROOT / "scripts"


def run_py(workdir: Path, script: str, *args) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, "-X", "utf8", str(SCRIPTS / script), *args],
        cwd=str(workdir), capture_output=True, text=True, encoding="utf-8",
        errors="replace", timeout=900)


def parse_json_docs(text: str) -> list[dict]:
    """Parse consecutive JSON documents on stdout (repair 先输出指纹结果,
    --patch-spec 成功后追加 SPEC_PATCHED 结果)."""
    dec = json.JSONDecoder()
    out = []
    i = 0
    while i < len(text):
        while i < len(text) and text[i] in " \r\n":
            i += 1
        if i >= len(text):
            break
        obj, i = dec.raw_decode(text, i)
        out.append(obj)
    return out


def build_gap_fixture(dirpath: Path) -> None:
    """4-row template (sheet TPL) with <row r=3> element removed from the
    sheet XML + 2-row source (sheet SRC). Sheet names are >1 char so officecli
    path parsing is unambiguous."""
    from openpyxl import Workbook
    src = Workbook()
    ws = src.active
    ws.title = "SRC"
    for i in range(2):
        ws.cell(row=i + 1, column=1, value=f"v{i}")
    src.save(dirpath / "source.xlsx")

    tpl = Workbook()
    ws = tpl.active
    ws.title = "TPL"
    ws["A1"] = "Title"
    ws["A2"] = "Header"
    ws["A3"] = "old data"
    ws["A4"] = "Total"
    p = dirpath / "template.xlsx"
    tpl.save(p)

    zf = zipfile.ZipFile(p)
    names = zf.namelist()
    sheet = next(n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
    xml = zf.read(sheet).decode("utf-8")
    new_xml = re.sub(r"<row r=\"3\"[^>]*>.*?</row>", "", xml, flags=re.DOTALL)
    assert new_xml != xml, "row r=3 element not found in fixture XML"
    entries = {n: zf.read(n) for n in names}
    zf.close()
    tmp = p.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zo:
        for n in names:
            zo.writestr(n, new_xml if n == sheet else entries[n])
    shutil.move(tmp, p)


@unittest.skipIf(shutil.which("officecli") is None, "officecli not on PATH")
class RepairGapsResyncTests(unittest.TestCase):
    """Issue 05 冒烟: repair 自动重 flatten 同步指纹; 指纹必然变化;
    --patch-spec 一步改 spec; 无空洞幂等."""

    def setUp(self):
        sys.path.insert(0, str(SCRIPTS))
        from _officecli import clean_residents  # noqa: PLC0415
        clean_residents()
        self.workdir = Path(tempfile.mkdtemp(prefix="gap_e2e_"))
        build_gap_fixture(self.workdir)

    def tearDown(self):
        from _officecli import clean_residents, unlink_retry  # noqa: PLC0415
        import time
        clean_residents()
        time.sleep(1.0)
        for p in sorted(self.workdir.rglob("*"), reverse=True):
            try:
                if p.is_file():
                    unlink_retry(p)
                else:
                    p.rmdir()
            except OSError:
                pass
        try:
            self.workdir.rmdir()
        except OSError:
            pass

    def _prepare(self) -> None:
        wd = self.workdir
        proc = run_py(wd, "prepare_run.py", "--workdir", ".",
                      "--files", "source.xlsx|source.xlsx,template.xlsx|template.xlsx",
                      "--outline")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        proc = run_py(wd, "prepare_run.py", "--workdir", ".",
                      "--flatten", "--sheets", "source.xlsx:SRC;template.xlsx:TPL",
                      "--target", "template.xlsx")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])

    def _manifest(self) -> dict:
        return json.loads(
            (self.workdir / "prepare_manifest.json").read_text(encoding="utf-8"))

    def _meta(self) -> dict:
        return json.loads(
            (self.workdir / "template_TPL_meta.json").read_text(encoding="utf-8"))

    def _write_spec(self, target_fp: str) -> Path:
        spec = {
            "task": {"intent": "row-gap resync 契约", "selected_mod": "NONE",
                     "selected_mod_revision": None},
            "inputs": {"sources": ["source.xlsx"], "target": "template.xlsx",
                       "source_sheets": [{"source": "source.xlsx", "sheets": ["SRC"]}],
                       "target_sheet": "TPL"},
            "fingerprints": {"source_structure": "unused-source-fp",
                             "target_structure": target_fp},
            "mapping": {"targets": [{
                "sheet": "TPL", "base_last_row": 4,
                "clone_roles": [{"role": "data", "template_row": 3}],
                "rows": {"source": "source_SRC"},
                "columns": [{"source": "A", "target": "A"}],
            }]},
            "decisions": [], "gaps": [],
            "lineage": [{"source": "source_SRC_flat.csv", "role": "primary",
                         "note": "row-gap resync fixture"}],
            "validation": {"required_coverage": [], "required_empty": [],
                           "key_outputs": ["A5"]},
        }
        p = self.workdir / "fill_spec.yaml"
        p.write_text(yaml.safe_dump(spec, allow_unicode=True, sort_keys=False),
                     encoding="utf-8")
        return p

    def test_repair_resyncs_manifest_fingerprints(self):
        """repair 后 manifest 指纹自动同步: == repair 输出 == 手工重 flatten
        结果; 且与新指纹必不等于修复前 (指纹必然变化的机械事实)."""
        self._prepare()
        fp_before = self._manifest()["fingerprints"]["target_structure"]
        src_fp_before = self._manifest()["fingerprints"]["source_structure"]
        self.assertEqual(self._meta()["row_gaps"], [3])

        proc = run_py(self.workdir, "repair_row_gaps.py", "--workdir", ".")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        out = json.loads(proc.stdout)
        self.assertEqual(out["code"], "ROW_GAPS_REPAIRED")
        self.assertEqual(out["repaired"], [3])
        self.assertTrue(out["fingerprints_synced"])
        new_fp = out["fingerprints"]["target_structure"]

        m = self._manifest()
        self.assertEqual(m["fingerprints"]["target_structure"], new_fp)
        self.assertEqual(m["fingerprints"]["source_structure"], src_fp_before,
                         "源侧指纹不受目标修复影响")
        self.assertNotEqual(new_fp, fp_before,
                            "行洞修复 = staged 文件修改 = 指纹必然变化")
        self.assertEqual(self._meta()["row_gaps"], [])

        # 手工重 flatten (仅目标) == repair 自动同步结果 (确定性)
        proc = run_py(self.workdir, "prepare_run.py", "--workdir", ".",
                      "--flatten", "--sheets", "template.xlsx:TPL",
                      "--target", "template.xlsx")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        m2 = self._manifest()
        self.assertEqual(m2["fingerprints"]["target_structure"], new_fp)

    def test_repair_patch_spec_one_step(self):
        """--patch-spec: 一步改写 fill_spec.yaml 的 target_structure 指纹
        (旧 spec 会以 FILLSPEC_FINGERPRINT_MISMATCH 拒绝 → 补丁后匹配)."""
        self._prepare()
        fp_before = self._manifest()["fingerprints"]["target_structure"]
        spec_path = self._write_spec(fp_before)

        proc = run_py(self.workdir, "repair_row_gaps.py", "--workdir", ".",
                      "--patch-spec", str(spec_path))
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        docs = parse_json_docs(proc.stdout)
        self.assertEqual(docs[0]["code"], "ROW_GAPS_REPAIRED")
        self.assertEqual(docs[1]["code"], "SPEC_PATCHED")
        new_fp = docs[0]["fingerprints"]["target_structure"]
        self.assertEqual(Path(docs[1]["spec_patched"]), spec_path)

        spec = yaml.safe_load(spec_path.read_text(encoding="utf-8"))
        self.assertEqual(spec["fingerprints"]["target_structure"], new_fp)
        self.assertNotEqual(new_fp, fp_before)
        self.assertEqual(spec["task"]["intent"], "row-gap resync 契约",
                         "--patch-spec 只改指纹, 不碰其余内容")

    def test_repair_patch_spec_failure_keeps_fingerprints_visible(self):
        """--patch-spec 失败 (exit 3) 时, 新指纹已在 stdout 首份 JSON 里 —
        Agent 仍可抄进 spec; 重跑 repair 只会 NO_ROW_GAPS 不再带指纹."""
        self._prepare()
        bad = self.workdir / "bad_spec.yaml"
        bad.write_text("task: {intent: no-fingerprints}\n", encoding="utf-8")
        proc = run_py(self.workdir, "repair_row_gaps.py", "--workdir", ".",
                      "--patch-spec", str(bad))
        self.assertEqual(proc.returncode, 3)
        self.assertIn("SPEC_FINGERPRINT_NOT_FOUND", proc.stderr)
        docs = parse_json_docs(proc.stdout)
        self.assertEqual(docs[0]["code"], "ROW_GAPS_REPAIRED")
        self.assertEqual(docs[0]["fingerprints"]["target_structure"],
                         self._manifest()["fingerprints"]["target_structure"])

    def test_repair_no_gaps_is_idempotent_noop(self):
        """无空洞时 NO_ROW_GAPS + fingerprints_synced=False (no-op, 幂等)."""
        self._prepare()
        run_py(self.workdir, "repair_row_gaps.py", "--workdir", ".")
        m1 = self._manifest()
        proc = run_py(self.workdir, "repair_row_gaps.py", "--workdir", ".")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        out = json.loads(proc.stdout)
        self.assertEqual(out["code"], "NO_ROW_GAPS")
        self.assertFalse(out["fingerprints_synced"])
        self.assertEqual(self._manifest()["fingerprints"], m1["fingerprints"])


if __name__ == "__main__":
    unittest.main()
