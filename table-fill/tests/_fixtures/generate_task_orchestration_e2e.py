#!/usr/bin/env python3
"""Dev-time generator for the issue 08 e2e synthetic fixture (spec S8 /
ticket 08 fixture 策略).

⛔ 测试运行时绝不 import 本脚本：issue 08 的验收是「预生成合成工作簿提交到
   tests/_fixtures/task_orchestration/e2e/，不用运行时生成（避免 openpyxl
   依赖）」。本文件只用于 (重新) 生成并提交工作簿，生成后测试只做复制。

Fixture 设计（ticket 08）：
  - 1 个源工作簿（3 sheets × ~30 行）：R32参数 / R410A参数 / R22参数；
  - 4 个 run：r32-cooling 与 r32-heating 共享 sheet A（R32参数），另两个
    run 用 B/C（R410A参数 / R22参数）——唯一源 sheet 需求 U_source = 3；
  - 共享目标模板 filling_template.xlsx（Sheet1）—— 任务内唯一目标需求 +1；
  - 合成（无客户真实数据），结构满足 append clone 填充（base_last_row=4、
    template_row=3 空样式行），execute 走最小映射（A→A, B→B, C→C, D→D）。

用法（生成后 git 提交工作簿）：
  python tests/_fixtures/generate_task_orchestration_e2e.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

OUT_DIR = Path(__file__).resolve().parent / "task_orchestration" / "e2e"
SOURCE = OUT_DIR / "sources" / "parameter_book.xlsx"
TEMPLATE = OUT_DIR / "templates" / "filling_template.xlsx"

ROWS = 30  # ~30 行数据（ticket 08: 3 sheets × ~30 行）

CAPACITIES = ("9000Btu", "12000Btu", "18000Btu")

# sheet 名 → (产品线族名, 型号前缀 tag)
SHEETS = {
    "R32参数": ("R32 变频", "R32"),
    "R410A参数": ("R410A 变频", "R410A"),
    "R22参数": ("R22 定频", "R22"),
}

HEADER = ["产品线", "型号", "容量", "数量", "备注"]

THIN = Side(style="thin", color="FF9E9E9E")


def build_source() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, (family, tag) in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADER)
        for i in range(1, ROWS + 1):
            ws.append([
                family,
                f"{tag}-{i:04d}",
                CAPACITIES[i % len(CAPACITIES)],
                (i % 12) + 1,
                f"批次 {((i - 1) // 10) + 1} 备注文本",
            ])
        for col, width in zip("ABCDE", (16, 16, 12, 10, 24)):
            ws.column_dimensions[col].width = width
    SOURCE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SOURCE)


def build_template() -> None:
    """Sheet1: A1 标题 / A2 表头 / A3 数据模板行（空 + 换行样式，无残留值）/
    A4 合计 = base_last_row。append clone 从 A3 继承格式。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "空调客户参数表"
    for col, header in zip("ABCDE", HEADER):
        ws[f"{col}2"] = header
    # 数据模板行：空值（validate_clone_residue 要求无残留）+ 换行 + 边框样式
    for col in "ABCDE":
        c = ws[f"{col}3"]
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws["A4"] = "合计"
    ws["A4"].font = ws["A4"].font.copy(bold=True)
    ws["A4"].border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    for col, width in zip("ABCDE", (18, 22, 12, 10, 26)):
        ws.column_dimensions[col].width = width
    TEMPLATE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TEMPLATE)


def main() -> None:
    build_source()
    build_template()
    print(f"wrote {SOURCE}")
    print(f"wrote {TEMPLATE}")


if __name__ == "__main__":
    main()