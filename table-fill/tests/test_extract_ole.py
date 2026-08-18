from __future__ import annotations

import io
import json
import re
import subprocess
import sys
import tempfile
import unittest
import zipfile
from io import StringIO
from pathlib import Path
from unittest import mock

SKILL_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SKILL_ROOT / "scripts"))

import extract_ole  # noqa: E402

OLE_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject"
IMAGE_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
NOTES_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/notesSlide"


def rels_xml(*rels: tuple[str, str, str]) -> str:
    """Build a slide rels document. Each rel: (Id, Type, Target)."""
    body = "".join(
        f'<Relationship Id="{rid}" Type="{typ}" Target="{target}"/>'
        for rid, typ, target in rels
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{body}</Relationships>"
    )


def make_xlsx_bytes() -> bytes:
    """Minimal valid xlsx with real cell data (fixture only, no external file)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["B2"] = 42
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def probe_payload(prog_id: str, rel_id: str) -> dict:
    """Shape of officecli `get /slide[N]/ole[K] --json` for a real OLE object."""
    return {
        "success": True,
        "data": {
            "matches": 1,
            "results": [
                {
                    "path": "/slide[1]/ole[1]",
                    "type": "ole",
                    "text": prog_id,
                    "childCount": 0,
                    "format": {
                        "objectType": "ole",
                        "progId": prog_id,
                        "name": "对象",
                        "display": "content",
                        "relId": rel_id,
                        "contentType": "application/vnd.openxmlformats-officedocument.oleObject",
                        "fileSize": 1234,
                    },
                    "children": [],
                }
            ],
        },
    }


class FakeOfficecli:
    """Adapter double: answers each `/ole[N]` probe from payload_by_index.

    Any unlisted index returns success:false (not_found), like real officecli.
    """

    def __init__(self, payload_by_index: dict[int, dict]):
        self.payload_by_index = payload_by_index
        self.probed: list[tuple] = []

    def __call__(self, *args, **kwargs):
        self.probed.append(args)
        m = re.search(r"/ole\[(\d+)\]", " ".join(str(a) for a in args))
        idx = int(m.group(1)) if m else 0
        payload = self.payload_by_index.get(
            idx, {"success": False, "error": {"code": "not_found"}}
        )
        return subprocess.CompletedProcess(
            args, 0, stdout=json.dumps(payload), stderr=""
        )


class OleFixtureMixin:
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.tmp = Path(self._tmp.name)

    def make_pptx(self, rels: str, embeddings: dict[str, bytes] | None = None,
                  slide: int = 1) -> Path:
        path = self.tmp / "fixture.pptx"
        with zipfile.ZipFile(path, "w") as z:
            z.writestr(f"ppt/slides/_rels/slide{slide}.xml.rels", rels)
            for name, data in (embeddings or {}).items():
                z.writestr(f"ppt/embeddings/{name}", data)
        return path


class FindOleMappingsTests(OleFixtureMixin, unittest.TestCase):
    """rels → OLEObject-typed embeddings only, in rels document order."""

    def test_filters_by_relationship_type(self):
        rels = rels_xml(
            ("rId1", IMAGE_REL, "../media/image1.emf"),
            ("rId2", OLE_REL, "../embeddings/oleObject2.bin"),
            ("rId3", NOTES_REL, "../notesSlides/notesSlide1.xml"),
            ("rId4", OLE_REL, "../embeddings/oleObject1.bin"),
        )
        pptx = self.make_pptx(rels)
        self.assertEqual(
            extract_ole.find_ole_mappings(str(pptx), 1),
            [
                {"rel_id": "rId2", "embedding": "embeddings/oleObject2.bin", "number": 2},
                {"rel_id": "rId4", "embedding": "embeddings/oleObject1.bin", "number": 1},
            ],
        )

    def test_missing_rels_returns_empty(self):
        pptx = self.tmp / "bare.pptx"
        with zipfile.ZipFile(pptx, "w") as z:
            z.writestr("ppt/presentation.xml", "<p/>")
        self.assertEqual(extract_ole.find_ole_mappings(str(pptx), 1), [])

    def test_no_oleobject_rels_returns_empty(self):
        rels = rels_xml(
            ("rId1", IMAGE_REL, "../media/image1.emf"),
            ("rId2", NOTES_REL, "../notesSlides/notesSlide1.xml"),
        )
        pptx = self.make_pptx(rels)
        self.assertEqual(extract_ole.find_ole_mappings(str(pptx), 1), [])


class SelectExcelOleTests(OleFixtureMixin, unittest.TestCase):
    """ProgId-validated selection: Excel.Sheet wins, never 'first by filename'."""

    def _pptx_with_two_ole(self):
        rels = rels_xml(
            ("rId1", OLE_REL, "../embeddings/oleObject1.bin"),  # Word, first in rels
            ("rId2", OLE_REL, "../embeddings/oleObject2.bin"),  # Excel, second
        )
        return self.make_pptx(rels)

    def test_selects_excel_not_first_ole(self):
        pptx = self._pptx_with_two_ole()
        fake = FakeOfficecli({
            1: probe_payload("Word.Document.12", "rId1"),
            2: probe_payload("Excel.Sheet.12", "rId2"),
        })
        with mock.patch("extract_ole.officecli", fake):
            chosen = extract_ole.select_excel_ole(str(pptx), 1)
        self.assertEqual(chosen["number"], 2)
        self.assertEqual(chosen["rel_id"], "rId2")
        self.assertEqual([a[2] for a in fake.probed], ["/slide[1]/ole[1]", "/slide[1]/ole[2]"])

    def test_matches_probe_slot_to_rels_by_relid(self):
        """officecli 的 ole 序号是位置序号；probe 返回的 relId 用于回映射 embedding。"""
        rels = rels_xml(
            ("rId1", OLE_REL, "../embeddings/oleObject5.bin"),
            ("rId2", OLE_REL, "../embeddings/oleObject3.bin"),
        )
        pptx = self.make_pptx(rels)
        fake = FakeOfficecli({
            1: probe_payload("Word.Document.12", "rId1"),
            2: probe_payload("Excel.Sheet.12", "rId2"),
        })
        with mock.patch("extract_ole.officecli", fake):
            chosen = extract_ole.select_excel_ole(str(pptx), 1)
        self.assertEqual(chosen["number"], 3)

    def test_no_excel_returns_none(self):
        pptx = self._pptx_with_two_ole()
        fake = FakeOfficecli({
            1: probe_payload("Word.Document.12", "rId1"),
            2: probe_payload("Word.Document.12", "rId2"),
        })
        with mock.patch("extract_ole.officecli", fake):
            self.assertIsNone(extract_ole.select_excel_ole(str(pptx), 1))

    def test_no_ole_returns_none(self):
        rels = rels_xml(("rId1", IMAGE_REL, "../media/image1.emf"))
        pptx = self.make_pptx(rels)
        with mock.patch("extract_ole.officecli", FakeOfficecli({})):
            self.assertIsNone(extract_ole.select_excel_ole(str(pptx), 1))


class MainFailuresTests(OleFixtureMixin, unittest.TestCase):
    """Only non-Excel OLE or no OLE → structured defect, no output file."""

    def run_main(self, pptx: Path, fake: FakeOfficecli) -> tuple[int, str]:
        out_dir = self.tmp / "out"
        buf = StringIO()
        old_stderr = sys.stderr
        sys.stderr = buf
        try:
            with mock.patch("extract_ole.officecli", fake), \
                 mock.patch.object(sys, "argv", ["extract_ole.py", "--input",
                                                 str(pptx), "--slide", "1",
                                                 "--output-dir", str(out_dir)]):
                try:
                    extract_ole.main()
                    return 0, buf.getvalue()
                except SystemExit as e:
                    return int(e.code or 0), buf.getvalue()
        finally:
            sys.stderr = old_stderr
        return 1, buf.getvalue()

    def test_non_excel_ole_fails_with_defect_code(self):
        rels = rels_xml(("rId1", OLE_REL, "../embeddings/oleObject1.bin"))
        pptx = self.make_pptx(rels, embeddings={"oleObject1.bin": b"not a zip"})
        fake = FakeOfficecli({1: probe_payload("Word.Document.12", "rId1")})
        code, stderr = self.run_main(pptx, fake)
        self.assertEqual(code, 3)
        self.assertIn('"code": "OLE_NO_EXCEL_EMBEDDING"', stderr)
        self.assertIn('"corrective_action"', stderr)
        self.assertEqual(list((self.tmp / "out").glob("*")), [])  # nothing produced

    def test_no_ole_fails_with_defect_code(self):
        rels = rels_xml(("rId1", IMAGE_REL, "../media/image1.emf"))
        pptx = self.make_pptx(rels)
        code, stderr = self.run_main(pptx, FakeOfficecli({}))
        self.assertEqual(code, 3)
        self.assertIn('"code": "OLE_NO_EXCEL_EMBEDDING"', stderr)
        self.assertIn('"corrective_action"', stderr)

    def test_missing_input_fails_with_defect_code(self):
        missing = self.tmp / "nope.pptx"
        code, stderr = self.run_main(missing, FakeOfficecli({}))
        self.assertEqual(code, 3)
        self.assertIn('"code": "INPUT_NOT_FOUND"', stderr)

    def test_extract_failure_emits_defect_code(self):
        rels = rels_xml(("rId1", OLE_REL, "../embeddings/oleObject1.bin"))
        pptx = self.make_pptx(rels, embeddings={"oleObject1.bin": b"no zip markers here"})
        fake = FakeOfficecli({1: probe_payload("Excel.Sheet.12", "rId1")})
        code, stderr = self.run_main(pptx, fake)
        self.assertEqual(code, 3)
        self.assertIn('"code": "OLE_EXTRACT_FAILED"', stderr)
        self.assertIn('"corrective_action"', stderr)
        self.assertEqual(list((self.tmp / "out").glob("*")), [])


class ExtractEndToEndTests(OleFixtureMixin, unittest.TestCase):
    """Multi-OLE fixture: the Excel.Sheet embedding is extracted, not the first."""

    def test_extracts_excel_embedding(self):
        xlsx = make_xlsx_bytes()
        rels = rels_xml(
            ("rId1", OLE_REL, "../embeddings/oleObject1.bin"),  # Word, first
            ("rId2", OLE_REL, "../embeddings/oleObject2.bin"),  # Excel
        )
        pptx = self.make_pptx(
            rels,
            embeddings={
                "oleObject1.bin": b"\x00" * 64 + b"word payload, not a zip",
                "oleObject2.bin": b"\x00" * 128 + xlsx,  # OLE 容器前缀 + 内嵌 xlsx
            },
        )
        fake = FakeOfficecli({
            1: probe_payload("Word.Document.12", "rId1"),
            2: probe_payload("Excel.Sheet.12", "rId2"),
        })
        out_dir = self.tmp / "out"
        buf = StringIO()
        old_stderr = sys.stderr
        sys.stderr = buf
        try:
            with mock.patch("extract_ole.officecli", fake), \
                 mock.patch.object(sys, "argv", ["extract_ole.py", "--input",
                                                 str(pptx), "--slide", "1",
                                                 "--output-dir", str(out_dir)]):
                try:
                    code = extract_ole.main()
                except SystemExit as e:
                    code = int(e.code or 0)
        finally:
            sys.stderr = old_stderr
        self.assertEqual(code, 0)
        produced = sorted(p.name for p in out_dir.glob("*"))
        self.assertEqual(produced, ["oleObject2_slide_extracted.xlsx"])
        extracted = (out_dir / "oleObject2_slide_extracted.xlsx").read_bytes()
        self.assertEqual(extracted, xlsx)

    def test_extract_xlsx_from_ole_regression(self):
        xlsx = make_xlsx_bytes()
        rels = rels_xml(("rId1", OLE_REL, "../embeddings/oleObject1.bin"))
        pptx = self.make_pptx(rels, embeddings={"oleObject1.bin": b"\x00" * 100 + xlsx})
        out_dir = self.tmp / "out"
        out_dir.mkdir(parents=True, exist_ok=True)  # main() 负责建目录，这里模拟真实用法
        out_path = extract_ole.extract_xlsx_from_ole(str(pptx), 1, out_dir)
        self.assertIsNotNone(out_path)
        self.assertEqual(Path(out_path).read_bytes(), xlsx)


class DocumentedFlattenEntriesTests(unittest.TestCase):
    """AC4: LAYER1_OLE_HANDLING.md 的命令引用真实脚本与真实参数（可执行性守卫）。"""

    def test_flatten_table_entry_exists_with_documented_flags(self):
        src = (SKILL_ROOT / "scripts" / "flatten_table.py").read_text(encoding="utf-8")
        for flag in ("--input", "--target", "--output", "--meta"):
            self.assertIn(flag, src)

    def test_flatten_workbook_entry_exists_with_documented_flags(self):
        src = (SKILL_ROOT / "scripts" / "flatten_workbook.py").read_text(encoding="utf-8")
        for flag in ("--input", "--plan", "--out-dir"):
            self.assertIn(flag, src)

    def test_layer1_doc_has_no_stale_script_and_documents_defect(self):
        doc = (SKILL_ROOT / "references" / "LAYER1_OLE_HANDLING.md").read_text(encoding="utf-8")
        self.assertNotIn("flatten_source", doc)
        self.assertIn("OLE_NO_EXCEL_EMBEDDING", doc)


if __name__ == "__main__":
    unittest.main()
