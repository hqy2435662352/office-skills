"""Compile/execute backing for Q7 (precision: keep vs round4).

The 2026-08-12 retrospective documented a cross-layer trap: a 15-digit cost
value written with `precision: keep` into a normal-width column passed
compile and failed at EXECUTE with text_overflow. Issue 04 converts that
trap into a COMPILE-TIME check: prepare now measures the template column
width (meta.column_width), and the compiler rejects `precision: keep` when
the widest rendered value exceeds the measured column width
(PRECISION_KEEP_NARROW_COLUMN, exit 3). These tests back the contract at
the layer that used to break:

- keep + narrow column  → compile rejected (PRECISION_KEEP_NARROW_COLUMN)
- keep + wide column    → compiles and executes green
- round4 + narrow column → the documented pattern, compiles and executes green

Skips cleanly when officecli is unavailable (unit-only environments).
Run with:
  python -m unittest tests.test_precision_e2e -v
"""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import yaml

SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = SKILL_ROOT / "scripts"

LONG_DECIMAL = "168.715100569657"


def run_py(workdir: Path, script: str, *args) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, "-X", "utf8", str(SCRIPTS / script), *args],
        cwd=str(workdir), capture_output=True, text=True, encoding="utf-8",
        errors="replace", timeout=900)


def build_fixtures(dirpath: Path, col_width: float = 12.0) -> None:
    """Tiny template + source: 4-row template (col A default width, data row
    wrapText + fixed height like the real quote template), 3 source rows
    carrying a 15-digit cost value."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    tpl = Workbook()
    ws = tpl.active
    ws.title = "S"
    ws["A1"] = "Quotation"       # title row
    ws["A2"] = "Type | Model"    # header row
    ws["A3"] = ""                # data template row (empty — no residue)
    ws["A4"] = "Total"           # base_last_row
    ws.row_dimensions[3].height = 20   # 固定行高 (真实模板 customHeight)
    ws["A3"].alignment = Alignment(wrap_text=True)  # 换行 → 长数值触发 text overflow
    ws.column_dimensions["A"].width = col_width  # 列宽实测 (prepare 采集到 meta.column_width)
    tpl.save(dirpath / "template.xlsx")

    src = Workbook()
    ws = src.active
    ws.title = "SRC"
    for i in range(3):
        ws.cell(row=i + 1, column=1, value=float(LONG_DECIMAL))
    src.save(dirpath / "source.xlsx")


@unittest.skipIf(shutil.which("officecli") is None, "officecli not on PATH")
class PrecisionExecutionContractTests(unittest.TestCase):
    """Q7 契约背书: keep 窄列 → 编译期拒绝 (不再执行期 text_overflow);
    keep 宽列 / round4 → 全链路通过."""

    def setUp(self):
        sys.path.insert(0, str(SCRIPTS))
        from _officecli import clean_residents  # noqa: PLC0415
        self.workdir = Path(tempfile.mkdtemp(prefix="prec_e2e_"))
        build_fixtures(self.workdir)

    def tearDown(self):
        from _officecli import clean_residents, unlink_retry  # noqa: PLC0415
        import time
        clean_residents()
        time.sleep(1.0)
        for p in sorted(self.workdir.rglob("*"), reverse=True):
            try:
                if p.is_file():
                    unlink_retry(p)
                else:
                    p.rmdir()
            except OSError:
                pass
        try:
            self.workdir.rmdir()
        except OSError:
            pass

    def _prepare(self) -> None:
        wd = self.workdir
        proc = run_py(wd, "prepare_run.py", "--workdir", ".",
                      "--files", "source.xlsx|source.xlsx,template.xlsx|template.xlsx",
                      "--outline")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])
        proc = run_py(wd, "prepare_run.py", "--workdir", ".",
                      "--flatten", "--sheets", "source.xlsx:SRC;template.xlsx:S",
                      "--target", "template.xlsx")
        self.assertEqual(proc.returncode, 0, proc.stderr[-800:])

    def _write_spec(self, column_extra: dict | None) -> Path:
        wd = self.workdir
        manifest = json.loads((wd / "prepare_manifest.json").read_text(encoding="utf-8"))
        fp = manifest["fingerprints"]
        mapping = {"source": "A", "target": "A"}
        if column_extra:
            mapping.update(column_extra)
        spec = {
            "task": {"intent": "precision 契约", "selected_mod": "NONE",
                     "selected_mod_revision": None},
            "inputs": {"sources": ["source.xlsx"], "target": "template.xlsx",
                       "source_sheets": [{"source": "source.xlsx", "sheets": ["SRC"]}],
                       "target_sheet": "S"},
            "fingerprints": {"source_structure": fp["source_structure"],
                             "target_structure": fp["target_structure"]},
            "mapping": {"targets": [{
                "sheet": "S", "base_last_row": 4,
                "clone_roles": [{"role": "data", "template_row": 3}],
                "rows": {"source": "source_SRC"},
                "columns": [mapping],
            }]},
            "decisions": [], "gaps": [],
            "lineage": [{"source": "source_SRC_flat.csv", "role": "primary",
                         "note": "精度契约 fixture"}],
            "validation": {"required_coverage": [], "required_empty": [],
                           "key_outputs": ["A5"]},
        }
        (wd / "fill_spec.yaml").write_text(
            yaml.safe_dump(spec, allow_unicode=True, sort_keys=False),
            encoding="utf-8")
        return wd / "fill_spec.yaml"

    def _compile(self, column_extra: dict | None) -> subprocess.CompletedProcess:
        self._prepare()
        spec_path = self._write_spec(column_extra)
        return run_py(self.workdir, "compile_fill.py", "--spec", spec_path.name,
                      "--workdir", ".")

    def _execute(self) -> subprocess.CompletedProcess:
        return run_py(self.workdir, "execute_batch.py", "--plan",
                      "execution_plan.json", "--template", "template.xlsx",
                      "--workdir", ".", "--round", "1", "--render", "html")

    def test_precision_keep_narrow_column_rejected_at_compile(self):
        """跨层陷阱已前移: 15 位值 + 12 宽列 + precision: keep → 编译期
        PRECISION_KEEP_NARROW_COLUMN (exit 3), 不再白烧一轮 execute."""
        proc = self._compile({"precision": "keep"})
        self.assertEqual(proc.returncode, 3, proc.stdout[-800:] + proc.stderr[-800:])
        self.assertIn("PRECISION_KEEP_NARROW_COLUMN", proc.stderr)
        self.assertFalse((self.workdir / "execution_plan.json").exists())

    def test_precision_keep_wide_column_passes_e2e(self):
        """keep 的机械前提成立: 列宽 20 > 16 渲染字符 → 编译 + 执行全链路通过."""
        (self.workdir / "template.xlsx").unlink(missing_ok=True)
        (self.workdir / "source.xlsx").unlink(missing_ok=True)
        build_fixtures(self.workdir, col_width=20.0)
        proc = self._compile({"precision": "keep"})
        self.assertEqual(proc.returncode, 0, proc.stdout[-800:] + proc.stderr[-800:])
        proc = self._execute()
        self.assertEqual(proc.returncode, 0, proc.stdout[-800:] + proc.stderr[-800:])
        receipt = json.loads(
            (self.workdir / "draft_receipt.json").read_text(encoding="utf-8"))
        self.assertEqual(receipt["issue_delta"]["new_issues"], 0)
        self.assertEqual(receipt["readback"]["passed"], receipt["readback"]["total"])

    def test_precision_round4_passes_at_execute(self):
        """文档推荐模式: transform: round4 → 编译 + 执行全链路通过, 无新增 issue."""
        proc = self._compile({"transform": "round4"})
        self.assertEqual(proc.returncode, 0, proc.stdout[-800:] + proc.stderr[-800:])
        proc = self._execute()
        self.assertEqual(proc.returncode, 0, proc.stdout[-800:] + proc.stderr[-800:])
        receipt = json.loads(
            (self.workdir / "draft_receipt.json").read_text(encoding="utf-8"))
        self.assertEqual(receipt["issue_delta"]["new_issues"], 0)
        self.assertEqual(receipt["readback"]["passed"], receipt["readback"]["total"])


if __name__ == "__main__":
    unittest.main()
